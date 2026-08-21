"""
Compliance Reports API — HTTP handler for the frontend.

All routes require Cognito JWT in Authorization header (validated by API Gateway).

Routes:
  GET  /reports             → list built-in + custom reports
  POST /execute             → execute a report async, returns run_id
  GET  /runs                → list recent runs (last 50)
  GET  /runs/{run_id}       → get run status + presigned download URL + result preview
  POST /queries             → save a custom SQL query to catalog
  DELETE /queries/{name}    → delete a custom query
  GET  /cluster/status      → get Redshift cluster status
  POST /cluster/wake        → resume a paused cluster
  POST /cluster/pause       → pause the cluster
  GET  /whitelist           → list active whitelist entries
  POST /whitelist           → add an entry to the whitelist
  DELETE /whitelist/{id}    → remove a whitelist entry
  GET  /alerts              → list active alert entries
  GET  /alerts/reviewed     → list reviewed (ya revisados) entries
  POST /alerts              → add an alert entry
  PUT  /alerts/{id}/review  → mark alert as reviewed (move to ya revisados)
  DELETE /alerts/{id}       → permanently remove an alert
  GET  /dashboard/stats        → submit 3 queries to Redshift; returns stmt_ids immediately
  GET  /dashboard/stats/result → poll results for stmt_ids (q0=, q1=, q2=); returns per-query
                                 rows when done, null when still running, all_done flag
  GET  /analytics/summary      → submit 5 CRM analytics queries; returns stmt_ids immediately
  GET  /analytics/result       → poll analytics results (q0..q4); all_done flag
  GET  /analytics/sla          → submit 3 SLA queries; returns stmt_ids
  GET  /analytics/sla/result   → poll SLA results (q0..q2); all_done flag
  GET  /users                  → list all CRM users with role name
  POST /users                  → create user {email, full_name, role_id}
  PUT  /users/{id}             → update user {full_name, role_id, is_active}
  DELETE /users/{id}           → deactivate user (soft delete)
  GET  /roles                  → list all roles
  GET  /rules                  → list auto-case rules (S3 JSON)
  POST /rules                  → create auto-case rule
  PUT  /rules/{id}             → update auto-case rule
  DELETE /rules/{id}           → delete auto-case rule
"""

from __future__ import annotations

import datetime as dt
import decimal
import hashlib
import html
import json
import os
import re
import time
import uuid
from pathlib import Path

import boto3
from boto3.dynamodb.conditions import Attr

try:
    from db_redshift import write_audit as _write_audit
except Exception:
    def _write_audit(**kwargs):  # noqa: ANN001
        pass

dynamodb = boto3.resource("dynamodb")
lambda_client = boto3.client("lambda")
s3 = boto3.client("s3")
redshift = boto3.client("redshift")
redshift_data = boto3.client("redshift-data")
secrets_client = boto3.client("secretsmanager")

SLACK_SECRET_ARN = os.environ.get("SLACK_WEBHOOK_SECRET_ARN", "")

def _get_slack_url() -> str:
    if not SLACK_SECRET_ARN:
        return ""
    try:
        val = secrets_client.get_secret_value(SecretId=SLACK_SECRET_ARN)
        raw = val.get("SecretString", "")
        try:
            return json.loads(raw).get("webhook_url", raw)
        except Exception:
            return raw.strip()
    except Exception:
        return ""

# Webhook dedicado al canal de alertas operativas (#watchtower_alertas). Un
# incoming webhook publica siempre en el canal con el que fue creado — no se
# puede elegir el canal por mensaje —, así que para separar canales hace falta
# un webhook por canal. Si este secreto no existe, se cae al webhook general
# para no perder avisos.
SLACK_ALERTS_SECRET_NAME = os.environ.get(
    "SLACK_ALERTS_SECRET_NAME", "compliance-redshift-reports/slack-webhook-alertas"
)
_slack_alerts_url_cache: str | None = None


def _get_slack_alerts_url() -> str:
    """Webhook del canal de alertas, con fallback al general."""
    global _slack_alerts_url_cache
    if _slack_alerts_url_cache is not None:
        return _slack_alerts_url_cache
    url = ""
    try:
        raw = secrets_client.get_secret_value(
            SecretId=SLACK_ALERTS_SECRET_NAME).get("SecretString", "")
        try:
            url = json.loads(raw).get("webhook_url", raw)
        except Exception:
            url = (raw or "").strip()
    except Exception:
        url = ""
    if not url:
        url = _get_slack_url()  # fallback: canal general
    _slack_alerts_url_cache = url
    return url


# Mapeo correo corporativo -> user ID de Slack, para poder arrobar de verdad.
# Escribir el correo en el mensaje NO genera notificación: Slack solo avisa
# con la sintaxis <@Uxxxx>. El mapeo vive en S3 para poder actualizarlo cuando
# entra gente nueva sin volver a desplegar.
SLACK_USERS_KEY = "config/slack_users.json"
_slack_users_cache: dict | None = None


def _slack_users() -> dict:
    global _slack_users_cache
    if _slack_users_cache is not None:
        return _slack_users_cache
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=SLACK_USERS_KEY)
        data = json.loads(obj["Body"].read())
        _slack_users_cache = {k.strip().lower(): v for k, v in (data or {}).items() if v}
    except Exception:
        _slack_users_cache = {}
    return _slack_users_cache


def _slack_mention(email: str, fallback: str = "_sin asignar_") -> str:
    """Devuelve <@Uxxxx> si conocemos el ID de Slack de ese correo; si no, el
    correo tal cual (mejor eso que nada, aunque no notifique)."""
    email = str(email or "").strip()
    if not email:
        return fallback
    uid = _slack_users().get(email.lower())
    return f"<@{uid}>" if uid else email


def _post_slack(text: str, blocks: list | None = None, canal: str = "alertas") -> None:
    """Publica en Slack. `blocks` permite mandar Block Kit (formato rico); el
    `text` queda igual como fallback para notificaciones push y clientes que
    no renderizan bloques.

    Por defecto va al canal de alertas operativas (#watchtower_alertas): es el
    canal único de WatchTower. Se deja como DEFAULT en vez de marcarlo en cada
    llamada para que cualquier aviso que se agregue más adelante caiga ahí
    solo, sin depender de acordarse. Con canal='general' se puede mandar al
    webhook viejo si alguna vez hiciera falta separar algo."""
    url = _get_slack_url() if canal == "general" else _get_slack_alerts_url()
    if not url:
        return
    import urllib.request as _ur
    cuerpo: dict = {"text": text}
    if blocks:
        cuerpo["blocks"] = blocks
    payload = json.dumps(cuerpo).encode()
    try:
        _ur.urlopen(_ur.Request(url, data=payload,
                                headers={"Content-Type": "application/json"},
                                method="POST"), timeout=5)
    except Exception:
        pass

WATCHTOWER_URL = os.environ.get(
    "WATCHTOWER_URL", "https://bmackenna-g66.github.io/compliance-redshift-reports"
)
_PRIORITY_LABEL_ES = {"high": "🔴 Alta", "medium": "🟡 Media", "low": "🟢 Baja"}
_CASE_STATUS_LABEL_ES = {
    "open": "Abierto", "in_progress": "En Investigación",
    "under_review": "Bajo Revisión", "closed": "Cerrado", "archived": "Archivado",
}
# Campos que ya se muestran aparte en el mensaje o que no aportan como métrica.
_SLACK_SKIP_FIELDS = {
    "customer_id", "company_id", "client_id", "beneficiary_id", "email",
    "nombre", "apellido", "nombre_completo", "correo", "name", "last_name",
    "prioridad", "concepto", "risk_score",
}


def _slack_alert_metrics(alert_data: dict, limite: int = 4) -> str:
    """Resumen corto de los números de la alerta, para que el mensaje de Slack
    diga POR QUÉ se abrió el caso y no solo que se abrió."""
    if not isinstance(alert_data, dict):
        return ""
    partes = []
    for k, v in alert_data.items():
        if k.lower() in _SLACK_SKIP_FIELDS:
            continue
        if v is None or v == "" or v == "—":
            continue
        val = f"{v:,.2f}".rstrip("0").rstrip(".") if isinstance(v, float) else (
            f"{v:,}" if isinstance(v, int) else str(v)
        )
        partes.append(f"`{k.replace('_', ' ')}`: {val}")
        if len(partes) >= limite:
            break
    return " · ".join(partes)


def _slack_client_line(case: dict) -> str:
    """Línea de identificación del cliente/empresa del caso."""
    nombre = str(case.get("entity_name") or "").strip()
    eid = str(case.get("entity_id") or "").strip()
    if not nombre and not eid:
        return ""
    etype = str(case.get("entity_type") or "").lower()
    campo = "company_id" if etype.startswith("comp") else "customer_id"
    if nombre and eid:
        return f"👤 Cliente: *{nombre}* (`{campo} {eid}`)"
    if nombre:
        return f"👤 Cliente: *{nombre}*"
    return f"👤 Cliente: `{campo} {eid}`"


def _slack_case_link(case_id: str) -> str:
    return f"🔗 <{WATCHTOWER_URL}|Abrir en WatchTower> · caso `{case_id[:8]}…`"


CLUSTER_ID = os.environ.get("CLUSTER_IDENTIFIER", "compliance-redshift-cluster")
# Decisión de negocio (2026-08-04): el cluster se deja encendido permanentemente
# en AWS — ningún proceso, ni siquiera el botón manual de Admin, debe pausarlo.
# Mismo criterio que AUTO_PAUSE en handler.py (report runner). Volver a poner en
# True solo si esa decisión cambia.
CLUSTER_MANUAL_PAUSE_ENABLED = False
DATABASE_NAME = os.environ.get("DATABASE_NAME", "dev")
DB_USER = os.environ.get("DB_USER", "awsuser")

RUNS_TABLE_NAME = os.environ["RUNS_TABLE"]
CATALOG_TABLE_NAME = os.environ["CATALOG_TABLE"]
REPORT_LAMBDA_NAME = os.environ["REPORT_LAMBDA"]
S3_BUCKET = os.environ["S3_BUCKET"]

# Phase 8 — Email notifications
GMAIL_USER = os.environ.get("GMAIL_USER", "benjamin.mackenna@global66.com")
# Fallback heredado: la app password como variable de entorno en texto plano.
# La fuente preferida es Secrets Manager (ver _get_gmail_password).
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
GMAIL_PASSWORD_SECRET_NAME = os.environ.get(
    "GMAIL_PASSWORD_SECRET_NAME", "compliance-redshift-reports/gmail-app-password"
)
_gmail_password_cache: str | None = None


def _get_gmail_password() -> str:
    """App password de Gmail, priorizando Secrets Manager sobre la variable de
    entorno en texto plano.

    Google muestra la clave en 4 grupos de 4 ("abcd efgh ijkl mnop") y es muy
    fácil guardarla tal cual, pero SMTP la rechaza con los espacios — así que
    se limpian acá en vez de depender de que se haya guardado bien.

    Se cachea por contenedor: si se rota el secreto, el valor nuevo entra
    cuando Lambda recicle el contenedor (o con un redeploy).
    """
    global _gmail_password_cache
    if _gmail_password_cache is not None:
        return _gmail_password_cache

    pw = ""
    try:
        sm = boto3.client("secretsmanager")
        raw = sm.get_secret_value(SecretId=GMAIL_PASSWORD_SECRET_NAME).get("SecretString", "")
        pw = "".join((raw or "").split())
        if pw:
            print(f"[email] app password leída de Secrets Manager ({len(pw)} chars)")
    except Exception as e:
        print(f"[email] no se pudo leer el secreto ({type(e).__name__}) — se usa la variable de entorno")

    if not pw:
        pw = "".join((GMAIL_APP_PASSWORD or "").split())

    _gmail_password_cache = pw
    return pw

# Phase 10 — Auto-case rules S3 key
AUTO_RULES_KEY = "config/auto_case_rules.json"

# Priorización de alertas — mantenedor de documentos a solicitar por alerta
ALERT_DOCS_CONFIG_KEY = "config/alert_document_config.json"
PRIORITY_TEST_TABLE = "compliance.alert_priority_test_data"
_PRIORITY_TO_CASE_PRIORITY = {"P1": "high", "P2": "medium", "P3": "low"}
_ALL_DOC_CATEGORIES = [
    "Origen de fondo", "Comprobantes/Soporte", "Relación/Beneficiario",
    "Domicilio", "Identidad/Datos personales",
]
# Interruptor general del envío automático de solicitudes de documentos.
# Apagado por defecto — se prende explícitamente desde el Admin cuando el
# proceso esté listo para correr sobre alertas reales (hoy solo se usa en el
# botón de prueba, que respeta este mismo interruptor).
PRIORITY_QUEUE_SETTINGS_KEY = "config/priority_queue_settings.json"
# Remitente "enviar como" — alias configurado en la cuenta de GMAIL_USER, no
# necesita una app password propia (ver _send_email).
ALERT_DOCS_FROM_ADDR = "compliance@global66.com"

# Escucha de respuestas del cliente (documentos adjuntos por correo).
# compliance@global66.com es un GRUPO de Workspace, no una casilla con login
# propio — no se puede hacer IMAP directo sobre el grupo. compliance.masivo@
# global66.com es una cuenta Gmail real que es miembro del grupo, así que una
# copia de cada respuesta le llega ahí también; ese es el buzón que se lee.
DOC_REPLY_IMAP_HOST = "imap.gmail.com"
DOC_REPLY_IMAP_USER = os.environ.get("DOC_REPLY_IMAP_USER", "compliance.masivo@global66.com")
DOC_REPLY_IMAP_SECRET_ARN = os.environ.get("DOC_REPLY_IMAP_SECRET_ARN", "")
_ATTACHMENT_EXTS_ALLOWED = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx", ".xls", ".xlsx", ".heic", ".webp"}
# Ventana hacia atrás del poller de respuestas. Acota la búsqueda a correo
# reciente: la casilla arrastra miles de mensajes viejos del grupo que no son
# respuestas a solicitudes, y no tiene sentido tocarlos.
DOC_REPLY_LOOKBACK_DAYS = int(os.environ.get("DOC_REPLY_LOOKBACK_DAYS", "10"))
_EMAIL_REF_RE = re.compile(r"\[ref:\s*([0-9a-f]{8})\]", re.IGNORECASE)

# ---------------------------------------------------------------------------
# Plantillas de correo — solicitud de documentación / comunicación con cliente
# ---------------------------------------------------------------------------
# Reemplazan al viejo esquema de fragmentos por categoría (solicitud_documentos_
# fragments.json, nunca llegó a poblarse). Ahora son 4 plantillas HTML
# completas, ya diseñadas y aprobadas por Compliance, cada una con su propio
# formulario adjunto (salvo texto_libre, que es para comunicación ad-hoc).
EMAIL_TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"
EMAIL_ATTACHMENTS_DIR = Path(__file__).resolve().parent / "attachments"
_SF_MERGE_FIELD_NOMBRE = "{!Account.first_name__c}"
_TEXTO_LIBRE_PLACEHOLDER = "TEXTO LIBRE<br>TEXTO LIBRE<br>TEXTO LIBRE"

EMAIL_TEMPLATE_CATALOG = {
    "texto_libre": {
        "label": "Texto libre",
        "file": "email_texto_libre.html",
        "attachment": None,
        "requires_custom_text": True,
    },
    "general_b2c": {
        "label": "Solicitud de documentos — B2C general",
        "file": "email_general_b2c.html",
        "attachment": "Formulario_KYC_Individual_Global.pdf",
        "requires_custom_text": False,
    },
    "argentina_b2c": {
        "label": "Solicitud de documentos — B2C Argentina",
        "file": "email_argentina_b2c.html",
        "attachment": "Formulario_KYC_Individual_Global.pdf",
        "requires_custom_text": False,
    },
    "b2b_generico": {
        "label": "Solicitud de documentos — B2B genérico",
        "file": "email_b2b_generico.html",
        "attachment": "Formulario_B2B.pdf",
        "requires_custom_text": False,
    },
    # Mismo diseño que el B2C general (cabecera, tipografía, saludo, cierre y
    # footer) pero con el cuerpo vacío: el analista escribe el contenido. Sirve
    # para pedidos que no encajan en el texto estándar sin perder la identidad
    # del correo. Lleva el formulario adjunto porque el uso previsto sigue
    # siendo pedir documentación.
    "general_b2c_blanco": {
        "label": "B2C general — cuerpo en blanco (texto libre)",
        "file": "email_general_b2c_blanco.html",
        "attachment": "Formulario_KYC_Individual_Global.pdf",
        "requires_custom_text": True,
    },
}

# Códigos de país (compliance.priority_queue_b2c.country_code, ISO2) que usan
# la plantilla de Argentina en vez de la general. Si Argentina llega a operar
# bajo otro código en algún origen de datos, agregarlo acá.
_ARGENTINA_COUNTRY_CODES = {"AR"}


def _pick_b2c_template_key(country_code: str | None) -> str:
    """Determina automáticamente qué plantilla B2C usar según el país de
    origen del cliente. Cualquier país que no sea Argentina cae en la
    plantilla general."""
    normalized = (country_code or "").strip().upper()
    return "argentina_b2c" if normalized in _ARGENTINA_COUNTRY_CODES else "general_b2c"


def _load_email_template_html(template_key: str) -> str:
    meta = EMAIL_TEMPLATE_CATALOG.get(template_key)
    if not meta:
        raise ValueError(f"Plantilla de correo desconocida: {template_key}")
    return (EMAIL_TEMPLATES_DIR / meta["file"]).read_text(encoding="utf-8")


# Bloques removibles de las plantillas de solicitud, en el orden en que
# aparecen en el correo. Cada punto numerado está envuelto en la plantilla con
# <!--B:nombre--> ... <!--/B--> y su número reemplazado por {{N}}.
_TEMPLATE_BLOCK_ORDER = ["domicilio", "formulario", "origen", "motivo"]

# Qué bloque del correo corresponde a cada categoría del checklist. Dos
# categorías caen en el mismo bloque (los comprobantes de respaldo son parte
# del punto de origen de fondos), así que alcanza con que una esté pedida.
_CATEGORY_TO_BLOCK = {
    "Domicilio": "domicilio",
    "Identidad/Datos personales": "formulario",
    "Origen de fondo": "origen",
    "Comprobantes/Soporte": "origen",
    "Relación/Beneficiario": "motivo",
}
_BLOCK_RE = re.compile(r"<!--B:([a-z]+)-->.*?<!--/B-->", re.DOTALL)


def _filtrar_bloques(html_body: str, documentos: list[str] | None) -> str:
    """Deja en el correo solo los puntos correspondientes a los documentos
    pedidos y renumera los que quedan.

    Si el analista destilda una categoría, ese punto no debe aparecer en el
    correo — antes se mandaba siempre el texto completo sin importar la
    selección. Sin selección conocida se manda todo (comportamiento seguro)."""
    if "<!--B:" not in html_body:
        return html_body  # plantilla sin bloques (texto libre / en blanco)

    pedidos = {_CATEGORY_TO_BLOCK[d] for d in (documentos or []) if d in _CATEGORY_TO_BLOCK}
    if not pedidos:
        pedidos = set(_TEMPLATE_BLOCK_ORDER)

    def _quitar(m):
        return m.group(0) if m.group(1) in pedidos else ""

    html_body = _BLOCK_RE.sub(_quitar, html_body)

    # Renumerar: los {{N}} que sobrevivieron pasan a 1, 2, 3... en orden.
    contador = [0]

    def _numerar(_m):
        contador[0] += 1
        return str(contador[0])

    html_body = re.sub(r"\{\{N\}\}", _numerar, html_body)
    # Se limpian los marcadores para no dejar comentarios en el correo final.
    return html_body.replace("<!--/B-->", "").replace("<!--B:", "<!--")


def _render_email_template(template_key: str, nombre_completo: str, texto_libre: str | None = None,
                           documentos: list[str] | None = None) -> str:
    """Renderiza una de las 4 plantillas oficiales, reemplazando el nombre
    del cliente y, para texto_libre, el contenido escrito a mano por el
    analista."""
    html_body = _load_email_template_html(template_key)
    html_body = html_body.replace(_SF_MERGE_FIELD_NOMBRE, nombre_completo or "cliente")
    if (EMAIL_TEMPLATE_CATALOG.get(template_key) or {}).get("requires_custom_text"):
        custom = (texto_libre or "").strip()
        custom_html = html.escape(custom).replace("\n", "<br>") if custom else "TEXTO LIBRE"
        html_body = html_body.replace(_TEXTO_LIBRE_PLACEHOLDER, custom_html)
    return _filtrar_bloques(html_body, documentos)


def _email_template_attachment(template_key: str) -> tuple[str, bytes] | None:
    """Devuelve (nombre_archivo, contenido) del formulario que corresponde a
    la plantilla, o None si esa plantilla no lleva adjunto (texto_libre)."""
    meta = EMAIL_TEMPLATE_CATALOG.get(template_key) or {}
    filename = meta.get("attachment")
    if not filename:
        return None
    return filename, (EMAIL_ATTACHMENTS_DIR / filename).read_bytes()


def preview_email_template(body: dict):
    """Devuelve el HTML del correo tal como se va a enviar, para que el
    analista lo vea antes de mandarlo."""
    template_key = (body.get("template_key") or "").strip()
    if template_key not in EMAIL_TEMPLATE_CATALOG:
        return resp(400, {"error": f"template_key desconocido: {template_key}"})
    meta = EMAIL_TEMPLATE_CATALOG[template_key]
    try:
        html_body = _render_email_template(
            template_key,
            (body.get("nombre") or "").strip(),
            body.get("texto_libre", ""),
            body.get("documentos") or None,
        )
    except Exception as e:
        return resp(200, {"error": f"No se pudo renderizar la plantilla: {e}"})
    return resp(200, {
        "html": html_body,
        "adjunto": meta.get("attachment") or "",
        "requiere_texto": meta.get("requires_custom_text", False),
        "label": meta.get("label", template_key),
    })


def list_email_templates(_qs: dict | None = None):
    """Catálogo de plantillas para el selector manual del frontend (modal de
    solicitar/reenviar documentos)."""
    return resp(200, {
        "templates": [
            {"key": key, "label": meta["label"], "has_attachment": bool(meta["attachment"]),
             "requires_custom_text": meta["requires_custom_text"]}
            for key, meta in EMAIL_TEMPLATE_CATALOG.items()
        ]
    })

runs_table = dynamodb.Table(RUNS_TABLE_NAME)
catalog_table = dynamodb.Table(CATALOG_TABLE_NAME)

# ---------------------------------------------------------------------------
# Built-in report definitions (mirrors REPORT_CONFIGS in handler.py)
# ---------------------------------------------------------------------------
BUILTIN_REPORTS = [
    # ─── Priorización de Alertas (datos de prueba) ──────────────────────────
    {
        "report_name": "priority_queue_test_alerts",
        "display_name": "Priorización de Alertas — Datos de Prueba",
        "description": "Priorización de alertas (datos de prueba): 5 filas ficticias con prioridad ya asignada, para probar el flujo manual completo (documentos, correo, caso) sin tocar datos reales.",
        "category": "priorizacion",
        "category_label": "Priorización (Pruebas)",
        "is_custom": False,
        "params": [],
    },
    # ─── AML Transaccional ───────────────────────────────────────────────────
    {
        "report_name": "high_risk_countries",
        "display_name": "Transacciones a Países Alto Riesgo",
        "description": "Transacciones outbound a jurisdicciones FATF/OFAC de alto riesgo. Incluye flag de mismatch SWIFT.",
        "category": "aml_transaccional",
        "category_label": "AML Transaccional",
        "is_custom": False,
        "params": [
            {"name": "since_date", "type": "date", "label": "Desde fecha", "default": "first_day_of_month"},
            {"name": "only_successful", "type": "bool", "label": "Solo transferencias exitosas", "default": False},
        ],
    },
    {
        "report_name": "amount_ranges_by_country",
        "display_name": "Rangos de Monto por País (7d)",
        "description": "Volumen y cantidad de transacciones agrupadas por rango USD × país destino, últimos 7 días.",
        "category": "aml_transaccional",
        "category_label": "AML Transaccional",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "top_customers_by_range_country",
        "display_name": "Top Clientes por Rango y País (7d)",
        "description": "Top 15 clientes por cantidad de transacciones para cada combinación país × rango USD, últimos 7 días.",
        "category": "aml_transaccional",
        "category_label": "AML Transaccional",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "tax_haven_transactions",
        "display_name": "Transacciones a Régimen Fiscal Preferencial (90d)",
        "description": "Transacciones exitosas hacia países con régimen fiscal preferencial o zonas francas en últimos 90 días.",
        "category": "aml_transaccional",
        "category_label": "AML Transaccional",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "tax_haven_funding",
        "display_name": "Fondeos desde Régimen Fiscal Preferencial (7d)",
        "description": "Cash calls entrantes (CR pagados) cuyo remitente proviene de países con régimen fiscal preferencial.",
        "category": "aml_transaccional",
        "category_label": "AML Transaccional",
        "is_custom": False,
        "params": [],
    },
    # ─── Patrones AML ────────────────────────────────────────────────────────
    {
        "report_name": "payin_payout_accumulation",
        "display_name": "Acumulación Pay In → Pay Out (7d)",
        "description": "Clientes con múltiples pay-ins seguidos de pay-outs en 7 días. Indica posible layering.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "small_payin_structuring",
        "display_name": "Pay In Pequeños → Pay Out (Smurfing, 7d)",
        "description": "Clientes con 5+ pay-ins < USD 1.000 seguidos de pay-outs. Patrón de estructuración/smurfing.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "velocity_payin_payout",
        "display_name": "Velocity Pay In ↔ Pay Out < 24h (7d)",
        "description": "Pares de pay-in y pay-out del mismo cliente separados por menos de 24 horas.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "external_funder_single",
        "display_name": "Tercero que Fondea Una Sola Cuenta (7d)",
        "description": "Personas externas que fondean repetidamente (3+) una única cuenta de cliente.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "external_funder_multiple",
        "display_name": "Tercero que Fondea Múltiples Cuentas (7d)",
        "description": "Personas externas que fondean 2+ cuentas distintas de clientes en los últimos 7 días.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "circular_transactions",
        "display_name": "Circularidad DNI Cliente ↔ Beneficiario (90d)",
        "description": "Clientes que envían fondos a personas que también les envían fondos — posible circularidad.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "structuring_detection",
        "display_name": "Estructuración / Fraccionamiento (7d)",
        "description": "Clientes con 5+ transacciones todas < USD 1.000 y volumen total > USD 3.000.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "shared_beneficiary",
        "display_name": "Beneficiario Compartido por Múltiples Remitentes (7d)",
        "description": "Beneficiarios que reciben fondos de 3+ clientes distintos en los últimos 7 días.",
        "category": "patrones_aml",
        "category_label": "Patrones AML",
        "is_custom": False,
        "params": [],
    },
    # ─── Comportamiento Clientes ──────────────────────────────────────────────
    {
        "report_name": "customer_metrics_7d",
        "display_name": "Métricas por Cliente B2C (7d)",
        "description": "Resumen de transacciones, beneficiarios únicos, montos y canales por cliente individual.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "beneficiary_concentration",
        "display_name": "Concentración de Beneficiarios (7d)",
        "description": "Clientes con 5+ transacciones pero solo 1-2 beneficiarios distintos — posible concentración sospechosa.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "beneficiary_dispersion",
        "display_name": "Dispersión de Beneficiarios (7d)",
        "description": "Clientes individuales que envían a 5+ beneficiarios distintos — posible dispersión de fondos.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "outbound_bank_change",
        "display_name": "Cambio de Banco Outbound (30d vs 7d)",
        "description": "Clientes que usaron un banco outbound nuevo en los últimos 7 días que no habían usado antes.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "new_corridor_detection",
        "display_name": "Corredor Nuevo para el Cliente (7d vs 90d)",
        "description": "Clientes que usaron una ruta origen/destino nueva en los últimos 7d que no habían usado en 90d.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "high_volume_vs_historical",
        "display_name": "Alto Volumen vs Histórico (7d vs 90d)",
        "description": "Clientes cuyo ticket promedio o volumen diario en 7d es 3x+ mayor que su histórico de 90 días.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "swift_mismatch_detection",
        "display_name": "Mismatch SWIFT vs País Beneficiario (30d)",
        "description": "Transacciones donde el código de país del SWIFT del banco no coincide con el país beneficiario.",
        "category": "comportamiento_clientes",
        "category_label": "Comportamiento Clientes",
        "is_custom": False,
        "params": [],
    },
    # ─── KYC / Jumio ─────────────────────────────────────────────────────────
    {
        "report_name": "jumio_kyc_approval_rates",
        "display_name": "Tasas de Aprobación / Rechazo KYC por Flujo",
        "description": "Estadísticas agregadas de aprobación y rechazo de validaciones Jumio por business flow y país.",
        "category": "kyc_jumio",
        "category_label": "KYC / Jumio",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "jumio_duplicate_flows",
        "display_name": "Documentos Jumio Duplicados / Flujos Múltiples",
        "description": "DNIs con múltiples clientes, cuentas Jumio o business flows — posible duplicación de identidad.",
        "category": "kyc_jumio",
        "category_label": "KYC / Jumio",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "b2c_as_legal_rep",
        "display_name": "Clientes B2C que son Representantes Legales (B2B)",
        "description": "Personas físicas con cuenta B2C activa que también son representantes legales de empresas B2B.",
        "category": "kyc_jumio",
        "category_label": "KYC / Jumio",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "top_companies_by_legal_reps",
        "display_name": "Top 15 Empresas con Más Representantes Legales",
        "description": "Empresas activas con mayor cantidad de representantes legales distintos registrados.",
        "category": "kyc_jumio",
        "category_label": "KYC / Jumio",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "age_anomaly_customers",
        "display_name": "Clientes con Anomalía de Edad (<18 o >90 años)",
        "description": "Clientes activos con fecha de nacimiento que indica menor de 18 años o mayor de 90 años.",
        "category": "kyc_jumio",
        "category_label": "KYC / Jumio",
        "is_custom": False,
        "params": [],
    },
    # ─── Crypto / Bridge ─────────────────────────────────────────────────────
    {
        "report_name": "crypto_bridge_transactions",
        "display_name": "Transacciones Bridge/Crypto (30d)",
        "description": "Resumen de transacciones involucrando Bridge o métodos crypto, agrupadas por método, estado y corredor.",
        "category": "crypto_bridge",
        "category_label": "Crypto / Bridge",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "crypto_bridge_cash_calls",
        "display_name": "Cash Calls Bridge/Crypto (30d)",
        "description": "Cash calls con método Bridge o moneda USDC/USDT/BTC/ETH en los últimos 30 días.",
        "category": "crypto_bridge",
        "category_label": "Crypto / Bridge",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "crypto_high_risk_destinations",
        "display_name": "Crypto hacia Países de Riesgo (30d)",
        "description": "Cash calls crypto cuyo beneficiario está en países de la lista de alto riesgo.",
        "category": "crypto_bridge",
        "category_label": "Crypto / Bridge",
        "is_custom": False,
        "params": [],
    },
    {
        "report_name": "crypto_full_bridge_activity",
        "display_name": "Actividad Completa Bridge (30d)",
        "description": "Vista completa de clientes Bridge: wallets, balances, transacciones crypto, transferencias y cash calls.",
        "category": "crypto_bridge",
        "category_label": "Crypto / Bridge",
        "is_custom": False,
        "params": [],
    },
    # ─── Institucional ───────────────────────────────────────────────────────
    {
        "report_name": "institutional_active_transactions",
        "display_name": "Transacciones Recientes — Clientes Institucionales Activos",
        "description": "Últimas transacciones de clientes institucionales que no están bloqueados ni fully blocked.",
        "category": "institucional",
        "category_label": "Institucional",
        "is_custom": False,
        "params": [],
    },
]


# ---------------------------------------------------------------------------
# Dashboard SQL queries (informational widgets, last 7 days of successful TRX)
# ---------------------------------------------------------------------------
_SQL_DAILY_EVOLUTION = """
SELECT CAST(t.start_date AS DATE) AS trx_date,
    COALESCE(t.payment_method, 'SIN_PAYMENT_METHOD') AS payment_method,
    COALESCE(t.outbound_bank_name, 'SIN_OUTBOUND_BANK') AS outbound_bank_name,
    t.origin_currency, t.destiny_currency,
    COUNT(*) AS total_transactions,
    COUNT(DISTINCT t.customer_id) AS unique_customers,
    SUM(t.destiny_amount_usd) AS total_amount_usd,
    AVG(t.destiny_amount_usd) AS avg_ticket_usd
FROM "db_prod"."transaction"."transaction" AS t
WHERE t.start_date >= DATEADD(day, -7, CURRENT_DATE)
  AND UPPER(t.tx_status) = 'TRANSFERENCIA_EXITOSA'
GROUP BY CAST(t.start_date AS DATE),
    COALESCE(t.payment_method, 'SIN_PAYMENT_METHOD'),
    COALESCE(t.outbound_bank_name, 'SIN_OUTBOUND_BANK'),
    t.origin_currency, t.destiny_currency
ORDER BY trx_date ASC, total_amount_usd DESC
"""

_SQL_OVER_300K = """
SELECT CAST(t.start_date AS DATE) AS trx_date,
    COALESCE(t.payment_method, 'SIN_PAYMENT_METHOD') AS payment_method,
    COALESCE(t.outbound_bank_name, 'SIN_OUTBOUND_BANK') AS outbound_bank_name,
    t.origin_currency, t.destiny_currency,
    COUNT(*) AS trx_over_300k,
    COUNT(DISTINCT t.customer_id) AS unique_customers_over_300k,
    SUM(t.destiny_amount_usd) AS total_amount_usd_over_300k,
    AVG(t.destiny_amount_usd) AS avg_ticket_usd_over_300k,
    MAX(t.destiny_amount_usd) AS max_ticket_usd
FROM "db_prod"."transaction"."transaction" AS t
WHERE t.start_date >= DATEADD(day, -7, CURRENT_DATE)
  AND UPPER(t.tx_status) = 'TRANSFERENCIA_EXITOSA'
  AND t.destiny_amount_usd >= 300000
GROUP BY CAST(t.start_date AS DATE),
    COALESCE(t.payment_method, 'SIN_PAYMENT_METHOD'),
    COALESCE(t.outbound_bank_name, 'SIN_OUTBOUND_BANK'),
    t.origin_currency, t.destiny_currency
ORDER BY trx_date ASC, total_amount_usd_over_300k DESC
"""

_SQL_BY_COUNTRY = """
SELECT t.beneficiary_country_code,
    MAX(t.beneficiary_country_name) AS beneficiary_country_name,
    COUNT(*) AS total_transactions,
    COUNT(DISTINCT t.customer_id) AS unique_customers,
    COUNT(DISTINCT t.beneficiary_id) AS unique_beneficiaries,
    SUM(t.destiny_amount_usd) AS total_amount_usd,
    AVG(t.destiny_amount_usd) AS avg_ticket_usd,
    MAX(t.destiny_amount_usd) AS max_ticket_usd
FROM "db_prod"."transaction"."transaction" AS t
WHERE t.start_date >= DATEADD(day, -7, CURRENT_DATE)
  AND UPPER(t.tx_status) = 'TRANSFERENCIA_EXITOSA'
GROUP BY t.beneficiary_country_code
ORDER BY total_amount_usd DESC
"""


# ---------------------------------------------------------------------------
# Redshift Data API helpers
# ---------------------------------------------------------------------------
def _esc(s) -> str:
    """Escape a value for safe inclusion in a Redshift SQL string literal."""
    return str(s).replace("'", "''")


def _rs_exec(sql: str) -> list[dict]:
    """Execute SQL via Redshift Data API; poll until done; return rows as list of dicts."""
    try:
        # Redshift Data API rejects trailing semicolons
        sql = sql.strip().rstrip(";").strip()
        resp_exec = redshift_data.execute_statement(
            ClusterIdentifier=CLUSTER_ID,
            Database=DATABASE_NAME,
            DbUser=DB_USER,
            Sql=sql,
        )
        statement_id = resp_exec["Id"]

        deadline = time.time() + 30
        while time.time() < deadline:
            desc = redshift_data.describe_statement(Id=statement_id)
            status = desc["Status"]
            if status == "FINISHED":
                if not desc.get("HasResultSet"):
                    return []
                rows: list[dict] = []
                columns: list[str] = []
                next_token = None
                while True:
                    kwargs = {"Id": statement_id}
                    if next_token:
                        kwargs["NextToken"] = next_token
                    result = redshift_data.get_statement_result(**kwargs)
                    if not columns:
                        columns = [c["name"] for c in result["ColumnMetadata"]]
                    for record in result["Records"]:
                        row = {}
                        for i, cell in enumerate(record):
                            if cell.get("isNull"):
                                row[columns[i]] = None
                            elif "stringValue" in cell:
                                row[columns[i]] = cell["stringValue"]
                            elif "longValue" in cell:
                                row[columns[i]] = cell["longValue"]
                            elif "doubleValue" in cell:
                                row[columns[i]] = cell["doubleValue"]
                            elif "booleanValue" in cell:
                                row[columns[i]] = cell["booleanValue"]
                            else:
                                row[columns[i]] = None
                        rows.append(row)
                    next_token = result.get("NextToken")
                    if not next_token:
                        break
                return rows
            if status in ("FAILED", "ABORTED"):
                raise RuntimeError(f"Redshift query {status}: {desc.get('Error', 'unknown error')}")
            time.sleep(0.5)

        raise RuntimeError("Redshift query timed out after 30s")

    except RuntimeError:
        raise
    except Exception as e:
        msg = str(e)
        if "paused" in msg.lower() or "unavailable" in msg.lower() or "not available" in msg.lower():
            raise RuntimeError(f"Redshift cluster is not available (may be paused): {msg}") from e
        raise


def _rs_get_rows(stmt_id: str) -> list[dict]:
    """Fetch all result rows from a FINISHED Redshift Data API statement (handles pagination)."""
    rows: list[dict] = []
    columns: list[str] = []
    next_token = None
    while True:
        kwargs: dict = {"Id": stmt_id}
        if next_token:
            kwargs["NextToken"] = next_token
        result = redshift_data.get_statement_result(**kwargs)
        if not columns:
            columns = [c["name"] for c in result["ColumnMetadata"]]
        for record in result["Records"]:
            row: dict = {}
            for i, cell in enumerate(record):
                if cell.get("isNull"):
                    row[columns[i]] = None
                elif "stringValue" in cell:
                    row[columns[i]] = cell["stringValue"]
                elif "longValue" in cell:
                    row[columns[i]] = cell["longValue"]
                elif "doubleValue" in cell:
                    row[columns[i]] = cell["doubleValue"]
                elif "booleanValue" in cell:
                    row[columns[i]] = cell["booleanValue"]
                else:
                    row[columns[i]] = None
            rows.append(row)
        next_token = result.get("NextToken")
        if not next_token:
            break
    return rows


def _rs_exec_multi(sqls: list, timeout_s: int = 90) -> list:
    """Submit multiple SQL statements in parallel via Redshift Data API and poll until all done.

    Exploits the async nature of the Data API: all statements are submitted first (no waiting),
    then polled together every second.  Individual failures return [] (non-blocking).
    Returns a list of row-lists in the same order as the input sqls.
    """
    if not sqls:
        return []

    # Submit all statements at once
    stmt_ids: list[str] = []
    for sql in sqls:
        r = redshift_data.execute_statement(
            ClusterIdentifier=CLUSTER_ID,
            Database=DATABASE_NAME,
            DbUser=DB_USER,
            Sql=sql.strip(),
        )
        stmt_ids.append(r["Id"])

    results: list = [[] for _ in sqls]
    pending: set = set(range(len(sqls)))
    deadline = time.time() + timeout_s

    while pending and time.time() < deadline:
        time.sleep(1.0)
        for i in list(pending):
            try:
                desc = redshift_data.describe_statement(Id=stmt_ids[i])
                status = desc["Status"]
                if status == "FINISHED":
                    pending.discard(i)
                    if desc.get("HasResultSet"):
                        results[i] = _rs_get_rows(stmt_ids[i])
                    # else results[i] stays []
                elif status in ("FAILED", "ABORTED"):
                    pending.discard(i)
                    # results[i] stays []
            except Exception:
                pending.discard(i)

    # Any still pending after timeout → left as []
    return results


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
class _Encoder(json.JSONEncoder):
    """Handle Decimal (DynamoDB numbers) and datetime."""
    def default(self, o):
        if isinstance(o, decimal.Decimal):
            return float(o) if o % 1 else int(o)
        if isinstance(o, (dt.datetime, dt.date)):
            return o.isoformat()
        return super().default(o)


def resp(status: int, body) -> dict:
    return {
        "statusCode": status,
        "headers": {
            "Content-Type": "application/json",
        },
        "body": json.dumps(body, cls=_Encoder),
    }


def get_user_email(event: dict) -> str:
    claims = (
        event.get("requestContext", {})
        .get("authorizer", {})
        .get("jwt", {})
        .get("claims", {})
    )
    return claims.get("email") or claims.get("cognito:username", "unknown")


def notify_alert(body: dict):
    """Send a notification email via Gmail SMTP when an alert is assigned."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    to_email = str(body.get("to_email", "")).strip()
    to_name = str(body.get("to_name", "")).strip()
    assigned_by = str(body.get("assigned_by", "")).strip()
    entity_value = str(body.get("entity_value", "")).strip()
    entity_field = str(body.get("entity_field", "")).strip()
    report_name = str(body.get("report_name", "")).strip()
    note = str(body.get("note", "")).strip()

    if not to_email or "@" not in to_email:
        return resp(400, {"error": "to_email is required"})

    # Misma fuente de credencial que _send_email (Secrets Manager con fallback
    # a la variable de entorno) — antes leía la env var por su cuenta y se
    # quedaba afuera cuando se rotaba la clave.
    gmail_user = GMAIL_USER
    gmail_app_password = _get_gmail_password()
    if not gmail_user or not gmail_app_password:
        return resp(500, {"error": "No hay credencial de Gmail configurada (Secrets Manager / GMAIL_APP_PASSWORD)"})

    subject = f"[WatchTower AML] Nueva alerta asignada: {entity_value}"
    note_row = (
        f'<tr><td style="padding:8px;border:1px solid #ddd;background:#f8f8f8"><strong>Nota</strong></td>'
        f'<td style="padding:8px;border:1px solid #ddd">{note}</td></tr>'
        if note else ""
    )
    body_html = f"""<html><body style="font-family:Arial,sans-serif;color:#333">
<h2 style="color:#1B3A6B">&#128737;&#65039; WatchTower AML &#8212; Nueva Alerta Asignada</h2>
<p>Hola {to_name or to_email},</p>
<p><strong>{assigned_by}</strong> te asign&#243; una alerta para revisar:</p>
<table style="border-collapse:collapse;width:100%;max-width:500px">
  <tr><td style="padding:8px;border:1px solid #ddd;background:#f8f8f8"><strong>Campo</strong></td>
      <td style="padding:8px;border:1px solid #ddd">{entity_field}</td></tr>
  <tr><td style="padding:8px;border:1px solid #ddd;background:#f8f8f8"><strong>Valor</strong></td>
      <td style="padding:8px;border:1px solid #ddd"><strong>{entity_value}</strong></td></tr>
  <tr><td style="padding:8px;border:1px solid #ddd;background:#f8f8f8"><strong>Reporte</strong></td>
      <td style="padding:8px;border:1px solid #ddd">{report_name.replace("_", " ")}</td></tr>
  {note_row}
</table>
<p style="margin-top:20px">
  <a href="https://bmackenna-g66.github.io/compliance-redshift-reports"
     style="background:#f97316;color:white;padding:10px 20px;text-decoration:none;border-radius:6px;font-weight:bold">
    Ir a WatchTower AML &rarr;
  </a>
</p>
<p style="color:#999;font-size:12px;margin-top:20px">Mensaje autom&#225;tico de WatchTower AML.</p>
</body></html>"""

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"Compliance Global66 <{gmail_user}>"
        msg["To"] = to_email
        msg.attach(MIMEText(body_html, "html", "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(gmail_user, gmail_app_password)
            smtp.sendmail(gmail_user, [to_email], msg.as_bytes())

        return resp(200, {"message": "Notification sent"})
    except Exception as e:
        return resp(500, {"error": f"Failed to send email: {e}"})


_B2C_QUERY = """
WITH latest_kyc_document AS (
    SELECT kd.customer_id, kd.document_number, kd.document_type,
        kd.country_code AS document_country_code, kd.approval_status,
        ROW_NUMBER() OVER (PARTITION BY kd.customer_id ORDER BY COALESCE(kd.updated_at, kd.created_at) DESC) AS rn
    FROM "db_prod"."customer"."kyc_document" kd WHERE kd.document_number IS NOT NULL
),
latest_customer_compliance AS (
    SELECT cc.customer_id, cc.compliance_status, cc.status_created_at, cc.compliance_agent, cc.compliance_channel,
        ROW_NUMBER() OVER (PARTITION BY cc.customer_id ORDER BY cc.status_created_at DESC) AS rn
    FROM "db_prod"."customer"."compliance" cc
),
latest_customer_kyc AS (
    SELECT ck.customer_id, ck.kyc_status, ck.step_onboarding_id, ck.created_at AS kyc_created_at,
        ROW_NUMBER() OVER (PARTITION BY ck.customer_id ORDER BY ck.created_at DESC) AS rn
    FROM "db_prod"."customer"."customer_kyc" ck
),
latest_customer_work AS (
    SELECT cw.customer_id, cw.profession, cw.work_position, cw.workplace,
        ROW_NUMBER() OVER (PARTITION BY cw.customer_id ORDER BY COALESCE(cw.updated_at, cw.created_at) DESC) AS rn
    FROM "db_prod"."customer"."customer_work" cw
),
latest_customer_segmentation AS (
    SELECT s.customer_id, s.segmentation,
        ROW_NUMBER() OVER (PARTITION BY s.customer_id ORDER BY s.last_updated_date DESC) AS rn
    FROM "db_prod"."customer"."segmentation" s
),
latest_virtual_account AS (
    SELECT cva.customer_id, cva.id AS customer_virtual_account_id, cva.account_id,
        cva.virtual_account_number, cva.global_account_number,
        cva.country_code AS virtual_account_country_code, cva.virtual_account_type,
        cva.is_enabled, cva.created_at AS virtual_account_created_at,
        ROW_NUMBER() OVER (PARTITION BY cva.customer_id ORDER BY cva.created_at DESC) AS rn
    FROM "db_prod"."product_gateway"."customer_virtual_account" cva
)
SELECT c.customer_id, c.name AS nombre, c.last_name AS apellido, c.email, c.phone_number, c.calling_code,
    c.country_code AS pais_residencia, c.nationality_code AS nacionalidad, c.risk_level, c.created_date AS fecha_onboarding,
    kd.document_number AS dni, kd.document_type AS tipo_documento, kd.document_country_code AS pais_documento, kd.approval_status AS estado_documento,
    lc.compliance_status, lc.status_created_at AS compliance_status_created_at, lc.compliance_agent, lc.compliance_channel,
    ck.kyc_status, sok.step AS onboarding_step, ck.kyc_created_at,
    seg.segmentation, cw.profession, cw.work_position, cw.workplace,
    va.customer_virtual_account_id, va.account_id, va.virtual_account_number, va.global_account_number,
    va.virtual_account_country_code, va.virtual_account_type, va.is_enabled AS virtual_account_active, va.virtual_account_created_at,
    DATEDIFF(day, c.created_date, va.virtual_account_created_at) AS dias_desde_onboarding_hasta_cuenta_virtual,
    DATEDIFF(day, c.created_date, CURRENT_DATE) AS dias_desde_onboarding
FROM "db_prod"."customer"."customer_v2" c
LEFT JOIN latest_kyc_document kd ON c.customer_id = kd.customer_id AND kd.rn = 1
LEFT JOIN latest_customer_compliance lc ON c.customer_id = lc.customer_id AND lc.rn = 1
LEFT JOIN latest_customer_kyc ck ON c.customer_id = ck.customer_id AND ck.rn = 1
LEFT JOIN "db_prod"."customer"."step_onboarding_kyc" sok ON ck.step_onboarding_id = sok.id
LEFT JOIN latest_customer_work cw ON c.customer_id = cw.customer_id AND cw.rn = 1
LEFT JOIN latest_customer_segmentation seg ON c.customer_id = seg.customer_id AND seg.rn = 1
LEFT JOIN latest_virtual_account va ON c.customer_id = va.customer_id AND va.rn = 1
WHERE sok.step = 'HOME' AND ck.kyc_status = 'APPROVED'
__FILTER__
ORDER BY c.created_date DESC
LIMIT 5
"""

_B2B_QUERY = """
SELECT co.company_id, co.name AS company_name, co.identification_number AS company_document_number,
    co.identification_type AS company_document_type, co.username, co.phone_country_code, co.phone_number,
    co.compliance_status, co.compliance_status_comment,
    co.kyc_stage_1, co.kyc_stage_1_approved_date, co.kyc_stage_1_rejected_date, co.kyc_stage_1_requested_date, co.kyc_stage_1_comment,
    co.kyc_stage_2, co.kyc_stage_2_approved_date, co.kyc_stage_2_rejected_date, co.kyc_stage_2_requested_date, co.kyc_stage_2_comment,
    co.kyc_stage_3, co.kyc_stage_3_approved_date, co.kyc_stage_3_rejected_date, co.kyc_stage_3_requested_date, co.kyc_stage_3_comment,
    co.risk_level, co.risk_level_regcheq, co.dni_regcheq,
    co.activity AS company_activity_raw, act.name AS activity_name, act.risk_level AS activity_risk_level, ind.name AS industry_name,
    co.activity_start_date, co.company_financial_activity_id, co.ind_activity,
    co.monthly_income, co.monthly_expenses, co.estimated_annual_billings, co.total_assets, co.total_liabilities,
    co.shipment_amounts, co.shipment_frequency, co.purpose_use,
    co.has_board_directors, co.has_joint_administration, co.has_partners_ten_sharedholding,
    co.legal_representatives_count, co.institutional, co.multi_user_enabled, co.crs, co.fatca,
    ac.country AS company_address_country, ac.state AS company_address_state, ac.city AS company_address_city,
    ac.district AS company_address_district, ac.street AS company_address_street, ac.number AS company_address_number,
    ac.apt AS company_address_apt, ac.floor AS company_address_floor, ac.postal_code AS company_address_postal_code,
    co.create_at AS company_created_at, co.record_created_at,
    DATEDIFF(day, co.create_at, CURRENT_DATE) AS dias_desde_creacion_empresa
FROM "db_prod"."company"."company" AS co
LEFT JOIN "db_prod"."company"."activity" AS act ON co.ind_activity = act.id
LEFT JOIN "db_prod"."company"."industry" AS ind ON act.industry_id = ind.id
LEFT JOIN "db_prod"."company"."address_country" AS ac ON co.company_address_country = ac.address_id
__WHERE__
ORDER BY co.create_at DESC
LIMIT 5
"""


def _lookup_customer_rows(identifier: str, kind: str = "b2c") -> list[dict]:
    """Ficha KYC/compliance de un cliente (b2c) o empresa (b2b).

    Es el mismo motor que usa la búsqueda de Análisis Individual; se extrajo
    a función aparte para poder reutilizarlo desde la ficha del caso sin
    duplicar las queries."""
    identifier = str(identifier or "").strip()
    if not identifier:
        raise ValueError("identifier vacío")
    if any(c in identifier for c in ("'", ";", "--", "\\")):
        raise ValueError("identifier inválido")
    safe = identifier.replace("'", "''")

    if kind == "b2b":
        if identifier.isdigit():
            where = f"WHERE co.company_id = {int(identifier)}"
        else:
            where = f"WHERE co.identification_number = '{safe}' OR co.username = '{safe}'"
        sql = _B2B_QUERY.replace("__WHERE__", where)
    else:
        if identifier.isdigit():
            extra = f"AND c.customer_id = {int(identifier)}"
        elif "@" in identifier:
            extra = f"AND LOWER(c.email) = LOWER('{safe}')"
        else:
            extra = f"AND kd.document_number = '{safe}'"
        sql = _B2C_QUERY.replace("__FILTER__", extra)

    return _rs_exec_multi([sql], timeout_s=90)[0]


def _display_name_from_profile(profile: dict, kind: str) -> str:
    """Nombre legible del cliente/empresa a partir de su ficha.

    Los alias de las queries están en español (nombre/apellido, razon_social),
    así que se prueban esos primero y los nombres crudos de columna después.
    El correo queda como último recurso, no como nombre."""
    if not isinstance(profile, dict):
        return ""

    def _v(*claves):
        for k in claves:
            v = str(profile.get(k) or "").strip()
            if v and v.lower() not in ("none", "null"):
                return v
        return ""

    if kind == "b2b":
        return _v("razon_social", "company_name", "nombre_empresa", "nombre",
                  "name", "rep_name", "username")

    nombre = " ".join(x for x in (_v("nombre", "name", "first_name"),
                                  _v("apellido", "last_name")) if x).strip()
    return nombre or _v("nombre_completo", "email")


def search_customer_b2c(body: dict):
    try:
        rows = _lookup_customer_rows(body.get("identifier", ""), "b2c")
    except ValueError as e:
        return resp(400, {"error": str(e)})
    return resp(200, {"rows": rows, "count": len(rows)})


def search_customer_b2b(body: dict):
    try:
        rows = _lookup_customer_rows(body.get("identifier", ""), "b2b")
    except ValueError as e:
        return resp(400, {"error": str(e)})
    return resp(200, {"rows": rows, "count": len(rows)})


def lookup_case_entity(body: dict):
    """Resuelve el nombre de un cliente/empresa desde su ID, para no tener
    que tipearlo a mano al crear un caso."""
    kind = "b2b" if str(body.get("entity_type", "")).lower().startswith("comp") else "b2c"
    try:
        rows = _lookup_customer_rows(body.get("entity_id", ""), kind)
    except ValueError as e:
        return resp(400, {"error": str(e)})
    except Exception as e:
        return resp(200, {"found": False, "error": f"No se pudo consultar: {e}"})
    if not rows:
        return resp(200, {"found": False, "error": "No se encontró el cliente con ese ID"})
    profile = rows[0]
    return resp(200, {
        "found": True,
        "entity_name": _display_name_from_profile(profile, kind),
        "profile": profile,
        "kind": kind,
    })


# ---------------------------------------------------------------------------
# AI generate proxy — calls Gemini using AI_API_KEY env var
# ---------------------------------------------------------------------------
_GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"

def ai_generate(body: dict):
    import urllib.request as _ur
    api_key = os.environ.get("AI_API_KEY", "")
    if not api_key:
        return resp(500, {"error": "AI_API_KEY not configured"})
    prompt = str(body.get("prompt", "")).strip()
    if not prompt:
        return resp(400, {"error": "prompt is required"})
    temperature = float(body.get("temperature", 0.3))
    max_tokens = int(body.get("max_tokens", 2048))
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": temperature, "maxOutputTokens": max_tokens},
    }).encode()
    url = f"{_GEMINI_URL}?key={api_key}"
    req = _ur.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with _ur.urlopen(req, timeout=60) as r:
            data = json.loads(r.read())
        text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        return resp(200, {"text": text})
    except _ur.HTTPError as e:
        err_body = e.read().decode(errors="ignore")
        try:
            err_msg = json.loads(err_body).get("error", {}).get("message", err_body)
        except Exception:
            err_msg = err_body
        return resp(502, {"error": err_msg})


# ---------------------------------------------------------------------------
# CUSTOMER CONTEXT — dossier para la IA (CRM siempre + transacciones si cluster on)
# ---------------------------------------------------------------------------
_ALLOWED_DAYS = (5, 15, 30, 60)


def _cluster_available() -> bool:
    try:
        r = redshift.describe_clusters(ClusterIdentifier=CLUSTER_ID)
        return r["Clusters"][0]["ClusterStatus"] == "available"
    except Exception:
        return False


def _customer_cashcall_sql(direction: str, customer_id: str, days: int) -> str:
    """Cash calls de un cliente. direction: 'DR' (pay out) | 'CR' (pay in).
    customer_id ya validado como dígitos; days dentro de _ALLOWED_DAYS."""
    return f"""
SELECT cc.cash_call_id, cc.external_reference_number,
    cc.customer_id, c.email, c.name, c.last_name,
    cc.creation_date, cc.paid_date, cc.type, cc.status, cc.currency_code,
    cc.amount, cc.origin_amount_usd, cc.destiny_amount_usd,
    cc.remitter_name, cc.remitter_lastname, cc.remitter_dni, cc.remitter_email,
    cc.business_bank_id, bb.bank_code, bb.bank_name
FROM "db_prod"."treasury"."cash_call" AS cc
LEFT JOIN "db_prod"."customer"."customer_v2" AS c
    ON cc.customer_id::VARCHAR = c.customer_id::VARCHAR
LEFT JOIN "db_prod"."treasury"."business_bank" AS bb
    ON cc.business_bank_id = bb.business_bank_id
WHERE cc.type = '{direction}'
  AND cc.customer_id::VARCHAR = '{customer_id}'
  AND cc.creation_date >= DATEADD(day, -{days}, CURRENT_DATE)
  AND cc.status = 'PAID'
ORDER BY cc.creation_date DESC
"""


def _tx_summary(rows: list, sample: int = 40) -> dict:
    """Resumen + muestra acotada (para no inflar el prompt de la IA)."""
    def _fnum(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
    total_origin = sum(_fnum(r.get("origin_amount_usd")) for r in rows)
    total_destiny = sum(_fnum(r.get("destiny_amount_usd")) for r in rows)
    return {
        "count": len(rows),
        "total_origin_usd": round(total_origin, 2),
        "total_destiny_usd": round(total_destiny, 2),
        "rows": rows[:sample],
        "truncated": len(rows) > sample,
    }


def _customer_crm_dossier(customer_id: str) -> dict:
    """Historial CRM del cliente desde S3 (siempre disponible)."""
    alerts = _crm_list("alerts")
    cust_alerts = [a for a in alerts
                   if str(a.get("entity_value", "")) == customer_id
                   and (a.get("entity_field") in ("customer_id", "", None) or True)]
    reports = sorted({a.get("report_name", "") for a in cust_alerts if a.get("report_name")})

    cases = _crm_list("cases")
    linked_ids = {a.get("case_id") for a in cust_alerts if a.get("case_id")}
    cust_cases = [c for c in cases
                  if str(c.get("entity_id", "")) == customer_id or c.get("case_id") in linked_ids]

    return {
        "alert_count": len(cust_alerts),
        "recurrent": len(cust_alerts) > 1,
        "combines_alerts": len(reports) > 1,
        "distinct_reports": reports,
        "alerts": [{
            "report_name": a.get("report_name", ""), "reason": a.get("reason", ""),
            "status": a.get("status", ""), "priority": a.get("priority", "medium"),
            "created_at": a.get("created_at", ""), "entity_field": a.get("entity_field", ""),
        } for a in sorted(cust_alerts, key=lambda x: x.get("created_at", ""), reverse=True)],
        "case_count": len(cust_cases),
        "cases": [{
            "case_id": c.get("case_id", ""), "title": c.get("title", ""),
            "status": c.get("status", ""), "priority": c.get("priority", ""),
            "created_at": c.get("created_at", ""),
        } for c in cust_cases],
    }


def customer_context(customer_id: str, days: int):
    customer_id = str(customer_id or "").strip()
    if not customer_id.isdigit():
        return resp(400, {"error": "customer_id debe ser numérico"})
    if days not in _ALLOWED_DAYS:
        days = 30

    crm = _customer_crm_dossier(customer_id)

    tx: dict = {"available": False, "reason": "cluster_paused"}
    if _cluster_available():
        try:
            payout = _rs_exec(_customer_cashcall_sql("DR", customer_id, days))
            payin = _rs_exec(_customer_cashcall_sql("CR", customer_id, days))
            tx = {"available": True, "days": days,
                  "payout": _tx_summary(payout), "payin": _tx_summary(payin)}
        except Exception as e:
            tx = {"available": False, "reason": "error", "error": str(e)}

    return resp(200, {"customer_id": customer_id, "days": days, "crm": crm, "transactions": tx})


# ---------------------------------------------------------------------------
# CASE EXCEL EXPORT
# ---------------------------------------------------------------------------

def export_case(case_id: str):
    """Generate an Excel workbook for a case and return a 1h presigned S3 URL."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return resp(500, {"error": "openpyxl not available"})

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": "Case not found"})

    notes = sorted(case.get("notes", []), key=lambda n: n.get("created_at", ""))
    alerts = [a for a in _crm_list("alerts") if a.get("case_id") == case_id]

    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1E293B")
    HEADER_FONT = Font(bold=True, color="F97316", size=11)
    LABEL_FONT  = Font(bold=True, color="94A3B8")

    # ── Sheet 1: Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    def add_row(label, value, row):
        c_a = ws.cell(row=row, column=1, value=label)
        c_a.font = LABEL_FONT
        c_b = ws.cell(row=row, column=2, value=str(value) if value else "—")
        c_b.alignment = Alignment(wrap_text=True)
        return row + 1

    r = 1
    ws.cell(r, 1, "REPORTE DE CASO — WatchTower AML").font = Font(bold=True, size=14, color="F97316")
    ws.merge_cells(f"A{r}:B{r}")
    r += 2
    for label, key in [
        ("ID del Caso", "case_id"), ("Título", "title"), ("Descripción", "description"),
        ("Estado", "status"), ("Prioridad", "priority"), ("Tipo entidad", "entity_type"),
        ("ID entidad", "entity_id"), ("Reporte origen", "report_name"),
        ("Asignado a", "assigned_to"), ("Creado por", "created_by"),
        ("Fecha creación", "created_at"), ("Última actualización", "updated_at"),
        ("Fecha cierre", "closed_at"),
    ]:
        r = add_row(label, case.get(key, ""), r)

    # ── Sheet 2: Notas ──
    ws2 = wb.create_sheet("Notas")
    ws2.sheet_view.showGridLines = False
    headers = ["#", "Autor", "Fecha", "Contenido"]
    widths   = [5, 30, 22, 80]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws2.column_dimensions[cell.column_letter].width = w
    for i, n in enumerate(notes, 1):
        ws2.append([i, n.get("author_email",""), n.get("created_at",""), n.get("content","")])
        ws2.cell(i+1, 4).alignment = Alignment(wrap_text=True)

    # ── Sheet 3: Alertas vinculadas ──
    ws3 = wb.create_sheet("Alertas vinculadas")
    ws3.sheet_view.showGridLines = False
    a_headers = ["ID Alerta", "Campo", "Valor", "Razón", "Reporte", "Fecha", "Estado"]
    a_widths   = [36, 16, 24, 40, 24, 22, 12]
    for col, (h, w) in enumerate(zip(a_headers, a_widths), 1):
        cell = ws3.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws3.column_dimensions[cell.column_letter].width = w
    for a in alerts:
        ws3.append([a.get("alert_id",""), a.get("entity_field",""), a.get("entity_value",""),
                    a.get("reason",""), a.get("report_name",""), a.get("created_at",""), a.get("status","")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    key = f"case-reports/{case_id}/caso_{case_id[:8]}_{dt.datetime.utcnow().strftime('%Y%m%dT%H%M%S')}.xlsx"
    s3.put_object(Bucket=S3_BUCKET, Key=key, Body=buf.getvalue(),
                  ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    url = s3.generate_presigned_url("get_object", Params={"Bucket": S3_BUCKET, "Key": key},
                                    ExpiresIn=3600)
    return resp(200, {"download_url": url, "filename": key.split("/")[-1]})


# ---------------------------------------------------------------------------
# ENTITY TIMELINE SEARCH
# ---------------------------------------------------------------------------

def search_entity_timeline(query: str, limit: int = 100):
    """Search alerts + cases for an entity value. Returns unified timeline sorted by date."""
    if not query or len(query) < 3:
        return resp(400, {"error": "query must be at least 3 characters"})
    q = query.strip().lower()
    try:
        alert_rows = []
        for a in _crm_list("alerts"):
            if q in (a.get("entity_value", "") or "").lower() or q in (a.get("reason", "") or "").lower():
                alert_rows.append({
                    "source_type": "alert", "source_id": a.get("alert_id", ""),
                    "entity_value": a.get("entity_value", ""), "detail": a.get("reason", ""),
                    "report_name": a.get("report_name", ""), "event_date": a.get("created_at", ""),
                    "status": a.get("status", ""),
                })
        case_rows = []
        for c in _crm_list("cases"):
            hay = (c.get("title", ""), c.get("entity_id", ""), c.get("description", ""))
            if any(q in (h or "").lower() for h in hay):
                case_rows.append({
                    "source_type": "case", "source_id": c.get("case_id", ""),
                    "entity_value": c.get("entity_id", ""), "detail": c.get("title", ""),
                    "report_name": c.get("report_name", ""), "event_date": c.get("created_at", ""),
                    "status": c.get("status", ""),
                })
        combined = sorted(alert_rows + case_rows,
                          key=lambda x: x.get("event_date", ""), reverse=True)
        return resp(200, {"results": combined[:int(limit)], "query": query,
                          "alert_count": len(alert_rows), "case_count": len(case_rows)})
    except Exception as e:
        return resp(200, {"results": [], "warning": str(e), "query": query})


# ---------------------------------------------------------------------------
# AUDIT LOG
# ---------------------------------------------------------------------------

def get_audit_log(limit: int = 200, entity_type: str | None = None,
                  user_email: str | None = None, action: str | None = None):
    """Read the audit log from S3 with optional filters. Most recent first."""
    try:
        rows = []
        for e in _crm_list("audit"):
            if entity_type and e.get("entity_type") != entity_type:
                continue
            if user_email and e.get("user_email") != user_email:
                continue
            if action and action.lower() not in (e.get("action", "") or "").lower():
                continue
            rows.append({
                "log_id": e.get("log_id", ""),
                "user_email": e.get("user_email", ""),
                "action": e.get("action", ""),
                "entity_type": e.get("entity_type", ""),
                "entity_id": e.get("entity_id", ""),
                "created_at": e.get("created_at", ""),
            })
        rows.sort(key=lambda x: x["created_at"], reverse=True)
        return resp(200, {"entries": rows[:int(limit)]})
    except Exception as e:
        return resp(200, {"entries": [], "warning": str(e)})


# ---------------------------------------------------------------------------
# SCHEDULER (EventBridge rules)
# ---------------------------------------------------------------------------

_events = boto3.client("events", region_name=os.environ.get("AWS_REGION", "us-east-1"))

def get_schedules():
    """List EventBridge rules tagged for this project."""
    try:
        result = _events.list_rules(NamePrefix="compliance-")
        rules = []
        for r in result.get("Rules", []):
            rules.append({
                "name":        r["Name"],
                "state":       r["State"],
                "description": r.get("Description", ""),
                "schedule":    r.get("ScheduleExpression", ""),
            })
        return resp(200, {"schedules": rules})
    except Exception as e:
        return resp(200, {"schedules": [], "warning": str(e)})


def toggle_schedule(name: str, body: dict):
    """Enable or disable an EventBridge rule."""
    action = body.get("action", "").lower()  # "enable" | "disable"
    if action not in ("enable", "disable"):
        return resp(400, {"error": "action must be 'enable' or 'disable'"})
    try:
        if action == "enable":
            _events.enable_rule(Name=name)
        else:
            _events.disable_rule(Name=name)
        return resp(200, {"message": f"Rule '{name}' {action}d"})
    except Exception as e:
        return resp(500, {"error": str(e)})


def update_schedule_expression(name: str, body: dict):
    """Update cron/rate expression of an EventBridge rule."""
    expression = body.get("schedule_expression", "").strip()
    if not expression:
        return resp(400, {"error": "schedule_expression is required"})
    try:
        existing = _events.describe_rule(Name=name)
        _events.put_rule(
            Name=name,
            ScheduleExpression=expression,
            State=existing.get("State", "ENABLED"),
            Description=existing.get("Description", ""),
        )
        return resp(200, {"message": f"Rule '{name}' expression updated to: {expression}"})
    except Exception as e:
        return resp(500, {"error": str(e)})


# ---------------------------------------------------------------------------
# Monitoreo de Clientes Institucionales
# ---------------------------------------------------------------------------
INSTITUTIONAL_SNAPSHOT_KEY = "compliance-data/institutional_clients_snapshot.json"
INSTITUTIONAL_RULES_KEY = "config/institutional_alert_rules.json"

# Orden de severidad para agrupar las mini-fichas — cubre los 6 valores
# reales de compliance_status vistos en clientes institucionales hoy.
INSTITUTIONAL_STATUS_ORDER = [
    "NORMAL", "PENDING_COMPLIANCE", "CUSTOMER_REVIEW",
    "UNDER_COMPLIANCE_REVIEW", "BLOCKED", "FULLY_BLOCKED",
]


def run_institutional_clients_refresh(body: dict | None = None):
    """Dispara el recálculo del snapshot de clientes institucionales
    (beneficiarios, monto total, transacciones). Async — igual patrón que
    wallet_search: crea un run y lo procesa el Report Runner."""
    body = body or {}
    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()
    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": "institutional_clients_snapshot",
        "status": "RUNNING",
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })
    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps({
            "report_name": "institutional_clients_snapshot",
            "run_id": run_id,
            "keep_session": False,
        }),
    )
    return resp(202, {"run_id": run_id, "status": "RUNNING"})


def get_institutional_clients():
    """Lee el snapshot ya calculado (fijo hasta el próximo 'Actualizar') —
    no toca Redshift, funciona incluso con el clúster pausado."""
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=INSTITUTIONAL_SNAPSHOT_KEY)
        snapshot = json.loads(obj["Body"].read())
    except Exception:
        return resp(200, {"computed_at": None, "clients": [], "never_computed": True})

    order = {s: i for i, s in enumerate(INSTITUTIONAL_STATUS_ORDER)}
    clients = snapshot.get("clients", [])
    clients.sort(key=lambda c: (order.get((c.get("compliance_status") or "").upper(), 99), c.get("company_name") or ""))
    return resp(200, {"computed_at": snapshot.get("computed_at"), "clients": clients})


def _load_institutional_rules() -> list[dict]:
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=INSTITUTIONAL_RULES_KEY)
        return json.loads(obj["Body"].read())
    except Exception:
        return []


def _save_institutional_rules(rules: list[dict]) -> None:
    s3.put_object(
        Bucket=S3_BUCKET, Key=INSTITUTIONAL_RULES_KEY,
        Body=json.dumps(rules, default=str).encode(), ContentType="application/json",
    )


def get_institutional_rules():
    return resp(200, {"rules": _load_institutional_rules()})


_INSTITUTIONAL_METRICS = {"monto_usd", "n_transacciones"}
_INSTITUTIONAL_WINDOWS = {"daily", "monthly", "custom_days"}
_INSTITUTIONAL_WINDOW_DAYS = {"daily": 1, "monthly": 30}
INSTITUTIONAL_LOOKBACK_DAYS = 90  # tope de la ventana custom_days


def _institutional_window_days(rule: dict) -> int:
    window = rule.get("window", "daily")
    if window == "custom_days":
        return max(1, min(INSTITUTIONAL_LOOKBACK_DAYS, int(rule.get("window_days") or 1)))
    return _INSTITUTIONAL_WINDOW_DAYS.get(window, 1)


def create_institutional_rule(body: dict):
    """body: {company_id (opcional — vacío/None = regla default/global
    para institucionales sin regla propia), company_name,
    metric: "monto_usd"|"n_transacciones",
    window: "daily"|"monthly"|"custom_days", window_days (solo si window
    es custom_days), umbral, enabled}."""
    umbral = body.get("umbral")
    if umbral is None:
        return resp(400, {"error": "umbral is required"})
    metric = body.get("metric", "monto_usd")
    if metric not in _INSTITUTIONAL_METRICS:
        return resp(400, {"error": f"metric debe ser uno de: {', '.join(sorted(_INSTITUTIONAL_METRICS))}"})
    window = body.get("window", "daily")
    if window not in _INSTITUTIONAL_WINDOWS:
        return resp(400, {"error": f"window debe ser uno de: {', '.join(sorted(_INSTITUTIONAL_WINDOWS))}"})
    company_id = body.get("company_id")
    rule = {
        "rule_id": str(uuid.uuid4()),
        "company_id": company_id if company_id not in (None, "") else None,
        "company_name": str(body.get("company_name", "")).strip(),
        "metric": metric,
        "window": window,
        "window_days": int(body.get("window_days") or 1) if window == "custom_days" else None,
        "umbral": float(umbral),
        "enabled": bool(body.get("enabled", True)),
        "created_at": _now_str(),
        "updated_at": _now_str(),
    }
    rules = _load_institutional_rules()
    rules.append(rule)
    _save_institutional_rules(rules)
    return resp(201, {"rule": rule})


def update_institutional_rule(rule_id: str, body: dict):
    rules = _load_institutional_rules()
    for r in rules:
        if r.get("rule_id") == rule_id:
            for field in ("company_id", "company_name", "metric", "window", "window_days", "umbral", "enabled"):
                if field in body:
                    r[field] = body[field]
            r["updated_at"] = _now_str()
            _save_institutional_rules(rules)
            return resp(200, {"rule": r})
    return resp(404, {"error": "Rule not found"})


def delete_institutional_rule(rule_id: str):
    rules = _load_institutional_rules()
    new_rules = [r for r in rules if r.get("rule_id") != rule_id]
    if len(new_rules) == len(rules):
        return resp(404, {"error": "Rule not found"})
    _save_institutional_rules(new_rules)
    return resp(200, {"message": "Regla eliminada"})


def apply_institutional_alert_rules(daily_rows: list[dict]) -> int:
    """daily_rows: agregados por (company_id, dia) de los últimos
    INSTITUTIONAL_LOOKBACK_DAYS días — {company_id, company_name, dia,
    monto_usd, n_transacciones}.

    Para cada cliente institucional con datos en la ventana, calcula el
    valor de SU regla asignada (o la regla default/global si no tiene una
    propia) sobre la métrica y ventana configuradas (diaria/mensual/X
    días) y crea una mini-alerta si se supera el umbral. Deduplica por
    company_id + rule_id + fecha de hoy — como mucho una alerta por
    cliente por regla por día, sin importar cuántas veces se corra el
    chequeo. Best-effort, nunca debería romper la corrida que la invoca."""
    try:
        rules = _load_institutional_rules()
        by_company = {r["company_id"]: r for r in rules if r.get("company_id") and r.get("enabled", True)}
        default_rule = next((r for r in rules if not r.get("company_id") and r.get("enabled", True)), None)
        if not by_company and not default_rule:
            return 0

        rows_by_company: dict = {}
        for row in daily_rows:
            rows_by_company.setdefault(row.get("company_id"), []).append(row)

        today = dt.date.today()
        today_str = today.isoformat()
        already_alerted = {
            (a.get("company_id"), a.get("rule_id"), a.get("fecha")) for a in _crm_list("institutional_alerts")
        }

        n_created = 0
        for company_id, rows in rows_by_company.items():
            rule = by_company.get(company_id) or default_rule
            if not rule:
                continue
            metric = rule.get("metric", "monto_usd")
            window_days = _institutional_window_days(rule)
            cutoff = today - dt.timedelta(days=window_days - 1)

            value = 0.0
            company_name = ""
            for r in rows:
                company_name = r.get("company_name") or company_name
                try:
                    dia = dt.datetime.strptime(str(r.get("dia"))[:10], "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    continue
                if dia < cutoff:
                    continue
                value += float(r.get(metric) or 0)

            umbral = float(rule.get("umbral", 0))
            if value <= umbral:
                continue
            if (company_id, rule.get("rule_id"), today_str) in already_alerted:
                continue

            aid = str(uuid.uuid4())
            _crm_put("institutional_alerts", aid, {
                "alert_id": aid,
                "company_id": company_id,
                "company_name": company_name,
                "fecha": today_str,
                "metric": metric,
                "window": rule.get("window"),
                "window_days": window_days,
                "valor": value,
                "umbral": umbral,
                "rule_id": rule.get("rule_id", ""),
                "revisado": False,
                "created_at": _now_str(),
            })
            n_created += 1
        return n_created
    except Exception:
        return 0


def run_institutional_alert_check(body: dict | None = None):
    """Dispara la revisión de umbral diario de todos los institucionales.
    Async — mismo patrón que el resto de los módulos que necesitan
    Redshift (puede tardar en encender el clúster)."""
    body = body or {}
    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()
    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": "institutional_alert_check",
        "status": "RUNNING",
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })
    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps({
            "report_name": "institutional_alert_check",
            "run_id": run_id,
            "keep_session": False,
        }),
    )
    return resp(202, {"run_id": run_id, "status": "RUNNING"})


def get_institutional_alerts():
    alerts = _crm_list("institutional_alerts")
    alerts.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    return resp(200, {"alerts": alerts})


def review_institutional_alert(alert_id: str, body: dict | None = None):
    body = body or {}
    updated = _crm_update("institutional_alerts", alert_id, {
        "revisado": True,
        "reviewed_by": str(body.get("reviewed_by", "")).strip(),
        "reviewed_at": _now_str(),
    })
    if updated is None:
        return resp(404, {"error": f"Alert '{alert_id}' not found"})
    return resp(200, {"alert": updated})


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------
# Admin — verifica capacidad de escritura en S3 (diagnóstico de permisos)
# ---------------------------------------------------------------------------
def admin_s3check():
    key = "crm/_s3check.json"
    result = {"bucket": S3_BUCKET}
    try:
        s3.put_object(Bucket=S3_BUCKET, Key=key, Body=b'{"ok":true}',
                      ContentType="application/json")
        result["write"] = True
    except Exception as e:
        result["write"] = False
        result["error"] = str(e)
        return resp(200, result)
    try:
        s3.get_object(Bucket=S3_BUCKET, Key=key)
        result["read"] = True
    except Exception as e:
        result["read"] = False
        result["read_error"] = str(e)
    try:
        s3.delete_object(Bucket=S3_BUCKET, Key=key)
        result["delete"] = True
    except Exception as e:
        result["delete"] = False
        result["delete_error"] = str(e)
    return resp(200, result)


def _str_to_epoch(ts) -> int:
    if not ts:
        return 0
    s = str(ts)[:19]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return int(dt.datetime.strptime(s, fmt).timestamp())
        except ValueError:
            continue
    return 0


def admin_migrate(body: dict):
    """One-time copy of CRM data from Redshift → S3. Runs inside the Lambda
    (has both redshift-data + S3 perms). Requires cluster ONLINE. Idempotent."""
    module = (body.get("module") or "").strip()
    migrators = {
        "whitelist": _migrate_whitelist_to_s3,
        "alerts": _migrate_alerts_to_s3,
        "cases": _migrate_cases_to_s3,
        "users": _migrate_users_to_s3,
    }
    if module == "all":
        out = {}
        for name, fn in migrators.items():
            try:
                out[name] = fn()
            except Exception as e:
                out[name] = f"error: {e}"
        return resp(200, {"migrated": out})
    fn = migrators.get(module)
    if not fn:
        return resp(400, {"error": f"unknown module '{module}'",
                          "available": list(migrators) + ["all"]})
    try:
        return resp(200, {"module": module, "migrated": fn()})
    except Exception as e:
        return resp(500, {"module": module, "error": str(e)})


def _migrate_whitelist_to_s3() -> int:
    rows = _rs_exec(
        "SELECT whitelist_id, entity_field, entity_value, duration_days, reason, scope, "
        "report_name, created_at::VARCHAR AS created_at, expires_at::VARCHAR AS expires_at "
        "FROM compliance.whitelist"
    )
    n = 0
    for r in rows:
        wid = r.get("whitelist_id")
        if not wid:
            continue
        exp_str = str(r.get("expires_at") or "")[:19]
        _crm_put("whitelist", str(wid), {
            "whitelist_id": str(wid),
            "entity_field": r.get("entity_field") or "",
            "entity_value": r.get("entity_value") or "",
            "duration_days": int(r.get("duration_days") or 0),
            "reason": r.get("reason") or "",
            "scope": r.get("scope") or "global",
            "report_name": r.get("report_name") or "",
            "created_at": str(r.get("created_at") or "")[:19],
            "expires_at": _str_to_epoch(exp_str),
            "expires_at_str": exp_str,
        })
        n += 1
    return n


def _migrate_alerts_to_s3() -> int:
    rows = _rs_exec(
        "SELECT alert_id, entity_field, entity_value, reason, report_name, row_data, "
        "created_at::VARCHAR AS created_at, status, "
        "COALESCE(reviewed_at::VARCHAR, '') AS reviewed_at, "
        "COALESCE(priority, 'medium') AS priority, "
        "COALESCE(assigned_to, '') AS assigned_to, "
        "COALESCE(reviewed_by, '') AS reviewed_by, "
        "COALESCE(notes, '') AS notes "
        "FROM compliance.alerts"
    )
    n = 0
    for r in rows:
        aid = r.get("alert_id")
        if not aid:
            continue
        _crm_put("alerts", str(aid), {
            "alert_id": str(aid),
            "entity_field": r.get("entity_field") or "",
            "entity_value": r.get("entity_value") or "",
            "reason": r.get("reason") or "",
            "report_name": r.get("report_name") or "",
            "row_data": r.get("row_data") or "",
            "created_at": str(r.get("created_at") or "")[:19],
            "status": r.get("status") or "active",
            "reviewed_at": str(r.get("reviewed_at") or "")[:19],
            "priority": r.get("priority") or "medium",
            "assigned_to": r.get("assigned_to") or "",
            "reviewed_by": r.get("reviewed_by") or "",
            "notes": r.get("notes") or "",
        })
        n += 1
    return n


def _migrate_cases_to_s3() -> int:
    cases = _rs_exec(
        "SELECT case_id, title, description, status, priority, entity_type, entity_id, "
        "report_name, COALESCE(assigned_to,'') AS assigned_to, created_by, "
        "created_at::VARCHAR AS created_at, updated_at::VARCHAR AS updated_at, "
        "COALESCE(closed_at::VARCHAR,'') AS closed_at FROM crm.cases"
    )
    notes = _rs_exec(
        "SELECT note_id, case_id, COALESCE(author_email,'') AS author_email, content, "
        "created_at::VARCHAR AS created_at FROM crm.case_notes"
    )
    notes_by_case: dict[str, list] = {}
    for nt in notes:
        notes_by_case.setdefault(str(nt.get("case_id")), []).append({
            "note_id": str(nt.get("note_id") or ""),
            "case_id": str(nt.get("case_id") or ""),
            "author_email": nt.get("author_email") or "",
            "content": nt.get("content") or "",
            "created_at": str(nt.get("created_at") or "")[:19],
        })
    n = 0
    for c in cases:
        cid = c.get("case_id")
        if not cid:
            continue
        cnotes = sorted(notes_by_case.get(str(cid), []), key=lambda x: x["created_at"])
        _crm_put("cases", str(cid), {
            "case_id": str(cid),
            "title": c.get("title") or "",
            "description": c.get("description") or "",
            "status": c.get("status") or "open",
            "priority": c.get("priority") or "medium",
            "entity_type": c.get("entity_type") or "",
            "entity_id": c.get("entity_id") or "",
            "report_name": c.get("report_name") or "",
            "assigned_to": c.get("assigned_to") or "",
            "created_by": c.get("created_by") or "",
            "created_at": str(c.get("created_at") or "")[:19],
            "updated_at": str(c.get("updated_at") or "")[:19],
            "closed_at": str(c.get("closed_at") or "")[:19],
            "notes": cnotes,
        })
        n += 1
    return n


def _migrate_users_to_s3() -> int:
    rows = _rs_exec(
        "SELECT u.email, COALESCE(u.full_name,'') AS full_name, u.is_active, "
        "COALESCE(r.name,'analyst') AS role_name, "
        "COALESCE(u.last_login_at::VARCHAR,'') AS last_login_at "
        "FROM crm.users u LEFT JOIN crm.roles r ON u.role_id = r.id"
    )
    n = 0
    for r in rows:
        email = (r.get("email") or "").strip()
        if not email:
            continue
        _crm_put("users", email, {
            "email": email,
            "full_name": r.get("full_name") or email,
            "is_active": bool(r.get("is_active", True)),
            "role_name": r.get("role_name") or "analyst",
            "last_login_at": str(r.get("last_login_at") or "")[:19],
        })
        n += 1
    return n


# ---------------------------------------------------------------------------
def handler(event, context):  # noqa: ARG001
    method = event.get("requestContext", {}).get("http", {}).get("method", "GET")
    path = event.get("rawPath", "/").rstrip("/") or "/"
    parts = [p for p in path.split("/") if p]

    try:
        body = json.loads(event.get("body") or "{}")
    except (json.JSONDecodeError, TypeError):
        body = {}

    try:
        # CORS preflight — return proper CORS headers so browser accepts the request.
        # API Gateway $default route routes OPTIONS to Lambda, so we handle CORS here.
        if method == "OPTIONS":
            req_headers = event.get("headers", {}) or {}
            origin = req_headers.get("origin") or req_headers.get("Origin", "")
            _allowed_origins = {
                "https://bmackenna-g66.github.io",
                "https://di7f123v3u2y5.cloudfront.net",
            }
            cors_origin = origin if origin in _allowed_origins else "https://bmackenna-g66.github.io"
            return {
                "statusCode": 200,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": cors_origin,
                    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                    "Access-Control-Allow-Headers": "Authorization, Content-Type",
                    "Access-Control-Max-Age": "300",
                },
                "body": "",
            }

        # GET /reports
        if method == "GET" and not parts:
            return resp(200, {"message": "Compliance Reports API"})

        if method == "GET" and parts == ["reports"]:
            return get_reports()

        # POST /execute
        if method == "POST" and parts == ["execute"]:
            return execute_report(body)

        # GET /runs
        if method == "GET" and parts == ["runs"]:
            qs = event.get("queryStringParameters") or {}
            return get_runs(qs.get("user_email", ""))

        # GET /runs/{run_id}/rows — todas las filas navegables de la corrida.
        # Va ANTES de GET /runs/{run_id} por especificidad.
        if method == "GET" and len(parts) == 3 and parts[0] == "runs" and parts[2] == "rows":
            return get_run_rows(parts[1])
        # GET /runs/{run_id}
        if method == "GET" and len(parts) == 2 and parts[0] == "runs":
            return get_run(parts[1])

        # POST /queries
        if method == "POST" and parts == ["queries"]:
            return save_query(body, get_user_email(event))

        # DELETE /queries/{name}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "queries":
            return delete_query(parts[1])

        # GET /cluster/status
        if method == "GET" and parts == ["cluster", "status"]:
            return get_cluster_status()

        # POST /analyze/individual
        if method == "POST" and parts == ["analyze", "individual"]:
            return run_individual_analysis(body)
        # POST /analyze/customer/b2c
        if method == "POST" and parts == ["analyze", "customer", "b2c"]:
            return search_customer_b2c(body)
        # POST /analyze/customer/b2b
        if method == "POST" and parts == ["analyze", "customer", "b2b"]:
            return search_customer_b2b(body)
        # POST /analyze/entity-name — resuelve el nombre desde el ID
        if method == "POST" and parts == ["analyze", "entity-name"]:
            return lookup_case_entity(body)

        # POST /search/transactions
        if method == "POST" and parts == ["search", "transactions"]:
            return run_transaction_search(body)
        # POST /remesas/search — versión EN VIVO (sincrónica) de la Búsqueda
        # de Remesas, pensada para ser llamada desde otro proyecto/sistema.
        if method == "POST" and parts == ["remesas", "search"]:
            return search_remesas_sync(body)
        # POST /search/wallet
        if method == "POST" and parts == ["search", "wallet"]:
            return run_wallet_search(body)

        # ── Institucional ────────────────────────────────────────────────
        # GET  /institutional/clients
        if method == "GET" and parts == ["institutional", "clients"]:
            return get_institutional_clients()
        # POST /institutional/clients/refresh
        if method == "POST" and parts == ["institutional", "clients", "refresh"]:
            return run_institutional_clients_refresh(body)
        # GET  /institutional/rules
        if method == "GET" and parts == ["institutional", "rules"]:
            return get_institutional_rules()
        # POST /institutional/rules
        if method == "POST" and parts == ["institutional", "rules"]:
            return create_institutional_rule(body)
        # PUT|POST /institutional/rules/{id}
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "institutional" and parts[1] == "rules":
            return update_institutional_rule(parts[2], body)
        # DELETE /institutional/rules/{id}
        if method == "DELETE" and len(parts) == 3 and parts[0] == "institutional" and parts[1] == "rules":
            return delete_institutional_rule(parts[2])
        # GET  /institutional/alerts
        if method == "GET" and parts == ["institutional", "alerts"]:
            return get_institutional_alerts()
        # POST /institutional/alerts/check
        if method == "POST" and parts == ["institutional", "alerts", "check"]:
            return run_institutional_alert_check(body)
        # POST /institutional/alerts/{id} — marca revisada (debe ir DESPUÉS
        # de /institutional/alerts/check para no interceptarla)
        if method == "POST" and len(parts) == 3 and parts[0] == "institutional" and parts[1] == "alerts":
            return review_institutional_alert(parts[2], body)

        # POST /cluster/wake
        if method == "POST" and parts == ["cluster", "wake"]:
            return wake_cluster()

        # POST /cluster/pause
        if method == "POST" and parts == ["cluster", "pause"]:
            return pause_cluster_api()

        # POST /admin/s3check — verifica si la Lambda puede escribir/leer/borrar en S3
        if method == "POST" and parts == ["admin", "s3check"]:
            return admin_s3check()

        # POST /admin/migrate — copia one-time de datos Redshift→S3 (cluster online)
        if method == "POST" and parts == ["admin", "migrate"]:
            return admin_migrate(body)

        # GET /whitelist
        if method == "GET" and parts == ["whitelist"]:
            return get_whitelist()
        # POST /whitelist
        if method == "POST" and parts == ["whitelist"]:
            return add_to_whitelist(body)
        # POST /whitelist/bulk — alta masiva desde la tabla de resultados (admin).
        # Va ANTES de cualquier ruta genérica /whitelist/{id}.
        if method == "POST" and parts == ["whitelist", "bulk"]:
            return bulk_add_to_whitelist(body)
        # DELETE /whitelist/{id}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "whitelist":
            return remove_from_whitelist(parts[1])

        # ---------------------------------------------------------------------------
        # PRIORIZACIÓN DE ALERTAS — mantenedor de documentos + corrida de prueba
        # ---------------------------------------------------------------------------
        # GET /alert-document-config
        if method == "GET" and parts == ["alert-document-config"]:
            return get_alert_document_config()
        # POST /alert-document-config
        if method == "POST" and parts == ["alert-document-config"]:
            return create_alert_document_config(body)
        # PUT /alert-document-config/{id}
        if method in ("PUT", "POST") and len(parts) == 2 and parts[0] == "alert-document-config":
            return update_alert_document_config(parts[1], body)
        # DELETE /alert-document-config/{id}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "alert-document-config":
            return delete_alert_document_config(parts[1])
        # POST /alert-prioritization/test-run
        if method == "POST" and parts == ["alert-prioritization", "test-run"]:
            return run_alert_prioritization_test(body)
        # GET /alert-prioritization/settings
        if method == "GET" and parts == ["alert-prioritization", "settings"]:
            return get_priority_queue_settings()
        # POST /alert-prioritization/settings
        if method == "POST" and parts == ["alert-prioritization", "settings"]:
            return update_priority_queue_settings(body)
        # POST /alert-prioritization/run — flujo real (no datos de prueba)
        if method == "POST" and parts == ["alert-prioritization", "run"]:
            return run_alert_prioritization_real(body)
        # POST /alert-prioritization/send-manual — boton manual, documentos a eleccion
        if method == "POST" and parts == ["alert-prioritization", "send-manual"]:
            return send_manual_document_request(body)
        # GET /email-templates — catálogo de plantillas para el selector manual
        if method == "GET" and parts == ["email-templates"]:
            return list_email_templates()
        # POST /email-templates/preview — HTML del correo tal como se enviará
        if method == "POST" and parts == ["email-templates", "preview"]:
            return preview_email_template(body)
        # POST /cases/{id}/documentos-checklist
        if method == "POST" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "documentos-checklist":
            return update_case_document_checklist(parts[1], body)

        # GET /alerts/reviewed
        if method == "GET" and parts == ["alerts", "reviewed"]:
            return get_alerts(status="reviewed")
        # GET /alerts
        if method == "GET" and parts == ["alerts"]:
            return get_alerts(status="active")
        # POST /alerts
        if method == "POST" and parts == ["alerts"]:
            return add_alert(body)
        # POST /alerts/bulk-distribute — repartir filas de un reporte como
        # alertas ya asignadas entre analistas (admin)
        if method == "POST" and parts == ["alerts", "bulk-distribute"]:
            return bulk_distribute_alerts(body)
        # PUT /alerts/{id}/review
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "alerts" and parts[2] == "review":
            return review_alert(parts[1], body)
        # PUT /alerts/{id}/assign
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "alerts" and parts[2] == "assign":
            return assign_alert(parts[1], body)
        # PUT /alerts/{id}/notes
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "alerts" and parts[2] == "notes":
            return update_alert_notes(parts[1], body)
        # DELETE /alerts/{id}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "alerts":
            return delete_alert(parts[1])

        # GET /crm/users
        if method == "GET" and parts == ["crm", "users"]:
            return get_crm_users()

        # ── Mapeo de usuarios de Slack (para poder arrobar en las alertas) ──
        if method == "GET" and parts == ["slack-users"]:
            return get_slack_users()
        if method == "POST" and parts == ["slack-users"]:
            return upsert_slack_user(body)
        if method == "DELETE" and len(parts) == 2 and parts[0] == "slack-users":
            return delete_slack_user(parts[1], body)

        # ---------------------------------------------------------------------------
        # CASES CRM
        # ---------------------------------------------------------------------------
        # GET /cases
        if method == "GET" and parts == ["cases"]:
            qs = event.get("queryStringParameters") or {}
            return get_cases(qs.get("status"), qs.get("priority"), qs.get("assigned_to"))
        # POST /cases
        if method == "POST" and parts == ["cases"]:
            return create_case(body)
        # POST /cases/bulk-assign — asignación masiva (admin).
        # OJO: tiene que ir ANTES de la ruta genérica POST /cases/{id}, que
        # también matchea 2 segmentos y se lo comería tratándolo como un id.
        if method == "POST" and parts == ["cases", "bulk-assign"]:
            return bulk_assign_cases(body)
        # GET /cases/{id}
        if method == "GET" and len(parts) == 2 and parts[0] == "cases":
            return get_case_detail(parts[1])
        # DELETE /cases/{id} — eliminar caso (admin)
        if method == "DELETE" and len(parts) == 2 and parts[0] == "cases":
            return delete_case(parts[1], body)
        # PUT /cases/{id}  (update title / description / priority)
        if method in ("PUT", "POST") and len(parts) == 2 and parts[0] == "cases":
            return update_case(parts[1], body)
        # POST /cases/{id}/take — el analista se auto-asigna el caso
        if method == "POST" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "take":
            return take_case(parts[1], body)
        # POST /cases/{id}/client-profile — ficha KYC del cliente (se cachea)
        if method == "POST" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "client-profile":
            return get_case_client_profile(parts[1], body)
        # PUT /cases/{id}/status
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "cases" and parts[2] == "status":
            return update_case_status(parts[1], body)
        # PUT /cases/{id}/assign
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "cases" and parts[2] == "assign":
            return update_case_assign(parts[1], body)
        # POST /cases/{id}/notes
        if method == "POST" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "notes":
            return add_case_note(parts[1], body)
        # POST /cases/{id}/attachments/upload-url
        if method == "POST" and len(parts) == 4 and parts[0] == "cases" and parts[2] == "attachments" and parts[3] == "upload-url":
            return get_attachment_upload_url(parts[1], body)
        # POST /cases/{id}/attachments
        if method == "POST" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "attachments":
            return add_case_attachment(parts[1], body)
        # GET /cases/{id}/attachments/{attachment_id}/download-url
        if method == "GET" and len(parts) == 5 and parts[0] == "cases" and parts[2] == "attachments" and parts[4] == "download-url":
            return get_attachment_download_url(parts[1], parts[3])
        # DELETE /cases/{id}/attachments/{attachment_id}
        if method == "DELETE" and len(parts) == 4 and parts[0] == "cases" and parts[2] == "attachments":
            return delete_case_attachment(parts[1], parts[3])
        # POST /alerts/{id}/link-case
        if method == "POST" and len(parts) == 3 and parts[0] == "alerts" and parts[2] == "link-case":
            return link_alert_to_case(parts[1], body)

        # POST /alerts/notify
        if method == "POST" and parts == ["alerts", "notify"]:
            return notify_alert(body)

        if method == "POST" and parts == ["ai", "generate"]:
            return ai_generate(body)

        # GET /customer/context?customer_id=X&days=N  — dossier CRM + transacciones para la IA
        if method == "GET" and parts == ["customer", "context"]:
            qs = event.get("queryStringParameters") or {}
            try:
                days = int(qs.get("days", 30) or 30)
            except (ValueError, TypeError):
                days = 30
            return customer_context(qs.get("customer_id", ""), days)

        # GET /cases/{id}/export
        if method == "GET" and len(parts) == 3 and parts[0] == "cases" and parts[2] == "export":
            return export_case(parts[1])

        # GET /search/entity
        if method == "GET" and parts == ["search", "entity"]:
            qs = event.get("queryStringParameters") or {}
            return search_entity_timeline(qs.get("q", ""), int(qs.get("limit", 100)))

        # GET /audit
        if method == "GET" and parts == ["audit"]:
            qs = event.get("queryStringParameters") or {}
            return get_audit_log(
                limit=int(qs.get("limit", 200)),
                entity_type=qs.get("entity_type"),
                user_email=qs.get("user_email"),
                action=qs.get("action"),
            )

        # GET /schedules
        if method == "GET" and parts == ["schedules"]:
            return get_schedules()
        # PUT /schedules/{name}/toggle
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "schedules" and parts[2] == "toggle":
            return toggle_schedule(parts[1], body)
        # PUT /schedules/{name}/expression
        if method in ("PUT", "POST") and len(parts) == 3 and parts[0] == "schedules" and parts[2] == "expression":
            return update_schedule_expression(parts[1], body)

        # GET /dashboard/stats (submit queries, returns stmt_ids)
        if method == "GET" and parts == ["dashboard", "stats"]:
            return get_dashboard_stats()
        # GET /dashboard/stats/result?q0=id&q1=id&q2=id (poll results)
        if method == "GET" and parts == ["dashboard", "stats", "result"]:
            qs = event.get("queryStringParameters") or {}
            return get_dashboard_stats_result(qs.get("q0", ""), qs.get("q1", ""), qs.get("q2", ""))

        # GET /analytics/summary — submit 5 CRM analytics queries, return stmt_ids
        if method == "GET" and parts == ["analytics", "summary"]:
            return get_analytics_summary()
        # GET /analytics/result?q0=&q1=&q2=&q3=&q4= — poll analytics results
        if method == "GET" and parts == ["analytics", "result"]:
            qs = event.get("queryStringParameters") or {}
            return get_analytics_result(
                qs.get("q0", ""), qs.get("q1", ""), qs.get("q2", ""),
                qs.get("q3", ""), qs.get("q4", ""),
            )

        # GET /analytics/sla — submit 3 SLA queries
        if method == "GET" and parts == ["analytics", "sla"]:
            return get_analytics_sla()
        # GET /analytics/sla/result?q0=&q1=&q2=
        if method == "GET" and parts == ["analytics", "sla", "result"]:
            qs = event.get("queryStringParameters") or {}
            return get_analytics_sla_result(qs.get("q0", ""), qs.get("q1", ""), qs.get("q2", ""))

        # GET /users
        if method == "GET" and parts == ["users"]:
            return get_users()
        # POST /users
        if method == "POST" and parts == ["users"]:
            return create_user(body)
        # PUT /users/{id}
        if method in ("PUT", "POST") and len(parts) == 2 and parts[0] == "users":
            return update_user(parts[1], body)
        # DELETE /users/{id}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "users":
            return deactivate_user(parts[1])

        # GET /roles
        if method == "GET" and parts == ["roles"]:
            return get_roles()

        # GET /rules
        if method == "GET" and parts == ["rules"]:
            return get_rules()
        # POST /rules
        if method == "POST" and parts == ["rules"]:
            return create_rule(body)
        # PUT /rules/{id}
        if method in ("PUT", "POST") and len(parts) == 2 and parts[0] == "rules":
            return update_rule(parts[1], body)
        # DELETE /rules/{id}
        if method == "DELETE" and len(parts) == 2 and parts[0] == "rules":
            return delete_rule(parts[1])

        return resp(404, {"error": "Not found", "path": path, "method": method})

    except Exception as e:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        return resp(500, {"error": str(e)})


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------
def get_reports():
    reports = list(BUILTIN_REPORTS)
    try:
        result = catalog_table.scan(
            FilterExpression=Attr("is_custom").eq(True)
        )
        for item in result.get("Items", []):
            item.setdefault("params", [])
            reports.append(item)
    except Exception:
        pass
    return resp(200, {"reports": reports})


def execute_report(body: dict):
    report_name = body.get("report_name", "").strip()
    if not report_name:
        return resp(400, {"error": "report_name is required"})

    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()

    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": report_name,
        "status": "RUNNING",
        "params": json.dumps({k: v for k, v in body.items() if k not in ("report_name", "user_email")}),
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })

    # Invoke report Lambda asynchronously (Event type = fire and forget)
    # Forward keep_session so the Lambda skips auto-pause when set
    payload = {**body, "run_id": run_id, "keep_session": bool(body.get("keep_session", False))}
    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps(payload),
    )

    return resp(202, {"run_id": run_id, "status": "RUNNING"})


def get_runs(user_email: str = ""):
    kwargs: dict = {
        "ProjectionExpression": (
            "run_id, report_name, #st, params, started_at, "
            "completed_at, s3_key, row_count, error_message, user_email"
        ),
        "ExpressionAttributeNames": {"#st": "status"},
    }
    if user_email:
        from boto3.dynamodb.conditions import Attr  # noqa: PLC0415
        kwargs["FilterExpression"] = Attr("user_email").eq(user_email)
    result = runs_table.scan(**kwargs)
    items = sorted(
        result.get("Items", []),
        key=lambda x: x.get("started_at", ""),
        reverse=True,
    )[:50]
    return resp(200, {"runs": items})


def get_run(run_id: str):
    result = runs_table.get_item(Key={"run_id": run_id})
    item = result.get("Item")
    if not item:
        return resp(404, {"error": "Run not found"})

    # Generate fresh presigned URL if s3_key exists
    if item.get("s3_key"):
        try:
            item["download_url"] = s3.generate_presigned_url(
                "get_object",
                Params={"Bucket": S3_BUCKET, "Key": item["s3_key"]},
                ExpiresIn=3600,
            )
        except Exception:
            pass

    # Parse result_preview / ai_summary if stored as JSON strings
    for fld, fallback in (("result_preview", []), ("ai_summary", None)):
        val = item.get(fld)
        if isinstance(val, str):
            try:
                item[fld] = json.loads(val)
            except Exception:
                item[fld] = fallback

    return resp(200, item)


def get_run_rows(run_id: str):
    """Todas las filas navegables de una corrida.

    El registro de la corrida en DynamoDB solo guarda 10 filas de muestra (por
    el límite de tamaño de item), por eso la tabla en pantalla mostraba "10 de
    N" sin forma de ver el resto. El runner deja el resultado completo en S3 y
    este endpoint lo sirve para que la tabla pueda paginarlo entero."""
    result = runs_table.get_item(Key={"run_id": run_id})
    item = result.get("Item")
    if not item:
        return resp(404, {"error": "Run not found"})

    key = item.get("rows_json_key")
    if not key:
        # Corrida vieja, anterior a este cambio: solo existe la muestra.
        preview = item.get("result_preview")
        if isinstance(preview, str):
            try:
                preview = json.loads(preview)
            except Exception:
                preview = []
        return resp(200, {
            "rows": preview or [],
            "count": len(preview or []),
            "total": int(item.get("row_count") or 0),
            "truncated": True,
            "reason": "corrida anterior a la vista completa — volvé a ejecutar el reporte para verlo todo",
        })

    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
        rows = json.loads(obj["Body"].read())
    except Exception as e:
        return resp(200, {"rows": [], "count": 0, "error": f"No se pudieron leer las filas: {e}"})

    total = int(item.get("row_count") or len(rows))
    return resp(200, {
        "rows": rows,
        "count": len(rows),
        "total": total,
        "truncated": len(rows) < total,
    })


def save_query(body: dict, created_by: str):
    report_name = body.get("report_name", "").strip().lower().replace(" ", "_")
    if not report_name or not body.get("sql", "").strip():
        return resp(400, {"error": "report_name and sql are required"})

    builtin_names = {r["report_name"] for r in BUILTIN_REPORTS}
    if report_name in builtin_names:
        return resp(400, {"error": f"'{report_name}' is a built-in report and cannot be overwritten"})

    catalog_table.put_item(Item={
        "report_name": report_name,
        "display_name": body.get("display_name", report_name).strip() or report_name,
        "description": body.get("description", "").strip(),
        "sql": body["sql"].strip(),
        "is_custom": True,
        "params": [],
        "created_at": dt.datetime.utcnow().isoformat(),
        "created_by": created_by,
    })
    return resp(201, {"report_name": report_name, "message": "Query guardada correctamente"})


def get_cluster_status():
    try:
        r = redshift.describe_clusters(ClusterIdentifier=CLUSTER_ID)
        status = r["Clusters"][0]["ClusterStatus"]
    except Exception as e:
        return resp(200, {"status": "unknown", "error": str(e)})
    return resp(200, {"status": status})


def wake_cluster():
    try:
        r = redshift.describe_clusters(ClusterIdentifier=CLUSTER_ID)
        status = r["Clusters"][0]["ClusterStatus"]
        if status == "paused":
            redshift.resume_cluster(ClusterIdentifier=CLUSTER_ID)
            return resp(200, {"status": "resuming", "message": "Cluster despertando (3-5 min)"})
        return resp(200, {"status": status, "message": "Cluster ya está disponible"})
    except Exception as e:
        return resp(500, {"error": str(e)})


def _do_pause_with_retry(max_attempts: int = 10, wait_sec: int = 15) -> None:
    """Attempt pause_cluster, retrying on transient InvalidClusterStateFault."""
    for attempt in range(max_attempts):
        try:
            redshift.pause_cluster(ClusterIdentifier=CLUSTER_ID)
            return
        except redshift.exceptions.InvalidClusterStateFault as e:
            if "operation running" in str(e).lower() and attempt < max_attempts - 1:
                time.sleep(wait_sec)
                continue
            raise


def pause_cluster_api():
    if not CLUSTER_MANUAL_PAUSE_ENABLED:
        return resp(409, {
            "status": "auto_pause_disabled",
            "error": "La pausa del cluster está bloqueada — se mantiene siempre encendido por decisión de negocio.",
        })
    try:
        r = redshift.describe_clusters(ClusterIdentifier=CLUSTER_ID)
        status = r["Clusters"][0]["ClusterStatus"]
        if status != "available":
            return resp(200, {"status": status, "message": f"Cluster en estado: {status}"})
        try:
            _do_pause_with_retry()
            return resp(200, {"status": "pausing", "message": "Cluster pausándose..."})
        except redshift.exceptions.InvalidClusterStateFault as e:
            if "backup" not in str(e).lower() and "recently available" not in str(e).lower():
                raise
            # No recent snapshot — create one, wait, then pause
            snap_id = f"watchtower-autopause-{int(time.time())}"
            redshift.create_cluster_snapshot(
                SnapshotIdentifier=snap_id,
                ClusterIdentifier=CLUSTER_ID,
            )
            # Wait for snapshot to be available (up to 5 min)
            for _ in range(60):
                time.sleep(5)
                s = redshift.describe_cluster_snapshots(SnapshotIdentifier=snap_id)
                if s["Snapshots"][0]["Status"] == "available":
                    break
            # Extra buffer — Redshift needs time after snapshot before accepting pause
            time.sleep(30)
            _do_pause_with_retry(max_attempts=12, wait_sec=15)
            return resp(200, {"status": "pausing", "message": "Snapshot creado y cluster pausándose..."})
    except Exception as e:
        return resp(500, {"error": str(e)})


def delete_query(report_name: str):
    builtin_names = {r["report_name"] for r in BUILTIN_REPORTS}
    if report_name in builtin_names:
        return resp(400, {"error": "No se pueden eliminar los reportes predefinidos"})
    catalog_table.delete_item(Key={"report_name": report_name})
    return resp(200, {"message": f"Query '{report_name}' eliminada"})


# ---------------------------------------------------------------------------
# S3 JSON store — datos operativos del CRM (always-on, sin depender de Redshift)
# Un objeto por registro: s3://<bucket>/crm/<kind>/<id>.json
# Reutilizable por whitelist, alertados, casos, usuarios, audit.
# ---------------------------------------------------------------------------
CRM_PREFIX = "crm"


def _crm_key(kind: str, item_id: str) -> str:
    return f"{CRM_PREFIX}/{kind}/{item_id}.json"


def _crm_put(kind: str, item_id: str, item: dict) -> None:
    s3.put_object(
        Bucket=S3_BUCKET, Key=_crm_key(kind, item_id),
        Body=json.dumps(item, default=str).encode("utf-8"),
        ContentType="application/json",
    )


def _crm_get(kind: str, item_id: str) -> dict | None:
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=_crm_key(kind, item_id))
        return json.loads(obj["Body"].read())
    except s3.exceptions.NoSuchKey:
        return None


def _crm_delete(kind: str, item_id: str) -> None:
    s3.delete_object(Bucket=S3_BUCKET, Key=_crm_key(kind, item_id))


def _crm_update(kind: str, item_id: str, changes: dict) -> dict | None:
    """Read-modify-write a single record. Returns the updated item, or None if
    it doesn't exist."""
    item = _crm_get(kind, item_id)
    if item is None:
        return None
    item.update(changes)
    _crm_put(kind, item_id, item)
    return item


def _safe_audit(*, user_email="unknown", action="", entity_type="",
                entity_id="", new_value=None, **_extra) -> None:
    """Best-effort audit write to the S3 store (always-on). Never breaks the
    calling operation if it fails."""
    try:
        aid = str(uuid.uuid4())
        _crm_put("audit", aid, {
            "log_id": aid,
            "user_email": user_email or "unknown",
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "new_value": new_value,
            "created_at": _now_str(),
        })
    except Exception:
        pass


def _crm_list(kind: str) -> list[dict]:
    """List all records of a kind. Lists keys then fetches each object in
    parallel (fine for the operational volumes here)."""
    prefix = f"{CRM_PREFIX}/{kind}/"
    keys: list[str] = []
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix):
        for o in page.get("Contents", []):
            if o["Key"].endswith(".json"):
                keys.append(o["Key"])
    if not keys:
        return []

    def _fetch(k):
        try:
            return json.loads(s3.get_object(Bucket=S3_BUCKET, Key=k)["Body"].read())
        except Exception:
            return None

    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=16) as ex:
        return [i for i in ex.map(_fetch, keys) if i is not None]


def get_whitelist():
    # S3-backed: works with the Redshift cluster paused.
    try:
        now = int(time.time())
        out = []
        for i in _crm_list("whitelist"):
            exp = int(i.get("expires_at", 0))
            if exp and exp <= now:
                continue  # vencida
            out.append({
                "whitelist_id": i.get("whitelist_id", ""),
                "entity_field": i.get("entity_field", ""),
                "entity_value": i.get("entity_value", ""),
                "duration_days": int(i.get("duration_days", 0)),
                "reason": i.get("reason", ""),
                "scope": i.get("scope", "global"),
                "report_name": i.get("report_name", ""),
                "created_at": i.get("created_at", ""),
                "expires_at": i.get("expires_at_str", ""),
            })
        out.sort(key=lambda x: x["created_at"], reverse=True)
        return resp(200, {"whitelist": out})
    except Exception as e:
        return resp(200, {"whitelist": [], "warning": str(e)})


def add_to_whitelist(body: dict):
    entity_field = body.get("entity_field", "").strip()
    entity_value = body.get("entity_value", "").strip()
    duration_days = int(body.get("duration_days", 30))
    reason = body.get("reason", "").strip()
    scope = body.get("scope", "global").strip()
    report_name = body.get("report_name", "").strip()

    if not entity_field or not entity_value:
        return resp(400, {"error": "entity_field and entity_value are required"})
    if duration_days not in (30, 60, 90):
        return resp(400, {"error": "duration_days must be 30, 60, or 90"})

    wid = str(uuid.uuid4())
    now = dt.datetime.utcnow()
    expires = now + dt.timedelta(days=duration_days)
    _crm_put("whitelist", wid, {
        "whitelist_id": wid,
        "entity_field": entity_field,
        "entity_value": entity_value,
        "duration_days": duration_days,
        "reason": reason,
        "scope": scope,
        "report_name": report_name if scope == "report" else "",
        "created_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "expires_at": int(expires.timestamp()),
        "expires_at_str": expires.strftime("%Y-%m-%d %H:%M:%S"),
    })
    return resp(201, {"whitelist_id": wid})


def bulk_add_to_whitelist(body: dict):
    """Alta masiva a whitelist desde la tabla de resultados de un reporte.

    Mismo criterio que el reparto de alertas: se seleccionan filas y se
    mandan todas juntas. Es admin-only porque poner clientes en whitelist
    silencia sus alertas — es una acción de control, no operativa.

    Salta las filas que no identifican a nadie y las que YA están en
    whitelist vigente, así se puede correr dos veces sin duplicar.

    body: {rows: [...], duration_days, reason, scope, report_name,
           actor_email}
    """
    denied = _require_admin(body)
    if denied:
        return denied

    rows = body.get("rows") or []
    if not rows:
        return resp(400, {"error": "rows es requerido (lista de al menos 1)"})
    if len(rows) > _BULK_ASSIGN_MAX:
        return resp(400, {"error": f"Máximo {_BULK_ASSIGN_MAX} filas por operación"})

    try:
        duration_days = int(body.get("duration_days", 30))
    except (TypeError, ValueError):
        duration_days = 30
    if duration_days not in (30, 60, 90):
        return resp(400, {"error": "duration_days debe ser 30, 60 o 90"})

    actor = (body.get("actor_email") or "").strip()
    reason = (body.get("reason") or "").strip()
    scope = (body.get("scope") or "global").strip()
    report_name = (body.get("report_name") or "").strip()
    if not reason:
        return resp(400, {"error": "reason es requerido — queda como justificación de por qué se silencian estas alertas"})

    # Whitelist vigente, para no duplicar entradas al repetir la operación.
    now = dt.datetime.utcnow()
    now_ts = int(now.timestamp())
    ya_vigentes = set()
    try:
        for w in _crm_list("whitelist"):
            if int(w.get("expires_at") or 0) > now_ts:
                ya_vigentes.add(((w.get("entity_field") or ""), str(w.get("entity_value") or "")))
    except Exception:
        pass

    expires = now + dt.timedelta(days=duration_days)
    added, skipped = [], []
    vistos = set()
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            skipped.append({"index": idx, "motivo": "fila inválida"})
            continue
        field, value = _extract_entity_from_row(row)
        if not value:
            skipped.append({"index": idx, "motivo": "la fila no identifica cliente/empresa"})
            continue
        clave = (field, value)
        if clave in ya_vigentes:
            skipped.append({"index": idx, "motivo": f"{field} {value} ya está en whitelist vigente"})
            continue
        if clave in vistos:
            skipped.append({"index": idx, "motivo": f"{field} {value} repetido en la misma selección"})
            continue
        vistos.add(clave)

        wid = str(uuid.uuid4())
        _crm_put("whitelist", wid, {
            "whitelist_id": wid,
            "entity_field": field,
            "entity_value": value,
            "duration_days": duration_days,
            "reason": reason,
            "scope": scope,
            "report_name": report_name if scope == "report" else "",
            "created_at": now.strftime("%Y-%m-%d %H:%M:%S"),
            "expires_at": int(expires.timestamp()),
            "expires_at_str": expires.strftime("%Y-%m-%d %H:%M:%S"),
        })
        added.append({"whitelist_id": wid, "entity_field": field, "entity_value": value})

    _safe_audit(user_email=actor or "unknown", action="whitelist.bulk_add",
                entity_type="whitelist", entity_id=f"{len(added)} entradas",
                new_value={"added": len(added), "skipped": len(skipped),
                           "duration_days": duration_days, "scope": scope,
                           "report_name": report_name, "reason": reason})

    return resp(200, {
        "added": len(added),
        "skipped": skipped,
        "expires_at": expires.strftime("%Y-%m-%d %H:%M:%S"),
        "detalle": added,
    })


def remove_from_whitelist(whitelist_id: str):
    _crm_delete("whitelist", whitelist_id)
    return resp(200, {"message": f"Whitelist entry '{whitelist_id}' removed"})


# ---------------------------------------------------------------------------
# Priorización de Alertas — mantenedor de documentos a solicitar por alerta
# (un solo JSON en S3, mismo patrón que auto_case_rules) + corrida de prueba
# end-to-end: evaluación → prioridad → email (Gmail SMTP) → caso sin asignar.
# ---------------------------------------------------------------------------
def _load_alert_document_config() -> list[dict]:
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=ALERT_DOCS_CONFIG_KEY)
        return json.loads(obj["Body"].read())
    except Exception:
        return []


def _save_alert_document_config(records: list[dict]) -> None:
    s3.put_object(
        Bucket=S3_BUCKET, Key=ALERT_DOCS_CONFIG_KEY,
        Body=json.dumps(records, ensure_ascii=False, default=str).encode("utf-8"),
        ContentType="application/json",
    )


def get_alert_document_config():
    return resp(200, {"config": _load_alert_document_config()})


def create_alert_document_config(body: dict):
    alerta = str(body.get("alerta", "")).strip()
    if not alerta:
        return resp(400, {"error": "alerta is required"})
    records = _load_alert_document_config()
    entry = {
        "config_id": str(uuid.uuid4()),
        "tipo_alerta": str(body.get("tipo_alerta", "")).strip(),
        "alerta": alerta,
        "documentos_b2b": body.get("documentos_b2b") or [],
        "documentos_b2c": body.get("documentos_b2c") or [],
    }
    records.append(entry)
    _save_alert_document_config(records)
    return resp(201, {"config": entry})


def update_alert_document_config(config_id: str, body: dict):
    records = _load_alert_document_config()
    for r in records:
        if r.get("config_id") == config_id:
            for field in ("tipo_alerta", "alerta", "documentos_b2b", "documentos_b2c"):
                if field in body:
                    r[field] = body[field]
            _save_alert_document_config(records)
            return resp(200, {"config": r})
    return resp(404, {"error": "Config not found"})


def delete_alert_document_config(config_id: str):
    records = _load_alert_document_config()
    new_records = [r for r in records if r.get("config_id") != config_id]
    if len(new_records) == len(records):
        return resp(404, {"error": "Config not found"})
    _save_alert_document_config(new_records)
    return resp(200, {"message": "Config eliminada"})


def _load_priority_queue_settings() -> dict:
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=PRIORITY_QUEUE_SETTINGS_KEY)
        return json.loads(obj["Body"].read())
    except Exception:
        return {"enabled": False, "updated_at": "", "updated_by": ""}


def get_priority_queue_settings():
    return resp(200, _load_priority_queue_settings())


def update_priority_queue_settings(body: dict):
    enabled = bool(body.get("enabled", False))
    settings = {
        "enabled": enabled,
        "updated_at": _now_str(),
        "updated_by": body.get("updated_by", "").strip(),
    }
    s3.put_object(
        Bucket=S3_BUCKET, Key=PRIORITY_QUEUE_SETTINGS_KEY,
        Body=json.dumps(settings, ensure_ascii=False).encode("utf-8"),
        ContentType="application/json",
    )
    _safe_audit(user_email=settings["updated_by"] or "unknown",
                action="priority_queue.toggle", entity_type="config",
                entity_id="priority_queue_settings", new_value={"enabled": enabled})
    return resp(200, settings)


def _register_email_ref(case_id: str) -> str:
    """Genera una referencia corta (8 hex) para un caso y la guarda como un
    índice separado (crm/email_refs/{ref}.json -> {case_id}) para que
    poll_document_replies pueda encontrar el caso a partir del asunto de una
    respuesta, sin tener que escanear todos los casos."""
    ref = case_id.replace("-", "")[:8]
    _crm_put("email_refs", ref, {"ref": ref, "case_id": case_id, "created_at": _now_str()})
    _crm_update("cases", case_id, {"email_ref": ref})
    return ref


def _subject_with_ref(base_subject: str, ref: str) -> str:
    return f"{base_subject} [ref: {ref}]" if ref else base_subject


def run_alert_prioritization_test(body: dict):
    """Prueba de concepto end-to-end: lee compliance.alert_priority_test_data
    (datos ficticios cargados a mano, con prioridad ya asignada) y por cada
    fila compone el correo de solicitud de documentos (modo prueba = TODAS las
    categorías, según lo pedido), lo manda vía _send_email (Gmail SMTP — hoy
    no-opea en silencio porque GMAIL_APP_PASSWORD no está configurado todavía,
    apenas se configure empieza a mandar de verdad sin tocar este código) y
    crea un caso automático sin asignar. Deja un registro de auditoría de cada
    solicitud de documentos en S3 (crm/document_requests/) independientemente
    de si el correo realmente salió o no.

    Respeta el interruptor general (config/priority_queue_settings.json) —
    si está apagado, no manda nada ni crea casos.
    """
    settings = _load_priority_queue_settings()
    if not settings.get("enabled"):
        return resp(409, {
            "error": "El proceso de priorización de alertas está apagado. "
                     "Prendelo desde Admin antes de correr la prueba.",
            "enabled": False,
        })
    try:
        rows = _rs_exec(
            "SELECT customer_id, total_payins_7d, total_payin_usd_7d, avg_payin_usd_7d, "
            "total_payouts_7d, total_payout_usd_7d, avg_payout_usd_7d, "
            "last_payin_date::VARCHAR AS last_payin_date, last_payout_date::VARCHAR AS last_payout_date, "
            "payout_vs_payin_ratio, nombre_completo, correo, prioridad, concepto "
            f"FROM {PRIORITY_TEST_TABLE} ORDER BY customer_id"
        )
    except Exception as e:
        return resp(500, {"error": f"No se pudo leer la tabla de prueba: {e}"})

    results = []
    for r in rows:
        customer_id = r.get("customer_id")
        nombre = r.get("nombre_completo") or ""
        correo = r.get("correo") or ""
        prioridad = str(r.get("prioridad") or "P3").strip().upper()
        concepto = r.get("concepto") or ""
        case_priority = _PRIORITY_TO_CASE_PRIORITY.get(prioridad, "low")

        # Modo prueba: se piden TODOS los documentos (los datos son ficticios,
        # "concepto" no es un nombre real de alerta del mantenedor). El cuerpo
        # del correo es la plantilla oficial de Global66 (solo se reemplaza el
        # nombre); la lista de documentos queda fija en la plantilla misma.
        documentos = _ALL_DOC_CATEGORIES

        # El caso se crea ANTES de mandar el correo para poder incluir su
        # referencia corta en el asunto — así una respuesta del cliente se
        # puede vincular de vuelta a este caso (ver poll_document_replies).
        case_resp = create_case({
            "title": f"[{prioridad}] {concepto} — cliente {customer_id}",
            "description": (
                f"Caso generado automáticamente por priorización de alertas (PRUEBA).\n"
                f"Cliente: {nombre} ({customer_id})\n"
                f"Pay-ins 7d: {r.get('total_payins_7d')} (USD {r.get('total_payin_usd_7d')})\n"
                f"Pay-outs 7d: {r.get('total_payouts_7d')} (USD {r.get('total_payout_usd_7d')})\n"
                f"Ratio payout/payin: {r.get('payout_vs_payin_ratio')}\n"
                f"Último pay-in: {r.get('last_payin_date')} | Último pay-out: {r.get('last_payout_date')}\n"
                f"Documentos solicitados: {', '.join(documentos)}"
            ),
            "priority": case_priority,
            "entity_type": "customer_test",
            "entity_id": str(customer_id),
            "report_name": "alert_prioritization_test",
            "alert_data": r,
            "alert_priority": prioridad,
            "assigned_to": "",
            "created_by": "alert_prioritization_test",
        })
        case_body = json.loads(case_resp["body"])
        case_id = case_body.get("case_id", "")
        if case_id:
            _crm_update("cases", case_id, {
                "documentos_checklist": [{"categoria": d, "estado": "pendiente"} for d in documentos],
            })

        ref = _register_email_ref(case_id) if case_id else ""
        subject = _subject_with_ref("Solicitud de información adicional — Global66", ref)
        # Los datos de prueba no traen país -> siempre plantilla B2C general.
        template_key = "general_b2c"
        html_body = _render_email_template(template_key, nombre, documentos=documentos)
        attachment = _email_template_attachment(template_key)
        envio = _send_email(correo, subject, html_body, from_addr=ALERT_DOCS_FROM_ADDR,
                            attachments=[attachment] if attachment else None)

        req_id = str(uuid.uuid4())
        _crm_put("document_requests", req_id, {
            "request_id": req_id,
            "customer_id": customer_id,
            "case_id": case_id,
            "correo": correo,
            "nombre_completo": nombre,
            "prioridad": prioridad,
            "concepto": concepto,
            "documentos_solicitados": documentos,
            "template_key": template_key,
            "subject": subject,
            "sent": envio["sent"],
            "send_error": envio["error"],
            "created_at": _now_str(),
            "test_mode": True,
        })

        results.append({
            "customer_id": customer_id,
            "prioridad": prioridad,
            "concepto": concepto,
            "case_id": case_id,
            "email_sent": envio["sent"],
            "email_error": envio["error"],
            "email_to": correo,
            "documentos_solicitados": documentos,
        })

    return resp(200, {"processed": len(results), "results": results})


_PRIORITY_QUEUE_VIEW = {"customer": "compliance.priority_queue_b2c", "company": "compliance.priority_queue_b2b"}


def _score_to_priority(score) -> str:
    """UMBRALES PLACEHOLDER — pendientes de confirmar con compliance, junto
    con los pesos reales de 'Riesgo Analizado'. Fáciles de ajustar acá."""
    try:
        score = float(score)
    except (TypeError, ValueError):
        return "P3"
    if score >= 75:
        return "P1"
    if score >= 50:
        return "P2"
    return "P3"


def _lookup_alert_documents(alerta: str, entity_type: str) -> list[str]:
    """Busca en el mantenedor (config/alert_document_config.json) los
    documentos a pedir para esta alerta + tipo de cliente. Si la alerta no
    está configurada, cae de vuelta a pedir todas las categorías (mismo
    comportamiento que el modo prueba)."""
    records = _load_alert_document_config()
    alerta_norm = (alerta or "").strip().lower()
    field = "documentos_b2b" if entity_type == "company" else "documentos_b2c"
    for r in records:
        if (r.get("alerta") or "").strip().lower() == alerta_norm:
            docs = r.get(field) or []
            return docs if docs else _ALL_DOC_CATEGORIES
    return _ALL_DOC_CATEGORIES


def run_alert_prioritization_real(body: dict):
    """Flujo real: recibe alertas ya gatilladas (entity_type, entity_id,
    alerta, concepto) — típicamente el resultado de correr uno de los 29
    reportes — y por cada una:
      1. Calcula prioridad real desde compliance.priority_queue_b2c/b2b
         (score PLACEHOLDER: promedio simple de los 4 componentes, pendiente
         de los pesos reales de 'Riesgo Analizado' — ver Notas de la matriz).
      2. Busca en el mantenedor qué documentos pedir para esa alerta + tipo
         de cliente.
      3. Manda el correo (plantilla completa — el recorte dinámico del HTML
         por categoría queda pendiente de confirmar el mapeo) y crea un caso
         automático sin asignar.
    Respeta el interruptor general, igual que el modo prueba.

    body: {"alerts": [{"entity_type": "customer"|"company", "entity_id": 123,
                        "alerta": "Transacciones a Países Alto Riesgo",
                        "concepto": "texto libre opcional"}]}
    """
    settings = _load_priority_queue_settings()
    if not settings.get("enabled"):
        return resp(409, {
            "error": "El proceso de priorización de alertas está apagado. "
                     "Prendelo desde Admin antes de correrlo.",
            "enabled": False,
        })

    alerts_in = body.get("alerts") or []
    if not alerts_in:
        return resp(400, {"error": "alerts es requerido (lista de {entity_type, entity_id, alerta})"})

    results = []
    for item in alerts_in:
        entity_type = (item.get("entity_type") or "customer").strip().lower()
        if entity_type not in ("customer", "company"):
            entity_type = "customer"
        entity_id = item.get("entity_id")
        alerta = item.get("alerta", "")
        concepto = item.get("concepto") or alerta

        view = _PRIORITY_QUEUE_VIEW[entity_type]
        id_col = "customer_id" if entity_type == "customer" else "company_id"
        try:
            rows = _rs_exec(f"SELECT * FROM {view} WHERE {id_col} = {int(entity_id)}")
        except Exception as e:
            results.append({"entity_type": entity_type, "entity_id": entity_id, "error": f"No se pudo calcular prioridad: {e}"})
            continue
        if not rows:
            results.append({"entity_type": entity_type, "entity_id": entity_id, "error": "Cliente/empresa no encontrado en la vista de priorización"})
            continue
        row = rows[0]

        score = row.get("risk_score")
        prioridad = _score_to_priority(score)
        case_priority = _PRIORITY_TO_CASE_PRIORITY.get(prioridad, "low")

        if entity_type == "customer":
            nombre = f"{row.get('name','') or ''} {row.get('last_name','') or ''}".strip()
            correo = row.get("email") or ""
            # B2C: la plantilla se elige sola según el país de origen del
            # cliente (compliance.priority_queue_b2c.country_code).
            template_key = _pick_b2c_template_key(row.get("country_code"))
        else:
            nombre = row.get("rep_name") or row.get("company_name") or ""
            correo = row.get("rep_email") or ""
            # B2B: una sola plantilla genérica, no varía por país.
            template_key = "b2b_generico"

        documentos = _lookup_alert_documents(alerta, entity_type)

        case_resp = create_case({
            "title": f"[{prioridad}] {alerta} — {'cliente' if entity_type=='customer' else 'empresa'} {entity_id}",
            "description": (
                f"Caso generado automáticamente por priorización de alertas.\n"
                f"{'Cliente' if entity_type=='customer' else 'Empresa'}: {nombre} ({entity_id})\n"
                f"Alerta: {alerta}\n"
                f"Score de riesgo: {score} (PLACEHOLDER — pendiente pesos reales de 'Riesgo Analizado')\n"
                f"Documentos solicitados: {', '.join(documentos)}"
            ),
            "priority": case_priority,
            "entity_type": entity_type,
            "entity_id": str(entity_id),
            "report_name": alerta,
            "alert_data": row,
            "alert_priority": prioridad,
            "assigned_to": "",
            "created_by": "alert_prioritization_real",
        })
        case_body = json.loads(case_resp["body"])
        case_id = case_body.get("case_id", "")
        if case_id:
            _crm_update("cases", case_id, {
                "documentos_checklist": [{"categoria": d, "estado": "pendiente"} for d in documentos],
            })

        ref = _register_email_ref(case_id) if case_id else ""
        subject = _subject_with_ref("Solicitud de información adicional — Global66", ref)
        html_body = _render_email_template(template_key, nombre, documentos=documentos)
        attachment = _email_template_attachment(template_key)
        envio = _send_email(correo, subject, html_body, from_addr=ALERT_DOCS_FROM_ADDR,
                            attachments=[attachment] if attachment else None)

        req_id = str(uuid.uuid4())
        _crm_put("document_requests", req_id, {
            "request_id": req_id,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "template_key": template_key,
            "case_id": case_id,
            "correo": correo,
            "nombre_completo": nombre,
            "prioridad": prioridad,
            "risk_score": score,
            "alerta": alerta,
            "concepto": concepto,
            "documentos_solicitados": documentos,
            "subject": subject,
            "sent": envio["sent"],
            "send_error": envio["error"],
            "created_at": _now_str(),
            "test_mode": False,
        })

        results.append({
            "entity_type": entity_type,
            "entity_id": entity_id,
            "alerta": alerta,
            "risk_score": score,
            "prioridad": prioridad,
            "case_id": case_id,
            "email_sent": envio["sent"],
            "email_error": envio["error"],
            "email_to": correo,
            "documentos_solicitados": documentos,
        })

    return resp(200, {"processed": len(results), "results": results})


_REPORT_DISPLAY_NAMES = {r["report_name"]: r["display_name"] for r in BUILTIN_REPORTS}


def maybe_trigger_auto_document_requests(report_name: str, rows: list[dict]) -> dict:
    """Cierra el ciclo de un reporte programado: si el interruptor maestro
    de Priorización está prendido (Admin > Priorización, apagado por
    defecto), dispara automáticamente el flujo real (prioridad + búsqueda
    de documentos en el mantenedor + correo + caso) para las filas de
    prioridad P1 de esta corrida — P2/P3 quedan con su prioridad visible en
    el Excel/tabla pero requieren acción manual del analista (decisión de
    producto: el modelo de scoring todavía usa pesos placeholder).

    Best-effort — nunca debe romper la ejecución del reporte que lo llama."""
    try:
        settings = _load_priority_queue_settings()
        if not settings.get("enabled"):
            return {"triggered": 0, "reason": "priorización automática apagada"}
        if not rows:
            return {"triggered": 0, "reason": "sin filas"}

        alertas_in = []
        for r in rows:
            if r.get("prioridad") != "P1":
                continue
            if r.get("company_id") is not None:
                entity_type, entity_id = "company", r["company_id"]
            elif r.get("customer_id") is not None:
                entity_type, entity_id = "customer", r["customer_id"]
            else:
                continue
            alertas_in.append({
                "entity_type": entity_type,
                "entity_id": entity_id,
                "alerta": _REPORT_DISPLAY_NAMES.get(report_name, report_name),
            })
        if not alertas_in:
            return {"triggered": 0, "reason": "sin filas P1"}

        result = run_alert_prioritization_real({"alerts": alertas_in})
        body = json.loads(result["body"])
        return {"triggered": body.get("processed", 0)}
    except Exception as e:
        return {"triggered": 0, "error": str(e)}


def send_manual_document_request(body: dict):
    """Botón manual — mismo correo/caso que el flujo automático, pero:
      - no depende del interruptor general (es una acción deliberada de un
        analista, uno a la vez, no el proceso masivo automático)
      - los documentos a pedir los elige el analista a mano (no el
        mantenedor), útil para casos puntuales fuera de lo estándar.

    body: {entity_type, entity_id, nombre, correo, prioridad, alerta,
           documentos: [...], case_id (opcional, para linkear a un caso
           existente en vez de crear uno nuevo), template_key (opcional —
           uno de EMAIL_TEMPLATE_CATALOG; si no se manda, se elige sola
           igual que en el flujo automático), texto_libre (solo si
           template_key='texto_libre')}
    """
    entity_type = (body.get("entity_type") or "customer").strip().lower()
    entity_id = body.get("entity_id", "")
    nombre = body.get("nombre", "").strip()
    correo = body.get("correo", "").strip()
    prioridad = (body.get("prioridad") or "P3").strip().upper()
    alerta = body.get("alerta", "").strip()
    documentos = body.get("documentos") or []
    existing_case_id = body.get("case_id", "").strip()
    template_key = (body.get("template_key") or "").strip()
    texto_libre = body.get("texto_libre", "")

    if not correo:
        return resp(400, {"error": "correo is required"})
    if not documentos:
        return resp(400, {"error": "documentos (lista, al menos 1) is required"})
    if not template_key:
        template_key = "b2b_generico" if entity_type == "company" else "general_b2c"
    if template_key not in EMAIL_TEMPLATE_CATALOG:
        return resp(400, {"error": f"template_key desconocido: {template_key}"})
    if EMAIL_TEMPLATE_CATALOG[template_key]["requires_custom_text"] and not texto_libre.strip():
        return resp(400, {"error": "texto_libre es requerido para la plantilla de texto libre"})

    checklist = [{"categoria": d, "estado": "pendiente"} for d in documentos]

    # El caso se crea/actualiza ANTES de mandar el correo para poder incluir
    # su referencia corta en el asunto (ver _register_email_ref).
    if existing_case_id:
        updated = _crm_update("cases", existing_case_id, {"documentos_checklist": checklist})
        case_id = existing_case_id if updated else ""
        ref = (updated or {}).get("email_ref") or (_register_email_ref(case_id) if case_id else "")
    else:
        case_priority = _PRIORITY_TO_CASE_PRIORITY.get(prioridad, "low")
        case_resp = create_case({
            "title": f"[{prioridad}] {alerta or 'Solicitud manual'} — {'cliente' if entity_type=='customer' else 'empresa'} {entity_id}",
            "description": (
                f"Caso generado por solicitud manual de documentos.\n"
                f"{'Cliente' if entity_type=='customer' else 'Empresa'}: {nombre} ({entity_id})\n"
                f"Alerta: {alerta or '(sin alerta asociada)'}\n"
                f"Documentos solicitados: {', '.join(documentos)}"
            ),
            "priority": case_priority,
            "entity_type": entity_type,
            "entity_id": str(entity_id),
            "report_name": alerta,
            "alert_data": body.get("alert_data") or {},
            "alert_priority": prioridad,
            "assigned_to": "",
            "created_by": "manual_document_request",
        })
        case_body = json.loads(case_resp["body"])
        case_id = case_body.get("case_id", "")
        if case_id:
            _crm_update("cases", case_id, {"documentos_checklist": checklist})
        ref = _register_email_ref(case_id) if case_id else ""

    subject = _subject_with_ref("Solicitud de información adicional — Global66", ref)
    html_body = _render_email_template(template_key, nombre, texto_libre, documentos)
    attachment = _email_template_attachment(template_key)
    envio = _send_email(correo, subject, html_body, from_addr=ALERT_DOCS_FROM_ADDR,
                        attachments=[attachment] if attachment else None)

    req_id = str(uuid.uuid4())
    _crm_put("document_requests", req_id, {
        "request_id": req_id,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "case_id": case_id,
        "correo": correo,
        "nombre_completo": nombre,
        "prioridad": prioridad,
        "alerta": alerta,
        "documentos_solicitados": documentos,
        "template_key": template_key,
        "subject": subject,
        "sent": envio["sent"],
        "send_error": envio["error"],
        "created_at": _now_str(),
        "manual": True,
    })

    return resp(200, {
        "case_id": case_id,
        "email_sent": envio["sent"],
        "email_error": envio["error"],
        "email_to": correo,
        "documentos_solicitados": documentos,
        "template_key": template_key,
    })


_CHECKLIST_ESTADOS = {"pendiente", "recibido", "entregado"}


def update_case_document_checklist(case_id: str, body: dict):
    """Actualiza el estado de un documento del checklist.
    body: {categoria, estado: "pendiente"|"recibido"|"entregado"}.

    "recibido" es un estado intermedio — lo pone el poller de correos cuando
    detecta una respuesta con adjuntos, antes de que un analista valide el
    contenido y lo pase a mano a "entregado" (ver poll_document_replies)."""
    categoria = body.get("categoria", "").strip()
    estado = (body.get("estado") or "").strip().lower()
    if not categoria:
        return resp(400, {"error": "categoria is required"})
    if estado not in _CHECKLIST_ESTADOS:
        return resp(400, {"error": f"estado debe ser uno de: {', '.join(sorted(_CHECKLIST_ESTADOS))}"})

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    checklist = case.get("documentos_checklist") or []
    found = False
    for item in checklist:
        if item.get("categoria") == categoria:
            item["estado"] = estado
            found = True
            break
    if not found:
        checklist.append({"categoria": categoria, "estado": estado})

    _crm_update("cases", case_id, {"documentos_checklist": checklist})
    return resp(200, {"documentos_checklist": checklist})


# ---------------------------------------------------------------------------
# ALERTS (Alertados / Ya Revisados)
# ---------------------------------------------------------------------------

_PRIORITY_RANK = {"high": 1, "medium": 2, "low": 3}


def get_alerts(status: str = "active"):
    # S3-backed: works with the Redshift cluster paused.
    try:
        out = []
        for i in _crm_list("alerts"):
            if i.get("status", "active") != status:
                continue
            out.append({
                "alert_id": i.get("alert_id", ""),
                "entity_field": i.get("entity_field", ""),
                "entity_value": i.get("entity_value", ""),
                "reason": i.get("reason", ""),
                "report_name": i.get("report_name", ""),
                "row_data": i.get("row_data", ""),
                "created_at": i.get("created_at", ""),
                "status": i.get("status", "active"),
                "reviewed_at": i.get("reviewed_at", ""),
                "priority": i.get("priority", "medium"),
                "assigned_to": i.get("assigned_to", ""),
                "reviewed_by": i.get("reviewed_by", ""),
                "notes": i.get("notes", ""),
                # Permite marcar en la tabla del reporte qué filas ya tienen
                # caso abierto, además de quién las tiene asignadas.
                "case_id": i.get("case_id", ""),
            })
        # Stable sort: first by created_at DESC, then by priority → within a
        # priority, newest first (matches the old SQL ORDER BY).
        out.sort(key=lambda a: a["created_at"], reverse=True)
        out.sort(key=lambda a: _PRIORITY_RANK.get(a["priority"], 2))
        return resp(200, {"alerts": out})
    except Exception as e:
        return resp(200, {"alerts": [], "warning": str(e)})


def add_alert(body: dict):
    entity_field = body.get("entity_field", "").strip()
    entity_value = body.get("entity_value", "").strip()
    reason = body.get("reason", "").strip()
    report_name = body.get("report_name", "").strip()
    row_data = body.get("row_data", {})
    priority = body.get("priority", "medium").strip()
    if priority not in ("high", "medium", "low"):
        priority = "medium"

    if not entity_field or not entity_value:
        return resp(400, {"error": "entity_field and entity_value are required"})

    aid = str(uuid.uuid4())
    _crm_put("alerts", aid, {
        "alert_id": aid,
        "entity_field": entity_field,
        "entity_value": entity_value,
        "reason": reason,
        "report_name": report_name,
        "row_data": json.dumps(row_data, default=str) if not isinstance(row_data, str) else row_data,
        "status": "active",
        "priority": priority,
        "assigned_to": "",
        "reviewed_by": "",
        "reviewed_at": "",
        "notes": "",
        "created_at": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    })
    return resp(201, {"alert_id": aid})


def review_alert(alert_id: str, body: dict | None = None):
    """Move an alert from 'active' to 'reviewed' (ya revisados)."""
    body = body or {}
    changes = {
        "status": "reviewed",
        "reviewed_at": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if body.get("reviewed_by", "").strip():
        changes["reviewed_by"] = body["reviewed_by"].strip()
    if body.get("notes", "").strip():
        changes["notes"] = body["notes"].strip()
    if _crm_update("alerts", alert_id, changes) is None:
        return resp(404, {"error": f"Alert '{alert_id}' not found"})
    _safe_audit(user_email=changes.get("reviewed_by", "unknown"), action="alert.review",
                entity_type="alert", entity_id=alert_id)
    return resp(200, {"message": f"Alert '{alert_id}' marked as reviewed"})


def delete_alert(alert_id: str):
    """Permanently remove an alert entry."""
    _crm_delete("alerts", alert_id)
    return resp(200, {"message": f"Alert '{alert_id}' permanently deleted"})


def assign_alert(alert_id: str, body: dict):
    """Assign an alert to a CRM user (by email)."""
    assigned_to = body.get("assigned_to", "").strip()
    if not assigned_to:
        return resp(400, {"error": "assigned_to is required"})
    if _crm_update("alerts", alert_id, {"assigned_to": assigned_to}) is None:
        return resp(404, {"error": f"Alert '{alert_id}' not found"})
    _safe_audit(user_email=body.get("actor_email", "unknown"), action="alert.assign",
                entity_type="alert", entity_id=alert_id, new_value={"assigned_to": assigned_to})
    return resp(200, {"message": f"Alert '{alert_id}' assigned to {assigned_to}"})


def update_alert_notes(alert_id: str, body: dict):
    """Update the analyst notes on an alert."""
    if _crm_update("alerts", alert_id, {"notes": body.get("notes", "").strip()}) is None:
        return resp(404, {"error": f"Alert '{alert_id}' not found"})
    return resp(200, {"message": "Notes updated"})


def _save_slack_users(mapping: dict) -> None:
    """Guarda el mapeo y limpia la caché del contenedor para que el cambio se
    vea de inmediato en esta Lambda."""
    global _slack_users_cache
    s3.put_object(
        Bucket=S3_BUCKET, Key=SLACK_USERS_KEY,
        Body=json.dumps(mapping, indent=2, ensure_ascii=False).encode("utf-8"),
        ContentType="application/json",
    )
    _slack_users_cache = {k.strip().lower(): v for k, v in mapping.items() if v}


def get_slack_users():
    """Mapeo correo -> user ID de Slack, con los analistas conocidos para que
    el mantenedor pueda mostrar quién todavía no tiene ID cargado."""
    mapping = dict(_slack_users())
    conocidos = []
    try:
        for u in _crm_list("users"):
            email = (u.get("email") or "").strip()
            if email and "@" in email:
                conocidos.append({
                    "email": email,
                    "full_name": u.get("full_name") or email,
                    "slack_user_id": mapping.get(email.lower(), ""),
                })
    except Exception:
        pass
    conocidos.sort(key=lambda x: x["full_name"].lower())
    # Entradas del mapeo que no corresponden a ningún analista del sistema.
    emails_conocidos = {c["email"].lower() for c in conocidos}
    huerfanas = [{"email": k, "full_name": "", "slack_user_id": v}
                 for k, v in mapping.items() if k not in emails_conocidos]
    return resp(200, {"analistas": conocidos, "sin_analista": huerfanas,
                      "total_con_id": sum(1 for v in mapping.values() if v)})


def upsert_slack_user(body: dict):
    """Carga o actualiza el ID de Slack de un correo."""
    denied = _require_admin(body)
    if denied:
        return denied
    email = str(body.get("email") or "").strip().lower()
    uid = str(body.get("slack_user_id") or "").strip()
    if not email or "@" not in email:
        return resp(400, {"error": "email válido es requerido"})
    # Los member ID de Slack son alfanuméricos y empiezan con U (personas) o
    # W (cuentas de Enterprise Grid).
    if uid and not re.fullmatch(r"[UW][A-Z0-9]{6,}", uid.upper()):
        return resp(400, {"error": "El ID de Slack debe empezar con U (o W) — ej. U0522F23D1V. "
                                   "Se copia desde el perfil de la persona en Slack: ⋮ → Copiar ID de miembro."})

    mapping = dict(_slack_users())
    if uid:
        mapping[email] = uid.upper() if uid.islower() else uid
    else:
        mapping.pop(email, None)  # ID vacío = quitar el mapeo
    _save_slack_users(mapping)
    _safe_audit(user_email=(body.get("actor_email") or "unknown"),
                action="slack_user.upsert", entity_type="config", entity_id=email,
                new_value={"slack_user_id": uid})
    return resp(200, {"email": email, "slack_user_id": uid, "total": len(mapping)})


def delete_slack_user(email: str, body: dict):
    denied = _require_admin(body)
    if denied:
        return denied
    email = str(email or "").strip().lower()
    mapping = dict(_slack_users())
    if email not in mapping:
        return resp(404, {"error": f"No hay mapeo cargado para {email}"})
    mapping.pop(email, None)
    _save_slack_users(mapping)
    _safe_audit(user_email=(body.get("actor_email") or "unknown"),
                action="slack_user.delete", entity_type="config", entity_id=email)
    return resp(200, {"message": f"Mapeo de {email} eliminado", "total": len(mapping)})


def get_crm_users():
    """Return active CRM users for the assignee dropdown (S3-backed)."""
    try:
        users = [
            {"email": u.get("email", ""), "full_name": u.get("full_name") or u.get("email", "")}
            for u in _crm_list("users") if u.get("is_active", True)
        ]
        users.sort(key=lambda u: u["full_name"])
        return resp(200, {"users": users})
    except Exception as e:
        return resp(200, {"users": [], "warning": str(e)})


# ---------------------------------------------------------------------------
# CASES CRM
# ---------------------------------------------------------------------------

_CASE_STATUS_RANK = {"open": 1, "in_progress": 2, "under_review": 3, "closed": 4, "archived": 5}


def _now_str() -> str:
    return dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def get_cases(status_filter=None, priority_filter=None, assigned_filter=None):
    """List cases with optional filters (S3-backed). Ordered by status urgency,
    priority, then updated_at DESC."""
    try:
        out = []
        for c in _crm_list("cases"):
            if status_filter and status_filter != "all" and c.get("status") != status_filter:
                continue
            if priority_filter and c.get("priority") != priority_filter:
                continue
            if assigned_filter and c.get("assigned_to") != assigned_filter:
                continue
            out.append({
                "case_id": c.get("case_id", ""),
                "title": c.get("title", ""),
                "description": c.get("description", ""),
                "status": c.get("status", "open"),
                "priority": c.get("priority", "medium"),
                "entity_type": c.get("entity_type", ""),
                "entity_id": c.get("entity_id", ""),
                "entity_name": c.get("entity_name", ""),
                "report_name": c.get("report_name", ""),
                "assigned_to": c.get("assigned_to", ""),
                "created_by": c.get("created_by", ""),
                "created_at": c.get("created_at", ""),
                "updated_at": c.get("updated_at", ""),
                "closed_at": c.get("closed_at", ""),
                "note_count": len(c.get("notes", [])),
            })
        out.sort(key=lambda x: x["updated_at"], reverse=True)
        out.sort(key=lambda x: (_CASE_STATUS_RANK.get(x["status"], 5),
                                _PRIORITY_RANK.get(x["priority"], 2)))
        return resp(200, {"cases": out})
    except Exception as e:
        return resp(200, {"cases": [], "warning": str(e)})


def create_case(body: dict):
    title = body.get("title", "").strip()
    if not title:
        return resp(400, {"error": "title is required"})

    priority = body.get("priority", "medium").strip()
    if priority not in ("high", "medium", "low"):
        priority = "medium"
    assigned_to = body.get("assigned_to", "").strip()
    created_by = body.get("created_by", "unknown").strip()

    # Datos de la alerta que originó el caso. Se guardan tal cual la fila del
    # reporte para que el analista vea en el caso los mismos números que
    # gatillaron la alerta, sin tener que volver a correr la consulta.
    alert_data = body.get("alert_data") or {}
    if isinstance(alert_data, str):
        try:
            alert_data = json.loads(alert_data)
        except Exception:
            alert_data = {}
    if not isinstance(alert_data, dict):
        alert_data = {}

    cid = str(uuid.uuid4())
    now = _now_str()
    _crm_put("cases", cid, {
        "case_id": cid,
        "title": title,
        "description": body.get("description", "").strip(),
        "status": "open",
        "priority": priority,
        "entity_type": body.get("entity_type", "").strip(),
        "entity_id": body.get("entity_id", "").strip(),
        "entity_name": body.get("entity_name", "").strip(),
        "report_name": body.get("report_name", "").strip(),
        "alert_data": alert_data,
        "alert_priority": (body.get("alert_priority") or "").strip(),
        "assigned_to": assigned_to,
        "created_by": created_by,
        "created_at": now,
        "updated_at": now,
        "closed_at": "",
        "notes": [],
    })
    _safe_audit(user_email=created_by, action="case.create", entity_type="case",
                entity_id=cid, new_value={"title": title, "priority": priority})
    # Mensaje a Slack con el contexto completo: quién es el cliente, qué
    # alerta lo originó y con qué números — para poder triar sin abrir la app.
    caso_slack = {
        "entity_name": body.get("entity_name", "").strip(),
        "entity_id": body.get("entity_id", "").strip(),
        "entity_type": body.get("entity_type", "").strip(),
    }
    lineas = [f"📁 *Nuevo caso creado*", f"*{title}*"]
    cliente = _slack_client_line(caso_slack)
    if cliente:
        lineas.append(cliente)

    reporte = body.get("report_name", "").strip()
    alert_prio = (body.get("alert_priority") or "").strip()
    if reporte or alert_prio:
        alerta_txt = " · ".join(x for x in (reporte, alert_prio) if x)
        lineas.append(f"🔔 Alerta: {alerta_txt}")

    metricas = _slack_alert_metrics(alert_data)
    if metricas:
        lineas.append(f"📊 {metricas}")

    lineas.append(
        f"⚡ Prioridad: {_PRIORITY_LABEL_ES.get(priority, priority)}  ·  "
        f"Asignado a: {_slack_mention(assigned_to)}"
    )
    lineas.append(f"✍️ Creado por: {created_by}")
    lineas.append(_slack_case_link(cid))
    _post_slack("\n".join(lineas))

    if assigned_to and "@" in assigned_to:
        _case_assignment_email(assigned_to, cid, title, priority, created_by)
    return resp(201, {"case_id": cid})


def get_case_detail(case_id: str):
    """Return full case data including notes and linked alerts (S3-backed)."""
    try:
        c = _crm_get("cases", case_id)
        if c is None:
            return resp(404, {"error": f"Case '{case_id}' not found"})
        notes = sorted(c.get("notes", []), key=lambda n: n.get("created_at", ""))
        # Linked alerts = alerts whose case_id points here
        alerts = []
        for a in _crm_list("alerts"):
            if a.get("case_id") == case_id:
                alerts.append({
                    "alert_id": a.get("alert_id", ""),
                    "entity_field": a.get("entity_field", ""),
                    "entity_value": a.get("entity_value", ""),
                    "reason": a.get("reason", ""),
                    "report_name": a.get("report_name", ""),
                    "created_at": a.get("created_at", ""),
                    "status": a.get("status", "active"),
                    "priority": a.get("priority", "medium"),
                })
        case_out = {k: v for k, v in c.items() if k != "notes"}

        # Última solicitud de documentos de este caso — se usa para prellenar
        # nombre/correo/plantilla al reenviar un correo desde el detalle del
        # caso (esos datos no viven en el caso mismo, solo en el audit trail).
        last_request = None
        for r in sorted(_crm_list("document_requests"), key=lambda r: r.get("created_at", "")):
            if r.get("case_id") == case_id:
                last_request = {
                    "correo": r.get("correo", ""),
                    "nombre_completo": r.get("nombre_completo", ""),
                    "template_key": r.get("template_key", ""),
                }

        return resp(200, {"case": case_out, "notes": notes, "alerts": alerts, "last_document_request": last_request})
    except Exception as e:
        return resp(500, {"error": str(e)})


def update_case(case_id: str, body: dict):
    """Update title, description, or priority."""
    changes = {}
    if "title" in body:
        changes["title"] = str(body["title"]).strip()
    if "description" in body:
        changes["description"] = str(body["description"]).strip()
    if "priority" in body and body["priority"] in ("high", "medium", "low"):
        changes["priority"] = body["priority"]
    if not changes:
        return resp(400, {"error": "No valid fields to update"})
    changes["updated_at"] = _now_str()
    if _crm_update("cases", case_id, changes) is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    return resp(200, {"message": "Case updated"})


def update_case_status(case_id: str, body: dict):
    """Change case status. Sets closed_at when status = 'closed'."""
    status = body.get("status", "").strip()
    valid = ("open", "in_progress", "under_review", "closed", "archived")
    if status not in valid:
        return resp(400, {"error": f"status must be one of {valid}"})

    changes = {"status": status, "updated_at": _now_str()}
    if status == "closed":
        changes["closed_at"] = _now_str()
    elif status != "archived":
        changes["closed_at"] = ""

    if _crm_update("cases", case_id, changes) is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    actor = body.get("actor_email", "unknown")
    _safe_audit(user_email=actor, action="case.status_change", entity_type="case",
                entity_id=case_id, new_value={"status": status})
    _STATUS_LABEL = {"under_review": "⚠️ Bajo Revisión", "closed": "✅ Cerrado", "open": "🔵 Abierto", "in_progress": "🔄 En Investigación"}
    if status in ("under_review", "closed"):
        _post_slack(
            f"{_STATUS_LABEL.get(status, status)} *Caso actualizado*\n"
            f"ID: {case_id[:8]}… | Nuevo estado: {_STATUS_LABEL.get(status, status)}\n"
            f"Por: {actor}"
        )
    return resp(200, {"message": f"Case status updated to {status}"})


def update_case_assign(case_id: str, body: dict):
    assigned_to = body.get("assigned_to", "").strip()
    actor = body.get("actor_email", "unknown")
    updated = _crm_update("cases", case_id, {"assigned_to": assigned_to, "updated_at": _now_str()})
    if updated is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    _safe_audit(user_email=actor, action="case.assign", entity_type="case",
                entity_id=case_id, new_value={"assigned_to": assigned_to})

    lineas = ["👤 *Caso asignado*", f"*{updated.get('title', '')}*"]
    cliente = _slack_client_line(updated)
    if cliente:
        lineas.append(cliente)
    if updated.get("report_name"):
        lineas.append(f"🔔 Alerta: {updated['report_name']}")
    lineas.append(
        f"⚡ Prioridad: {_PRIORITY_LABEL_ES.get(updated.get('priority', ''), updated.get('priority', ''))}"
        f"  ·  Estado: {_CASE_STATUS_LABEL_ES.get(updated.get('status', ''), updated.get('status', ''))}"
    )
    lineas.append(f"➡️ Asignado a: {_slack_mention(assigned_to)}  ·  por {actor}")
    lineas.append(_slack_case_link(case_id))
    _post_slack("\n".join(lineas))

    if assigned_to and "@" in assigned_to:
        _case_assignment_email(assigned_to, case_id, updated.get("title", ""),
                               updated.get("priority", "medium"), actor)
    return resp(200, {"message": f"Case assigned to {assigned_to}"})


# ---------------------------------------------------------------------------
# Permisos de administrador
# ---------------------------------------------------------------------------
# El frontend ya esconde las acciones de admin según el rol, pero eso es solo
# cosmético: la API no valida tokens (el header Authorization se omite a
# propósito por la config de CORS — ver api() en el frontend). Este guard
# agrega una barrera real del lado del servidor usando el store de usuarios en
# S3 (crm/users), que es el mismo que administra el panel Admin.
# Falla CERRADO: si el usuario no existe, está inactivo, o no se puede leer el
# store, no es admin.
_ADMIN_ROLE_NAMES = {"ADMIN", "SUPER_ADMIN", "SUPERADMIN"}


def _is_admin_email(email: str) -> bool:
    email = (email or "").strip().lower()
    if not email:
        return False
    try:
        for u in _crm_list("users"):
            if (u.get("email") or "").strip().lower() != email:
                continue
            if not u.get("is_active", True):
                return False
            role = (u.get("role_name") or "").strip().upper().replace("-", "_")
            return role in _ADMIN_ROLE_NAMES
    except Exception:
        return False
    return False


def _require_admin(body: dict):
    """Devuelve una respuesta 403 si el actor no es admin, o None si puede seguir."""
    actor = (body.get("actor_email") or "").strip()
    if not _is_admin_email(actor):
        return resp(403, {
            "error": "Esta acción está reservada para administradores.",
            "actor_email": actor,
        })
    return None


def delete_case(case_id: str, body: dict):
    """Elimina un caso definitivamente — reservado para administradores.

    Borra también los adjuntos del caso en S3 y desvincula las alertas que
    apuntaban a él (vuelven a quedar sin caso), para no dejar referencias
    colgando a un caso que ya no existe.
    """
    denied = _require_admin(body)
    if denied:
        return denied

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    actor = (body.get("actor_email") or "").strip()

    # Adjuntos: se borran del bucket (best-effort, uno por uno).
    deleted_files = 0
    for att in case.get("attachments", []) or []:
        key = att.get("s3_key")
        if not key:
            continue
        try:
            s3.delete_object(Bucket=S3_BUCKET, Key=key)
            deleted_files += 1
        except Exception:
            pass

    # Alertas vinculadas: vuelven a quedar sin caso, no se borran.
    unlinked = 0
    try:
        for a in _crm_list("alerts"):
            if a.get("case_id") == case_id and a.get("alert_id"):
                _crm_update("alerts", a["alert_id"], {"case_id": ""})
                unlinked += 1
    except Exception:
        pass

    _crm_delete("cases", case_id)
    _safe_audit(user_email=actor or "unknown", action="case.delete", entity_type="case",
                entity_id=case_id,
                new_value={"title": case.get("title", ""), "status": case.get("status", ""),
                           "attachments_deleted": deleted_files, "alerts_unlinked": unlinked})
    return resp(200, {
        "message": "Caso eliminado",
        "case_id": case_id,
        "attachments_deleted": deleted_files,
        "alerts_unlinked": unlinked,
    })


_BULK_ASSIGN_MAX = 500


def _bulk_assignment_email(to_email: str, items: list[dict], assigned_by: str,
                           kind: str = "casos") -> dict:
    """Un solo correo con todo lo que le tocó al analista en un reparto masivo
    — mandar un correo por caso/alerta sería spam.

    `items` son dicts {id, title}; `kind` es solo el sustantivo que se muestra
    ("casos" o "alertas")."""
    rows = "".join(
        '<tr>'
        f'<td style="padding:6px 10px;border-bottom:1px solid #1e293b;font-family:monospace;font-size:12px;color:#64748b">{html.escape(str(i.get("id", ""))[:8])}</td>'
        f'<td style="padding:6px 10px;border-bottom:1px solid #1e293b;font-size:13px;color:#e2e8f0">{html.escape(str(i.get("title", "")))}</td>'
        '</tr>'
        for i in items
    )
    body_html = f"""
<div style="font-family:system-ui,sans-serif;background:#0f172a;color:#e2e8f0;padding:24px;border-radius:12px;max-width:640px">
  <div style="background:#1e293b;border-radius:8px;padding:16px 20px;margin-bottom:16px">
    <p style="margin:0;font-size:13px;color:#94a3b8">WatchTower AML &middot; Global66 Compliance</p>
    <h2 style="margin:8px 0 0;font-size:18px;color:#fff">Se te asignaron {len(items)} {kind}</h2>
  </div>
  <table style="width:100%;border-collapse:collapse">{rows}</table>
  <p style="font-size:12px;color:#475569;margin-top:16px">Asignado por: <strong style="color:#94a3b8">{html.escape(assigned_by or "—")}</strong></p>
</div>
"""
    return _send_email(to_email, f"[WatchTower] Se te asignaron {len(items)} {kind}", body_html)


# Orden de preferencia para identificar a quién apunta una fila de un reporte.
# Es el mismo criterio que usa el botón individual "Marcar como Alertado" en
# el frontend, extendido con company_id para los reportes B2B.
_ROW_ENTITY_PATTERNS = [
    ("customer_id", re.compile(r"^(customer_id|client_id)$", re.I)),
    ("company_id", re.compile(r"^company_id$", re.I)),
    ("beneficiary_id", re.compile(r"^beneficiary_id$", re.I)),
    ("email", re.compile(r"email", re.I)),
]


def _extract_entity_from_row(row: dict) -> tuple[str, str]:
    """Devuelve (entity_field, entity_value) de una fila de reporte, o
    ("", "") si la fila no identifica a ningún cliente/empresa."""
    for field, pattern in _ROW_ENTITY_PATTERNS:
        for key in row:
            if pattern.match(str(key)) or (field == "email" and pattern.search(str(key))):
                value = str(row[key] if row[key] is not None else "").strip()
                if value and value.lower() not in ("none", "nan", "—"):
                    return field, value
    return "", ""


def bulk_distribute_alerts(body: dict):
    """Reparte filas de un reporte como alertas ya asignadas entre analistas.

    Es el flujo real de trabajo: se corre un reporte (ej. Patrones AML) y
    desde la misma tabla de resultados se seleccionan las filas y se reparten
    entre el equipo — cada fila queda como una alerta asignada, sin tener que
    marcarlas y asignarlas una por una.

    body: {rows: [...], report_name, assignees: [...], actor_email,
           priority, reason, notify}
    """
    denied = _require_admin(body)
    if denied:
        return denied

    rows = body.get("rows") or []
    if not rows:
        return resp(400, {"error": "rows es requerido (lista de al menos 1)"})
    if len(rows) > _BULK_ASSIGN_MAX:
        return resp(400, {"error": f"Máximo {_BULK_ASSIGN_MAX} filas por operación"})

    assignees = [str(a).strip() for a in (body.get("assignees") or []) if str(a).strip()]
    if not assignees:
        return resp(400, {"error": "assignees es requerido (lista de analistas)"})

    report_name = (body.get("report_name") or "").strip()
    actor = (body.get("actor_email") or "").strip()
    priority = (body.get("priority") or "medium").strip()
    if priority not in ("high", "medium", "low"):
        priority = "medium"
    reason = (body.get("reason") or "").strip() or f"Repartido desde: {report_name or 'reporte'}"
    notify = bool(body.get("notify", True))

    per_assignee: dict[str, list[dict]] = {a: [] for a in assignees}
    created, skipped = [], []
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            skipped.append({"index": idx, "motivo": "fila inválida"})
            continue
        field, value = _extract_entity_from_row(row)
        if not value:
            skipped.append({"index": idx, "motivo": "la fila no identifica cliente/empresa"})
            continue
        # El round-robin va sobre las filas efectivamente creadas, no sobre el
        # índice original — así las filas descartadas no dejan huecos en el
        # reparto (si no, alguien podría recibir menos por pura casualidad).
        target = assignees[len(created) % len(assignees)]
        aid = str(uuid.uuid4())
        _crm_put("alerts", aid, {
            "alert_id": aid,
            "entity_field": field,
            "entity_value": value,
            "reason": reason,
            "report_name": report_name,
            "row_data": json.dumps(row, default=str),
            "status": "active",
            "priority": priority,
            "assigned_to": target,
            "reviewed_by": "",
            "reviewed_at": "",
            "notes": "",
            "created_at": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        })
        per_assignee[target].append({"id": aid, "title": f"{field} {value}"})
        created.append({"alert_id": aid, "assigned_to": target, "entity_value": value})

    # Notificaciones: se reporta el resultado real por destinatario, para que
    # el analista que reparte sepa si el correo salió o no (antes fallaba en
    # silencio y nadie se enteraba).
    notificaciones = []
    if notify:
        for email, items in per_assignee.items():
            if items and "@" in email:
                r = _bulk_assignment_email(email, items, actor, kind="alertas")
                notificaciones.append({"email": email, **r})

    _safe_audit(user_email=actor or "unknown", action="alert.bulk_distribute",
                entity_type="alert", entity_id=f"{len(created)} alertas",
                new_value={"report_name": report_name, "assignees": assignees,
                           "created": len(created), "skipped": len(skipped),
                           "notificaciones": notificaciones})

    return resp(200, {
        "created": len(created),
        "skipped": skipped,
        "por_analista": {k: len(v) for k, v in per_assignee.items()},
        "notificaciones": notificaciones,
        "correos_fallidos": [n for n in notificaciones if not n["sent"]],
    })


def bulk_assign_cases(body: dict):
    """Asignación masiva de casos — reservado para administradores.

    Dos modos:
      - directo:    {"case_ids": [...], "assigned_to": "ana@global66.com"}
                    todos los casos seleccionados van al mismo analista.
      - equitativo: {"case_ids": [...], "assignees": ["ana@...", "luis@..."]}
                    los casos se reparten en round-robin entre la lista — es
                    el modo pensado para repartir N alertas por persona.

    Manda UN correo resumen por analista, no uno por caso.
    """
    denied = _require_admin(body)
    if denied:
        return denied

    case_ids = body.get("case_ids") or []
    if not case_ids:
        return resp(400, {"error": "case_ids es requerido (lista de al menos 1)"})
    if len(case_ids) > _BULK_ASSIGN_MAX:
        return resp(400, {"error": f"Máximo {_BULK_ASSIGN_MAX} casos por operación"})

    assigned_to = (body.get("assigned_to") or "").strip()
    assignees = [str(a).strip() for a in (body.get("assignees") or []) if str(a).strip()]
    if assigned_to:
        assignees = [assigned_to]
    if not assignees:
        return resp(400, {"error": "Indicá assigned_to (un analista) o assignees (lista para repartir)"})

    actor = (body.get("actor_email") or "").strip()
    notify = bool(body.get("notify", True))

    # Round-robin: el caso i va al analista i % len(assignees). Con un solo
    # analista en la lista es equivalente al modo directo.
    per_assignee: dict[str, list[dict]] = {a: [] for a in assignees}
    assigned, not_found = [], []
    for idx, cid in enumerate(case_ids):
        cid = str(cid).strip()
        target = assignees[idx % len(assignees)]
        updated = _crm_update("cases", cid, {"assigned_to": target, "updated_at": _now_str()})
        if updated is None:
            not_found.append(cid)
            continue
        per_assignee[target].append({"id": cid, "title": updated.get("title", "")})
        assigned.append({"case_id": cid, "assigned_to": target})

    notificaciones = []
    if notify:
        for email, items in per_assignee.items():
            if items and "@" in email:
                r = _bulk_assignment_email(email, items, actor, kind="casos")
                notificaciones.append({"email": email, **r})

    _safe_audit(user_email=actor or "unknown", action="case.bulk_assign", entity_type="case",
                entity_id=f"{len(assigned)} casos",
                new_value={"assignees": assignees, "assigned": len(assigned),
                           "not_found": len(not_found),
                           "notificaciones": notificaciones})

    if assigned:
        modo = "a un analista" if len(assignees) == 1 else f"entre {len(assignees)} analistas"
        lineas = [
            "👥 *Reparto masivo de casos*",
            f"*{len(assigned)} caso(s)* repartidos {modo} por {actor or 'desconocido'}",
        ]
        for email, items in per_assignee.items():
            if items:
                lineas.append(f"   • {_slack_mention(email)} — *{len(items)}* caso(s)")
        if not_found:
            lineas.append(f"⚠️ {len(not_found)} no encontrado(s)")
        fallidos = [n["email"] for n in notificaciones if not n.get("sent")]
        if fallidos:
            lineas.append(f"📭 No se pudo avisar por correo a: {', '.join(fallidos)}")
        lineas.append(f"🔗 <{WATCHTOWER_URL}|Abrir en WatchTower>")
        _post_slack("\n".join(lineas))

    return resp(200, {
        "assigned": len(assigned),
        "not_found": not_found,
        "por_analista": {k: len(v) for k, v in per_assignee.items()},
        "notificaciones": notificaciones,
        "correos_fallidos": [n for n in notificaciones if not n["sent"]],
    })


def get_case_client_profile(case_id: str, body: dict):
    """Trae la ficha KYC/compliance del cliente del caso y la deja GUARDADA.

    Se consulta una sola vez: queda dentro del caso y las siguientes aperturas
    la leen de ahí sin volver a pegarle al cluster. Con refresh=true se fuerza
    una consulta nueva (por ejemplo si el cliente actualizó sus datos).
    """
    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    if case.get("client_profile") and not body.get("refresh"):
        return resp(200, {
            "profile": case["client_profile"],
            "cached": True,
            "consultado_at": case.get("client_profile_at", ""),
            "kind": case.get("client_profile_kind", "b2c"),
        })

    entity_id = str(case.get("entity_id") or "").strip()
    if not entity_id:
        return resp(400, {"error": "El caso no tiene un ID de cliente asociado."})

    entity_type = str(case.get("entity_type") or "").lower()
    kind = "b2b" if entity_type.startswith("comp") else "b2c"

    try:
        rows = _lookup_customer_rows(entity_id, kind)
    except ValueError as e:
        return resp(400, {"error": str(e)})
    except Exception as e:
        return resp(200, {"error": f"No se pudo consultar la ficha: {e}"})

    if not rows:
        return resp(200, {"error": f"No se encontró información para el ID {entity_id}."})

    profile = rows[0]
    ahora = _now_str()
    cambios = {
        "client_profile": profile,
        "client_profile_at": ahora,
        "client_profile_kind": kind,
        "updated_at": ahora,
    }
    # Si el caso no tenía nombre de cliente, se aprovecha para completarlo.
    if not case.get("entity_name"):
        nombre = _display_name_from_profile(profile, kind)
        if nombre:
            cambios["entity_name"] = nombre
    _crm_update("cases", case_id, cambios)

    _safe_audit(user_email=(body.get("actor_email") or "unknown"),
                action="case.client_profile", entity_type="case", entity_id=case_id,
                new_value={"entity_id": entity_id, "kind": kind})

    return resp(200, {
        "profile": profile,
        "cached": False,
        "consultado_at": ahora,
        "kind": kind,
        "entity_name": cambios.get("entity_name", case.get("entity_name", "")),
    })


def take_case(case_id: str, body: dict):
    """"Tomar caso" — el analista se auto-asigna el caso.

    NO es admin-only a propósito: es justamente la vía para que cada analista
    levante trabajo por su cuenta. Si el caso ya está tomado por otra persona
    no lo pisa (409), salvo que se mande force=true — así nadie le saca un
    caso a un compañero sin querer.
    """
    actor = (body.get("actor_email") or "").strip()
    if not actor:
        return resp(400, {"error": "actor_email es requerido"})

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    current = (case.get("assigned_to") or "").strip()
    if current and current.lower() != actor.lower() and not body.get("force"):
        return resp(409, {
            "error": f"El caso ya está asignado a {current}.",
            "assigned_to": current,
        })

    _crm_update("cases", case_id, {"assigned_to": actor, "updated_at": _now_str()})
    _safe_audit(user_email=actor, action="case.take", entity_type="case",
                entity_id=case_id, new_value={"assigned_to": actor, "previous": current})

    lineas = ["✋ *Caso tomado*", f"*{case.get('title', '')}*"]
    cliente = _slack_client_line(case)
    if cliente:
        lineas.append(cliente)
    lineas.append(f"➡️ Lo tomó: {_slack_mention(actor)}" + (f"  ·  antes: {current}" if current else ""))
    lineas.append(_slack_case_link(case_id))
    _post_slack("\n".join(lineas))

    return resp(200, {"message": f"Caso tomado por {actor}", "assigned_to": actor})


def add_case_note(case_id: str, body: dict):
    content = body.get("content", "").strip()
    if not content:
        return resp(400, {"error": "content is required"})
    author_email = body.get("author_email", "").strip()
    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    note = {
        "note_id": str(uuid.uuid4()),
        "case_id": case_id,
        "author_email": author_email,
        "content": content,
        "created_at": _now_str(),
    }
    case.setdefault("notes", []).append(note)
    case["updated_at"] = _now_str()
    _crm_put("cases", case_id, case)
    _safe_audit(user_email=author_email or "unknown", action="case.note_add",
                entity_type="case", entity_id=case_id)
    return resp(201, {"message": "Note added"})


def link_alert_to_case(alert_id: str, body: dict):
    """Link an alert to a case (sets case_id on the alert)."""
    case_id = body.get("case_id", "").strip()
    if not case_id:
        return resp(400, {"error": "case_id is required"})
    alerta = _crm_update("alerts", alert_id, {"case_id": case_id})
    if alerta is None:
        return resp(404, {"error": f"Alert '{alert_id}' not found"})

    cambios = {"updated_at": _now_str()}
    # Si el caso todavía no tiene los datos de una alerta, se copian los de
    # esta — así al abrir el caso se ven los números que la gatillaron.
    caso = _crm_get("cases", case_id) or {}
    if not caso.get("alert_data"):
        row = alerta.get("row_data")
        if isinstance(row, str):
            try:
                row = json.loads(row)
            except Exception:
                row = None
        if isinstance(row, dict) and row:
            cambios["alert_data"] = row
            cambios["alert_priority"] = alerta.get("priority", "")
    _crm_update("cases", case_id, cambios)
    return resp(200, {"message": f"Alert '{alert_id}' linked to case '{case_id}'"})


# ---------------------------------------------------------------------------
# CASE ATTACHMENTS — documentos adjuntos a un caso, organizados en S3 por
# cliente (customer_id o company_id), no solo por caso, para que todo lo que
# se le pidió/recibió de un cliente sea encontrable aunque abra varios casos
# a lo largo del tiempo:
#
#   client-documents/{entity_type}/{entity_id}/{case_id}/{ts}_{filename}
#
# Subida vía presigned PUT (el navegador sube directo a S3, sin pasar por el
# Lambda) — evita el límite de payload del Lambda/API Gateway para archivos
# grandes (PDFs escaneados, etc). Mismo patrón que ya usa el proyecto para
# descargas de reportes (presigned URL).
# ---------------------------------------------------------------------------
def _safe_filename(name: str) -> str:
    name = (name or "archivo").strip().replace("/", "_").replace("\\", "_")
    name = re.sub(r"[^A-Za-z0-9._\-]", "_", name)
    return name[:200] or "archivo"


def _attachment_s3_key(entity_type: str, entity_id: str, case_id: str, ts: str, filename: str) -> str:
    entity_type = (entity_type or "customer").strip().lower()
    if entity_type not in ("customer", "company"):
        entity_type = "customer"
    entity_id = re.sub(r"[^A-Za-z0-9_\-]", "_", str(entity_id or "sin_id"))
    return f"client-documents/{entity_type}/{entity_id}/{case_id}/{ts}_{_safe_filename(filename)}"


def get_attachment_upload_url(case_id: str, body: dict):
    """Paso 1 de la subida: devuelve una URL PUT firmada + la key donde va a
    quedar el archivo. El navegador sube directo a S3 con esa URL."""
    filename = body.get("filename", "").strip()
    if not filename:
        return resp(400, {"error": "filename is required"})
    content_type = body.get("content_type", "").strip() or "application/octet-stream"

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    key = _attachment_s3_key(case.get("entity_type", ""), case.get("entity_id", ""), case_id, ts, filename)

    upload_url = s3.generate_presigned_url(
        "put_object",
        Params={"Bucket": S3_BUCKET, "Key": key, "ContentType": content_type},
        ExpiresIn=300,
    )
    return resp(200, {"upload_url": upload_url, "s3_key": key, "content_type": content_type})


def add_case_attachment(case_id: str, body: dict):
    """Paso 2: el navegador ya subió el archivo a S3 con la URL del paso 1;
    esto registra la metadata en el caso (el archivo en sí no pasa por acá)."""
    s3_key = body.get("s3_key", "").strip()
    filename = body.get("filename", "").strip()
    if not s3_key or not filename:
        return resp(400, {"error": "s3_key and filename are required"})

    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})

    uploaded_by = body.get("uploaded_by", "").strip()
    attachment = {
        "attachment_id": str(uuid.uuid4()),
        "filename": filename,
        "s3_key": s3_key,
        "size": int(body.get("size") or 0),
        "content_type": body.get("content_type", "").strip(),
        "uploaded_by": uploaded_by,
        "uploaded_at": _now_str(),
    }
    case.setdefault("attachments", []).append(attachment)
    case["updated_at"] = _now_str()
    _crm_put("cases", case_id, case)
    _safe_audit(user_email=uploaded_by or "unknown", action="case.attachment_add",
                entity_type="case", entity_id=case_id, new_value={"filename": filename})
    return resp(201, {"attachment": attachment})


def get_attachment_download_url(case_id: str, attachment_id: str):
    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    for a in case.get("attachments", []):
        if a.get("attachment_id") == attachment_id:
            url = s3.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": S3_BUCKET, "Key": a["s3_key"],
                    "ResponseContentDisposition": f'attachment; filename="{a.get("filename","archivo")}"',
                },
                ExpiresIn=300,
            )
            return resp(200, {"download_url": url, "filename": a.get("filename", "")})
    return resp(404, {"error": "Attachment not found"})


def delete_case_attachment(case_id: str, attachment_id: str):
    case = _crm_get("cases", case_id)
    if case is None:
        return resp(404, {"error": f"Case '{case_id}' not found"})
    attachments = case.get("attachments", [])
    target = next((a for a in attachments if a.get("attachment_id") == attachment_id), None)
    if target is None:
        return resp(404, {"error": "Attachment not found"})
    try:
        s3.delete_object(Bucket=S3_BUCKET, Key=target["s3_key"])
    except Exception:
        pass
    case["attachments"] = [a for a in attachments if a.get("attachment_id") != attachment_id]
    case["updated_at"] = _now_str()
    _crm_put("cases", case_id, case)
    _safe_audit(action="case.attachment_delete", entity_type="case", entity_id=case_id,
                new_value={"filename": target.get("filename", "")})
    return resp(200, {"message": "Attachment deleted"})


# ---------------------------------------------------------------------------
# Escucha de respuestas del cliente (documentos por correo)
# ---------------------------------------------------------------------------
def _get_imap_password() -> str:
    if not DOC_REPLY_IMAP_SECRET_ARN:
        return ""
    try:
        val = secrets_client.get_secret_value(SecretId=DOC_REPLY_IMAP_SECRET_ARN)
        return val["SecretString"].strip()
    except Exception:
        return ""


def _decode_mime_words(s: str) -> str:
    from email.header import decode_header
    out = []
    for text, enc in decode_header(s or ""):
        out.append(text.decode(enc or "utf-8", errors="replace") if isinstance(text, bytes) else text)
    return "".join(out)


_QUOTE_HEADER_RE = re.compile(
    r"^\s*(El .+escribi[oó]:|On .+wrote:|>+.*)\s*$",
    re.IGNORECASE | re.MULTILINE,
)


def _strip_quoted_reply(text: str) -> str:
    """Corta el texto en la primera línea que arranca el thread citado
    (encabezado tipo "El ... escribió:" / "On ... wrote:") o una línea de
    cita ("> ..."), para que la nota del caso muestre solo lo que el
    cliente realmente escribió, no el correo completo de vuelta."""
    m = _QUOTE_HEADER_RE.search(text)
    return (text[:m.start()] if m else text).strip()


def _extract_email_text(msg) -> str:
    """Cuerpo de texto plano del mensaje (preferido) o HTML despojado de
    tags si no hay texto plano, recortado del thread citado. Usado para
    dejar como nota en el caso la respuesta de un cliente que no adjuntó
    nada."""
    for part in msg.walk():
        if part.get_filename() or part.get_content_type() != "text/plain":
            continue
        payload = part.get_payload(decode=True)
        if payload:
            charset = part.get_content_charset() or "utf-8"
            try:
                text = payload.decode(charset, errors="replace").strip()
            except LookupError:
                text = payload.decode("utf-8", errors="replace").strip()
            return _strip_quoted_reply(text)
    for part in msg.walk():
        if part.get_filename() or part.get_content_type() != "text/html":
            continue
        payload = part.get_payload(decode=True)
        if payload:
            charset = part.get_content_charset() or "utf-8"
            html = payload.decode(charset, errors="replace")
            text = re.sub(r"<[^>]+>", " ", html)
            text = _strip_quoted_reply(text)
            return re.sub(r"[ \t]+", " ", text).strip()
    return ""


def _notify_client_reply(case: dict, adjuntos: list[str], estado_anterior: str,
                         from_addr: str = "", texto: str = "") -> None:
    """Avisa en Slack que un cliente respondió una solicitud de documentos.

    Es el aviso que faltaba: hasta ahora la respuesta se archivaba en el caso
    pero nadie se enteraba hasta abrirlo a mano. Va al canal de alertas
    operativas con formato Block Kit."""
    case_id = case.get("case_id", "")
    titulo = case.get("title", "(sin título)")
    con_docs = bool(adjuntos)
    estado_actual = case.get("status", "open")
    cambio_estado = estado_actual != estado_anterior

    nombre = str(case.get("entity_name") or "").strip()
    eid = str(case.get("entity_id") or "").strip()
    etype = str(case.get("entity_type") or "").lower()
    campo = "company_id" if etype.startswith("comp") else "customer_id"
    cliente = nombre or "(sin nombre)"
    if eid:
        cliente += f"\n`{campo} {eid}`"

    asignado = _slack_mention(case.get("assigned_to"))

    encabezado = ("📥 Documentos recibidos del cliente" if con_docs
                  else "💬 Respuesta del cliente (sin documentos)")

    checklist = case.get("documentos_checklist") or []
    recibidos = sum(1 for i in checklist if i.get("estado") == "recibido")
    entregados = sum(1 for i in checklist if i.get("estado") == "entregado")
    pendientes = sum(1 for i in checklist if i.get("estado") == "pendiente")

    blocks: list[dict] = [
        {"type": "header", "text": {"type": "plain_text", "text": encabezado, "emoji": True}},
        {"type": "section", "text": {"type": "mrkdwn", "text": f"*{titulo}*"}},
        {"type": "section", "fields": [
            {"type": "mrkdwn", "text": f"*Cliente*\n{cliente}"},
            {"type": "mrkdwn", "text": f"*Analista*\n{asignado}"},
        ]},
    ]

    if con_docs:
        lista = "\n".join(f"• `{a}`" for a in adjuntos[:8])
        if len(adjuntos) > 8:
            lista += f"\n• _y {len(adjuntos) - 8} más_"
        blocks.append({"type": "section",
                       "text": {"type": "mrkdwn",
                                "text": f"*Archivos adjuntos ({len(adjuntos)})*\n{lista}"}})
        blocks.append({"type": "section",
                       "text": {"type": "mrkdwn",
                                "text": (f"*Checklist* — {recibidos} por validar · "
                                         f"{entregados} entregado(s) · {pendientes} pendiente(s)\n"
                                         "⚠️ Los documentos quedan como *Recibido*: hay que abrirlos "
                                         "y marcarlos como *Entregado* a mano.")}})
    else:
        extracto = (texto or "").strip().replace("\n", " ")[:280]
        if extracto:
            blocks.append({"type": "section",
                           "text": {"type": "mrkdwn", "text": f"> {extracto}"}})
        blocks.append({"type": "section",
                       "text": {"type": "mrkdwn",
                                "text": ("El cliente respondió pero *no adjuntó documentos*. "
                                         "El checklist sigue en *Pendiente* — puede que haya que "
                                         "reinsistir o aclararle qué falta.")}})

    contexto = [f"Caso `{case_id[:8]}…`"]
    if cambio_estado:
        contexto.append(f"Estado: {_CASE_STATUS_LABEL_ES.get(estado_anterior, estado_anterior)} → "
                        f"*{_CASE_STATUS_LABEL_ES.get(estado_actual, estado_actual)}*")
    else:
        contexto.append(f"Estado: {_CASE_STATUS_LABEL_ES.get(estado_actual, estado_actual)} (sin cambios)")
    if from_addr:
        contexto.append(f"De: {from_addr}")
    blocks.append({"type": "context",
                   "elements": [{"type": "mrkdwn", "text": "  ·  ".join(contexto)}]})
    blocks.append({"type": "actions", "elements": [
        {"type": "button", "text": {"type": "plain_text", "text": "Abrir en WatchTower", "emoji": True},
         "url": WATCHTOWER_URL, "style": "primary"},
    ]})
    blocks.append({"type": "divider"})

    resumen = f"{encabezado} — {titulo}"
    _post_slack(resumen, blocks=blocks)


def _email_ya_procesado(message_id: str) -> bool:
    """¿Ya procesamos este correo antes?

    Antes el poller usaba el flag \\Seen de IMAP como control de duplicados:
    buscaba solo UNSEEN. El problema es que ese flag lo puede cambiar
    cualquiera — si una persona abría la casilla y leía la respuesta de un
    cliente antes de que corriera el poller, el mensaje quedaba \\Seen y NUNCA
    se procesaba: el documento no llegaba al caso y nadie se enteraba.

    Ahora el control es propio: se guarda el Message-ID de cada correo ya
    procesado, así da igual quién lo haya leído."""
    if not message_id:
        return False
    clave = hashlib.sha256(message_id.encode("utf-8", "replace")).hexdigest()[:32]
    return _crm_get("email_processed", clave) is not None


def _marcar_email_procesado(message_id: str, case_id: str, outcome: str) -> None:
    if not message_id:
        return
    clave = hashlib.sha256(message_id.encode("utf-8", "replace")).hexdigest()[:32]
    _crm_put("email_processed", clave, {
        "message_id": message_id, "case_id": case_id,
        "outcome": outcome, "processed_at": _now_str(),
    })


def _process_reply_message(imap, num, email_lib, move_matched_from_spam=False) -> str:
    """Procesa un único mensaje IMAP. Retorna 'matched' (adjuntos subidos,
    checklist a "recibido"), 'matched_no_attachment' (el caso matcheó pero
    el cliente no adjuntó nada válido — se deja una nota con el texto de su
    respuesta, el checklist NO se toca), 'unmatched' (no matchea ningún
    caso) o 'error'. No decide por sí solo si hay que marcarlo \\Seen — eso
    lo hace el caller, porque la política difiere entre INBOX (todo se
    marca leído) y Spam (solo lo que matchea, para no tocar spam real)."""
    # BODY.PEEK[] (no RFC822 plano) — leer un mensaje por IMAP marca \Seen como
    # efecto secundario a menos que se use PEEK; sin esto, cualquier lectura
    # (incluso solo para inspeccionar) marcaría como leído un spam real que
    # no matcheó ningún caso, rompiendo la garantía de "no tocar nada".
    status, msg_data = imap.fetch(num, "(BODY.PEEK[])")
    if status != "OK":
        return "error"
    msg = email_lib.message_from_bytes(msg_data[0][1])
    subject = _decode_mime_words(msg.get("Subject", ""))
    from_addr = email_lib.utils.parseaddr(msg.get("From", ""))[1]
    message_id = (msg.get("Message-ID") or "").strip()
    if _email_ya_procesado(message_id):
        return "duplicado"

    m = _EMAIL_REF_RE.search(subject)
    ref_record = _crm_get("email_refs", m.group(1).lower()) if m else None
    case = _crm_get("cases", ref_record["case_id"]) if ref_record else None
    if case is None:
        return "unmatched"
    case_id = ref_record["case_id"]

    saved_any = False
    nuevos_adjuntos: list[str] = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename = part.get_filename()
        if not filename:
            continue
        filename = _decode_mime_words(filename)
        if os.path.splitext(filename)[1].lower() not in _ATTACHMENT_EXTS_ALLOWED:
            continue
        payload = part.get_payload(decode=True)
        if not payload or len(payload) < 200:
            continue  # descarta íconos/firmas embebidas, no documentos reales

        ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        s3_key = _attachment_s3_key(case.get("entity_type", ""), case.get("entity_id", ""), case_id, ts, filename)
        s3.put_object(Bucket=S3_BUCKET, Key=s3_key, Body=payload,
                      ContentType=part.get_content_type() or "application/octet-stream")
        case.setdefault("attachments", []).append({
            "attachment_id": str(uuid.uuid4()),
            "filename": filename,
            "s3_key": s3_key,
            "size": len(payload),
            "content_type": part.get_content_type() or "",
            "uploaded_by": from_addr,
            "uploaded_at": _now_str(),
            "source": "email_reply",
        })
        nuevos_adjuntos.append(filename)
        saved_any = True

    if saved_any:
        checklist = case.get("documentos_checklist") or []
        for item in checklist:
            if item.get("estado") == "pendiente":
                item["estado"] = "recibido"
        case["documentos_checklist"] = checklist
        case["updated_at"] = _now_str()
        # El texto de la respuesta también se guarda como nota: muchas veces el
        # cliente explica ahí qué mandó, qué le falta o algo del origen de los
        # fondos. Antes se descartaba y solo quedaban los archivos, perdiendo
        # ese contexto.
        cuerpo = _extract_email_text(msg)[:2000].strip()
        if cuerpo:
            case.setdefault("notes", []).append({
                "note_id": str(uuid.uuid4()),
                "case_id": case_id,
                "author_email": from_addr or "cliente",
                "content": ("Respuesta del cliente por correo (con "
                            f"{len(nuevos_adjuntos)} documento(s) adjunto(s)):\n\n{cuerpo}"),
                "created_at": _now_str(),
            })
        # Llegó material: el caso pasa a "En Investigación" para que se vea que
        # hay algo que trabajar. Solo se promueve desde "open" — si ya estaba
        # en revisión o cerrado, no se pisa el criterio del analista.
        estado_anterior = case.get("status", "open")
        if estado_anterior == "open":
            case["status"] = "in_progress"
        _crm_put("cases", case_id, case)
        _safe_audit(user_email=from_addr or "cliente", action="case.email_reply_attachments",
                    entity_type="case", entity_id=case_id,
                    new_value={"n_attachments": len(case.get("attachments", [])),
                               "status": case.get("status"), "status_anterior": estado_anterior})
        _notify_client_reply(case, nuevos_adjuntos, estado_anterior, from_addr)
        outcome = "matched"
    else:
        # El caso matcheó pero no vino ningún adjunto válido — igual es una
        # respuesta real del cliente, así que queda registrada como nota
        # (no se toca el checklist: no se recibió ningún documento).
        body_text = _extract_email_text(msg)[:2000] or "(sin contenido de texto legible)"
        note = {
            "note_id": str(uuid.uuid4()),
            "case_id": case_id,
            "author_email": from_addr or "cliente",
            "content": f"Respuesta del cliente por correo, sin documentos adjuntos:\n\n{body_text}",
            "created_at": _now_str(),
        }
        case.setdefault("notes", []).append(note)
        case["updated_at"] = _now_str()
        estado_anterior = case.get("status", "open")
        if estado_anterior == "open":
            case["status"] = "in_progress"
        _crm_put("cases", case_id, case)
        _safe_audit(user_email=from_addr or "cliente", action="case.email_reply_no_attachment",
                    entity_type="case", entity_id=case_id,
                    new_value={"note_id": note["note_id"], "status": case.get("status"),
                               "status_anterior": estado_anterior})
        _notify_client_reply(case, [], estado_anterior, from_addr, texto=body_text)
        outcome = "matched_no_attachment"

    _marcar_email_procesado(message_id, case_id, outcome)

    if move_matched_from_spam:
        try:
            imap.copy(num, "INBOX")
            imap.store(num, "+FLAGS", "\\Deleted")
            imap.expunge()
        except Exception:
            pass  # no bloquea el match si Gmail no deja mover — igual queda guardado en el caso

    return outcome


def poll_document_replies() -> dict:
    """Revisa compliance.masivo@global66.com (cuenta Gmail real, miembro del
    grupo compliance@global66.com) por respuestas de clientes a una
    solicitud de documentos, sube los adjuntos al caso correspondiente en S3
    y pasa el checklist de "pendiente" a "recibido" — nunca a "entregado"
    directamente, eso lo confirma un analista a mano tras revisar el
    contenido (decisión de producto).

    La correlación caso <-> respuesta es por el token [ref: xxxxxxxx] en el
    asunto (ver _register_email_ref) — los clientes de correo preservan el
    asunto original al responder. Si no se encuentra el token o no matchea
    ningún caso en INBOX, el mensaje se marca leído igual (para no
    reprocesarlo en cada corrida) pero no se toca ningún caso.

    También revisa Spam — el primer correo de un remitente externo
    respondiendo a un alias/grupo cae ahí seguido — pero SOLO actúa sobre
    mensajes cuyo token matchea un caso real (y los mueve a INBOX); todo lo
    demás en Spam se deja intacto, nunca se marca leído ni se toca.

    Se invoca periódicamente vía EventBridge -> Report Runner Lambda con
    {"report_name": "poll_document_replies"}."""
    import imaplib
    import email as email_lib

    password = _get_imap_password()
    if not password:
        return {"status": "skipped", "reason": "DOC_REPLY_IMAP_SECRET_ARN no configurado"}

    try:
        # timeout explícito: sin esto, un stall de red deja la Lambda colgada
        # hasta su límite de 900s y las corridas programadas se apilan.
        imap = imaplib.IMAP4_SSL(DOC_REPLY_IMAP_HOST, timeout=30)
        imap.login(DOC_REPLY_IMAP_USER, password)
    except Exception as e:
        return {"status": "error", "error": f"No se pudo conectar/autenticar por IMAP: {e}"}

    processed = matched = matched_no_doc = unmatched = errors = duplicados = 0
    pendientes = 0
    # Topes por corrida: recorrer toda la casilla en una sola invocación
    # colgaba la Lambda. Se procesa de a tandas; lo que sobra queda para la
    # corrida siguiente (cada 10 minutos).
    MAX_POR_CARPETA = 40
    deadline = time.time() + 240  # 4 min, muy por debajo del límite de 900s

    # Se busca SOLO lo que trae nuestro token de referencia en el asunto
    # ("[ref: xxxxxxxx]", que el cliente conserva al responder), acotado a los
    # últimos días.
    #
    # Es clave que el filtro sea así de específico: compliance.masivo espeja al
    # grupo compliance@ y tiene ~1.500 mensajes sin leer que no tienen nada que
    # ver con esto. Filtrando por asunto no se los toca — antes se los abría y
    # se los marcaba como leídos solo para descartarlos, ensuciando una casilla
    # corporativa real.
    # OJO: NO se filtra por UNSEEN. El control de duplicados es el ledger de
    # Message-ID (ver _email_ya_procesado): si dependiéramos de \Seen, una
    # respuesta que alguien abrió a mano antes del poller no se procesaría
    # nunca y el documento no llegaría al caso.
    desde = (dt.datetime.utcnow() - dt.timedelta(days=DOC_REPLY_LOOKBACK_DAYS)).strftime("%d-%b-%Y")
    criterio = f'(SINCE {desde} SUBJECT "ref:")'

    try:
        # ── INBOX: todo lo no matcheado se marca leído (comportamiento normal) ──
        status, _ = imap.select("INBOX")
        if status == "OK":
            status, data = imap.search(None, criterio)
            nums = data[0].split() if status == "OK" else []
            pendientes += max(0, len(nums) - MAX_POR_CARPETA)
            for num in nums[:MAX_POR_CARPETA]:
                if time.time() > deadline:
                    pendientes += 1
                    continue
                try:
                    outcome = _process_reply_message(imap, num, email_lib)
                    if outcome == "duplicado":
                        duplicados += 1
                        continue
                    if outcome == "error":
                        errors += 1
                        continue
                    imap.store(num, "+FLAGS", "\\Seen")
                    processed += 1
                    matched += (outcome == "matched")
                    matched_no_doc += (outcome == "matched_no_attachment")
                    unmatched += (outcome == "unmatched")
                except Exception as e:
                    print(f"poll_document_replies: error procesando mensaje INBOX {num}: {e}")
                    errors += 1

        # ── Spam: solo se toca lo que matchea un token real; el resto ni se mira ──
        status, _ = imap.select('"[Gmail]/Spam"')
        if status == "OK":
            status, data = imap.search(None, criterio)
            nums = data[0].split() if status == "OK" else []
            pendientes += max(0, len(nums) - MAX_POR_CARPETA)
            for num in nums[:MAX_POR_CARPETA]:
                if time.time() > deadline:
                    pendientes += 1
                    continue
                try:
                    outcome = _process_reply_message(imap, num, email_lib, move_matched_from_spam=True)
                    if outcome in ("matched", "matched_no_attachment"):
                        processed += 1
                        matched += (outcome == "matched")
                        matched_no_doc += (outcome == "matched_no_attachment")
                    # unmatched/error en Spam: no se toca, no se marca, no se cuenta —
                    # es spam real de otra gente, no nuestro.
                except Exception as e:
                    print(f"poll_document_replies: error procesando mensaje Spam {num}: {e}")
    finally:
        try:
            imap.close()
            imap.logout()
        except Exception:
            pass

    resultado = {"status": "ok", "processed": processed, "matched": matched,
                 "matched_no_attachment": matched_no_doc, "unmatched": unmatched,
                 "errors": errors, "duplicados": duplicados, "pendientes": pendientes}
    print(f"poll_document_replies: {resultado}")
    return resultado


# ---------------------------------------------------------------------------
# DASHBOARD STATS
# ---------------------------------------------------------------------------

def get_dashboard_stats():
    """Submit 3 dashboard queries to Redshift Data API and return the statement IDs immediately.

    Uses a two-phase async pattern to avoid API Gateway's 30s timeout:
      1. This endpoint submits all 3 queries and returns stmt_ids in < 1s.
      2. The frontend polls /dashboard/stats/result?q0=id&q1=id&q2=id until all done.
    """
    try:
        stmt_ids: list[str] = []
        for sql in [_SQL_DAILY_EVOLUTION, _SQL_OVER_300K, _SQL_BY_COUNTRY]:
            r = redshift_data.execute_statement(
                ClusterIdentifier=CLUSTER_ID,
                Database=DATABASE_NAME,
                DbUser=DB_USER,
                Sql=sql.strip(),
            )
            stmt_ids.append(r["Id"])
        return resp(200, {"stmt_ids": stmt_ids})
    except Exception as e:
        msg = str(e)
        if "paused" in msg.lower() or "unavailable" in msg.lower() or "not available" in msg.lower():
            return resp(200, {
                "error": "cluster_paused",
                "message": "El cluster está pausado. Enciéndelo para cargar las estadísticas.",
            })
        return resp(200, {"error": str(e), "message": "Error enviando consultas al cluster."})


def run_transaction_search(body: dict):
    """Submit a transaction search by list of transaction_ids (remesas)."""
    transaction_ids = body.get("transaction_ids", [])
    if not transaction_ids:
        return resp(400, {"error": "transaction_ids is required"})
    if len(transaction_ids) > 5000:
        return resp(400, {"error": "Maximum 5000 transaction_ids per search"})

    clean_ids = []
    for tid in transaction_ids:
        try:
            clean_ids.append(int(str(tid).strip()))
        except (ValueError, TypeError):
            return resp(400, {"error": f"Invalid transaction_id: {tid!r}"})

    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()
    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": "transaction_search",
        "status": "RUNNING",
        "params": json.dumps({"transaction_ids": clean_ids, "n_transactions": len(clean_ids)}),
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })

    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps({
            "report_name": "transaction_search",
            "transaction_ids": clean_ids,
            "run_id": run_id,
            "keep_session": False,
        }),
    )
    return resp(202, {"run_id": run_id, "status": "RUNNING", "n_transactions": len(clean_ids)})


_SQL_REMESA_SEARCH = """
SELECT
    transaction_id,
    customer_id,
    beneficiary_country_name,
    CASE
        WHEN beneficiary_country_name = 'Chile' THEN 'Envío nacional'
        ELSE 'Envío internacional'
    END AS tipo_envio,
    beneficiary_dni,
    beneficiary_dni_type,
    beneficiary_name,
    beneficiary_first_name,
    beneficiary_last_name,
    beneficiary_email,
    beneficiary_id,
    origin_country,
    destiny_country,
    destiny_amount_usd,
    tx_status,
    start_date
FROM "db_prod"."transaction"."transaction"
WHERE transaction_id IN ({ids_sql})
ORDER BY start_date DESC
"""

# Máximo de remesas por consulta EN VIVO — deliberadamente bajo (a diferencia
# del límite de 5000 de la búsqueda async/Excel de arriba) porque esta corre
# sincrónica dentro de la misma invocación HTTP: hay que devolver la respuesta
# antes de que el otro sistema se canse de esperar.
_REMESA_SEARCH_SYNC_MAX = 50


def search_remesas_sync(body: dict):
    """Búsqueda de Remesas EN VIVO — misma query y misma tabla que
    run_transaction_search (Paso 2 del módulo de Análisis Individual), pero
    sincrónica: pensada para que OTRO proyecto mande el/los número(s) de
    remesa y reciba el detalle en la misma respuesta HTTP, sin polling de
    run_id ni generación de Excel.

    body: {transaction_id: 123} o {transaction_ids: [123, 456, ...]}
          (máx. 50 — para lotes más grandes usar /search/transactions, que
          es async y genera un Excel).
    """
    raw_ids = body.get("transaction_ids")
    if raw_ids is None:
        single = body.get("transaction_id")
        raw_ids = [single] if single is not None else []
    if not raw_ids:
        return resp(400, {"error": "transaction_id o transaction_ids es requerido"})
    if len(raw_ids) > _REMESA_SEARCH_SYNC_MAX:
        return resp(400, {
            "error": f"Máximo {_REMESA_SEARCH_SYNC_MAX} transaction_ids por consulta en vivo "
                     f"(para lotes más grandes usar /search/transactions, que es async)."
        })

    clean_ids = []
    for tid in raw_ids:
        try:
            clean_ids.append(int(str(tid).strip()))
        except (ValueError, TypeError):
            return resp(400, {"error": f"transaction_id inválido: {tid!r}"})

    ids_sql = ", ".join(str(i) for i in clean_ids)
    sql = _SQL_REMESA_SEARCH.format(ids_sql=ids_sql)
    try:
        rows = _rs_exec(sql)
    except RuntimeError as e:
        return resp(200, {"error": "cluster_unavailable", "message": str(e)})

    found_ids = {r.get("transaction_id") for r in rows}
    not_found = [tid for tid in clean_ids if tid not in found_ids]
    return resp(200, {"rows": rows, "count": len(rows), "not_found": not_found})


def run_wallet_search(body: dict):
    """Submit a wallet search by list of partner_account_ids (ej. 'CL-KNNW-8795')."""
    partner_account_ids = body.get("partner_account_ids", [])
    entity_type = body.get("entity_type", "b2c")
    if entity_type not in ("b2c", "b2b"):
        entity_type = "b2c"
    if not partner_account_ids:
        return resp(400, {"error": "partner_account_ids is required"})
    if len(partner_account_ids) > 5000:
        return resp(400, {"error": "Maximum 5000 partner_account_ids per search"})

    clean_ids = []
    for pid in partner_account_ids:
        pid = str(pid).strip()
        if pid:
            clean_ids.append(pid)
    if not clean_ids:
        return resp(400, {"error": "partner_account_ids is required"})

    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()
    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": "wallet_search",
        "status": "RUNNING",
        "params": json.dumps({"partner_account_ids": clean_ids, "n_ids": len(clean_ids), "entity_type": entity_type}),
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })

    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps({
            "report_name": "wallet_search",
            "partner_account_ids": clean_ids,
            "entity_type": entity_type,
            "run_id": run_id,
            "keep_session": False,
        }),
    )
    return resp(202, {"run_id": run_id, "status": "RUNNING", "n_ids": len(clean_ids)})


def run_individual_analysis(body: dict):
    """Submit an individual AML analysis for a list of customer_ids."""
    customer_ids = body.get("customer_ids", [])
    if not customer_ids:
        return resp(400, {"error": "customer_ids is required"})
    if len(customer_ids) > 1000:
        return resp(400, {"error": "Maximum 1000 customer_ids per analysis"})

    # Sanitize: accept integers or numeric strings
    clean_ids = []
    for cid in customer_ids:
        try:
            clean_ids.append(int(str(cid).strip()))
        except (ValueError, TypeError):
            return resp(400, {"error": f"Invalid customer_id: {cid!r}"})

    # Período configurable: 5/15/30/60/90 días, o None/omitido = histórico (sin límite)
    days = body.get("days")
    try:
        days = int(days)
        if days not in (5, 15, 30, 60, 90):
            days = None
    except (TypeError, ValueError):
        days = None

    # Tipo de entidad: 'b2c' (default, customer_v2) o 'b2b' (company.company).
    entity_type = "b2b" if str(body.get("entity_type", "b2c")).lower() == "b2b" else "b2c"

    run_id = str(uuid.uuid4())
    now = dt.datetime.utcnow().isoformat()
    user_email = str(body.get("user_email", "")).strip()[:200]
    runs_table.put_item(Item={
        "run_id": run_id,
        "report_name": "individual_aml_analysis",
        "status": "RUNNING",
        "params": json.dumps({"customer_ids": clean_ids, "n_customers": len(clean_ids), "days": days, "entity_type": entity_type}),
        "started_at": now,
        "user_email": user_email,
        "ttl": int((dt.datetime.utcnow() + dt.timedelta(days=90)).timestamp()),
    })

    lambda_client.invoke(
        FunctionName=REPORT_LAMBDA_NAME,
        InvocationType="Event",
        Payload=json.dumps({
            "report_name": "individual_aml_analysis",
            "customer_ids": clean_ids,
            "days": days,
            "entity_type": entity_type,
            "run_id": run_id,
            "keep_session": False,
        }),
    )
    return resp(202, {"run_id": run_id, "status": "RUNNING", "n_customers": len(clean_ids), "days": days, "entity_type": entity_type})


def get_dashboard_stats_result(q0: str, q1: str, q2: str):
    """Check status of 3 previously submitted statements; return results for done ones.

    Response keys:
      daily_evolution / over_300k / by_country → list[dict] if done, null if still running
      all_done → True when all 3 are finished (or failed)
    Each done/failed statement is fetched once and never polled again by the Lambda.
    """
    stmt_ids = [q0, q1, q2]
    keys = ["daily_evolution", "over_300k", "by_country"]
    result: dict = {}
    all_done = True

    for stmt_id, key in zip(stmt_ids, keys):
        if not stmt_id:
            result[key] = []
            continue
        try:
            desc = redshift_data.describe_statement(Id=stmt_id)
            status = desc["Status"]
            if status == "FINISHED":
                result[key] = _rs_get_rows(stmt_id) if desc.get("HasResultSet") else []
            elif status in ("FAILED", "ABORTED"):
                result[key] = []
            else:
                # SUBMITTED / PICKED / STARTED — still running
                all_done = False
                result[key] = None   # null signals "still pending" to the frontend
        except Exception:
            result[key] = []         # treat errors as done-empty

    result["all_done"] = all_done
    return resp(200, result)


# ---------------------------------------------------------------------------
# Analytics CRM — Phase 6
# ---------------------------------------------------------------------------

_SQL_CASES_BY_STATUS = """
SELECT status, COUNT(*) AS n
FROM crm.cases
GROUP BY status
ORDER BY n DESC
"""

_SQL_CASES_BY_WEEK = """
SELECT DATE_TRUNC('week', created_at)::DATE AS week_start, COUNT(*) AS n
FROM crm.cases
WHERE created_at >= DATEADD(week, -8, GETDATE())
GROUP BY 1
ORDER BY 1
"""

_SQL_ALERTS_BY_REPORT = """
SELECT report_name, COUNT(*) AS n
FROM compliance.alerts
WHERE created_at >= DATEADD(day, -90, GETDATE())
GROUP BY report_name
ORDER BY n DESC
LIMIT 10
"""

_SQL_ALERTS_DAILY_30D = """
SELECT created_at::DATE AS day, COUNT(*) AS n
FROM compliance.alerts
WHERE created_at >= DATEADD(day, -30, GETDATE())
GROUP BY 1
ORDER BY 1
"""

_SQL_TOP_ENTITIES = """
SELECT entity_value, entity_type, COUNT(*) AS n
FROM compliance.alerts
WHERE entity_value IS NOT NULL AND TRIM(entity_value) <> ''
GROUP BY entity_value, entity_type
ORDER BY n DESC
LIMIT 5
"""


def _parse_dt(s):
    """Parse a stored 'YYYY-MM-DD HH:MM:SS' (or ISO) string to datetime, or None."""
    if not s:
        return None
    txt = str(s)[:19]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return None


def get_analytics_summary():
    """S3-backed analytics → no Redshift needed. Returns placeholder ids so the
    frontend's 2-phase flow keeps working; real data comes from /analytics/result."""
    return resp(200, {"stmt_ids": ["s3", "s3", "s3", "s3", "s3"]})


def get_analytics_result(q0: str = "", q1: str = "", q2: str = "", q3: str = "", q4: str = ""):
    """Compute the 5 CRM analytics datasets from the S3 store."""
    from collections import Counter
    try:
        now = dt.datetime.utcnow()
        cases = _crm_list("cases")
        alerts = _crm_list("alerts")

        # cases_by_status
        st = Counter(c.get("status", "open") for c in cases)
        cases_by_status = sorted([{"status": k, "n": v} for k, v in st.items()],
                                 key=lambda x: x["n"], reverse=True)

        # cases_by_week (last 8 weeks, Monday-anchored)
        wk = Counter()
        cutoff_8w = now - dt.timedelta(weeks=8)
        for c in cases:
            d = _parse_dt(c.get("created_at"))
            if d and d >= cutoff_8w:
                wstart = (d - dt.timedelta(days=d.weekday())).strftime("%Y-%m-%d")
                wk[wstart] += 1
        cases_by_week = [{"week_start": k, "n": wk[k]} for k in sorted(wk)]

        # alerts_by_report (last 90d, top 10)
        cutoff_90 = now - dt.timedelta(days=90)
        rep = Counter()
        for a in alerts:
            d = _parse_dt(a.get("created_at"))
            if d and d >= cutoff_90:
                rep[a.get("report_name", "") or "—"] += 1
        alerts_by_report = [{"report_name": k, "n": v} for k, v in rep.most_common(10)]

        # alerts_daily_30d
        cutoff_30 = now - dt.timedelta(days=30)
        day = Counter()
        for a in alerts:
            d = _parse_dt(a.get("created_at"))
            if d and d >= cutoff_30:
                day[d.strftime("%Y-%m-%d")] += 1
        alerts_daily_30d = [{"day": k, "n": day[k]} for k in sorted(day)]

        # top_entities (top 5)
        ent = Counter()
        for a in alerts:
            ev = (a.get("entity_value", "") or "").strip()
            if ev:
                ent[(ev, a.get("entity_type", "") or a.get("entity_field", ""))] += 1
        top_entities = [{"entity_value": k[0], "entity_type": k[1], "n": v}
                        for k, v in ent.most_common(5)]

        return resp(200, {
            "cases_by_status": cases_by_status,
            "cases_by_week": cases_by_week,
            "alerts_by_report": alerts_by_report,
            "alerts_daily_30d": alerts_daily_30d,
            "top_entities": top_entities,
            "all_done": True,
        })
    except Exception as e:
        return resp(200, {"cases_by_status": [], "cases_by_week": [], "alerts_by_report": [],
                          "alerts_daily_30d": [], "top_entities": [], "all_done": True,
                          "warning": str(e)})


# ---------------------------------------------------------------------------
# Phase 8 — Email notifications
# ---------------------------------------------------------------------------

def _send_email(
    to: str, subject: str, html_body: str, from_addr: str | None = None,
    attachments: list[tuple[str, bytes]] | None = None,
) -> dict:
    """Manda un correo por SMTP de Gmail. NO lanza excepciones (para no tumbar
    la operación que lo llamó), pero SÍ devuelve qué pasó:

        {"sent": bool, "error": str | None}

    Antes se tragaba los errores en silencio y no quedaba rastro: si el correo
    no salía, nadie se enteraba. Ahora además queda registrado en CloudWatch.

    Siempre se autentica como GMAIL_USER (la cuenta que tiene la app password),
    pero `from_addr` permite salir como uno de sus alias "enviar como"
    (ej. compliance@global66.com) sin necesitar una app password propia.

    `attachments` es una lista opcional de (nombre, bytes).
    """
    if not to or not to.strip():
        return {"sent": False, "error": "destinatario vacío"}
    gmail_password = _get_gmail_password()
    if not gmail_password:
        msg_err = ("No hay app password de Gmail configurada (ni en Secrets Manager "
                   f"'{GMAIL_PASSWORD_SECRET_NAME}' ni en la variable GMAIL_APP_PASSWORD)")
        print(f"[email] NO ENVIADO a {to}: {msg_err}")
        return {"sent": False, "error": msg_err}

    import smtplib
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    sender = from_addr or GMAIL_USER

    body_part = MIMEMultipart("alternative")
    body_part.attach(MIMEText(html_body, "html"))

    if attachments:
        msg = MIMEMultipart("mixed")
        msg.attach(body_part)
        for filename, content in attachments:
            part = MIMEApplication(content, Name=filename)
            part["Content-Disposition"] = f'attachment; filename="{filename}"'
            msg.attach(part)
    else:
        msg = body_part

    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to
    try:
        # 8s era muy justo: el handshake TLS + login contra Gmail desde una
        # Lambda fría se pasaba del límite y el correo se perdía sin aviso.
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20) as server:
            server.login(GMAIL_USER, gmail_password)
            server.sendmail(sender, [to], msg.as_string())
        print(f"[email] enviado a {to} (asunto: {subject!r})")
        return {"sent": True, "error": None}
    except Exception as e:
        detalle = f"{type(e).__name__}: {e}"
        print(f"[email] FALLÓ el envío a {to}: {detalle}")
        return {"sent": False, "error": detalle}


def _case_assignment_email(to_email: str, case_id: str, title: str, priority: str, assigned_by: str) -> None:
    priority_color = {"critical": "#ef4444", "high": "#f97316", "medium": "#eab308", "low": "#22c55e"}.get(priority, "#94a3b8")
    html = f"""
<div style="font-family:system-ui,sans-serif;background:#0f172a;color:#e2e8f0;padding:24px;border-radius:12px;max-width:560px">
  <div style="background:#1e293b;border-radius:8px;padding:16px 20px;margin-bottom:16px">
    <p style="margin:0;font-size:13px;color:#94a3b8">WatchTower AML &middot; Global66 Compliance</p>
    <h2 style="margin:8px 0 0;font-size:18px;color:#fff">Caso asignado a ti</h2>
  </div>
  <p style="font-size:14px;color:#cbd5e1">Se te ha asignado el siguiente caso de investigaci&#xf3;n:</p>
  <div style="background:#1e293b;border-left:4px solid {priority_color};border-radius:4px;padding:14px 18px;margin:12px 0">
    <p style="margin:0 0 4px;font-size:12px;color:#64748b;font-family:monospace">{case_id}</p>
    <p style="margin:0;font-size:15px;font-weight:600;color:#f1f5f9">{title}</p>
    <span style="display:inline-block;margin-top:8px;background:{priority_color}22;color:{priority_color};border-radius:999px;padding:2px 10px;font-size:11px;font-weight:600;text-transform:uppercase">{priority}</span>
  </div>
  <p style="font-size:12px;color:#475569;margin-top:16px">Asignado por: <strong style="color:#94a3b8">{assigned_by}</strong></p>
</div>
"""
    _send_email(to_email, f"[WatchTower] Caso asignado: {title}", html)


# ---------------------------------------------------------------------------
# Phase 9 — SLA Analytics
# ---------------------------------------------------------------------------

_SQL_SLA_AVG_RESOLUTION = """
SELECT priority,
       COUNT(*) AS total_closed,
       AVG(DATEDIFF(hour, created_at, closed_at)) AS avg_hours
FROM crm.cases
WHERE status = 'closed' AND closed_at IS NOT NULL
GROUP BY priority
ORDER BY priority
"""

_SQL_SLA_OVERDUE = """
SELECT
  SUM(CASE WHEN priority='critical' AND created_at < DATEADD(day,-1,GETDATE())  THEN 1 ELSE 0 END) AS critical_overdue,
  SUM(CASE WHEN priority='high'     AND created_at < DATEADD(day,-3,GETDATE())  THEN 1 ELSE 0 END) AS high_overdue,
  SUM(CASE WHEN priority='medium'   AND created_at < DATEADD(day,-7,GETDATE())  THEN 1 ELSE 0 END) AS medium_overdue,
  SUM(CASE WHEN priority='low'      AND created_at < DATEADD(day,-30,GETDATE()) THEN 1 ELSE 0 END) AS low_overdue,
  COUNT(*) AS total_open
FROM crm.cases
WHERE status NOT IN ('closed')
"""

_SQL_SLA_BY_PRIORITY = """
SELECT priority, status, COUNT(*) AS n
FROM crm.cases
GROUP BY priority, status
ORDER BY
  CASE priority WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 ELSE 4 END,
  status
"""


def get_analytics_sla():
    """S3-backed → placeholder ids, real data from /analytics/sla/result."""
    return resp(200, {"stmt_ids": ["s3", "s3", "s3"]})


def get_analytics_sla_result(q0: str = "", q1: str = "", q2: str = ""):
    """Compute the 3 SLA datasets from the S3 cases store."""
    from collections import Counter
    try:
        now = dt.datetime.utcnow()
        cases = _crm_list("cases")

        # avg_resolution: por prioridad, horas promedio entre creación y cierre
        closed_by_pri: dict[str, list] = {}
        for c in cases:
            if c.get("status") == "closed" and c.get("closed_at"):
                cd, dd = _parse_dt(c.get("created_at")), _parse_dt(c.get("closed_at"))
                if cd and dd:
                    closed_by_pri.setdefault(c.get("priority", "medium"), []).append(
                        (dd - cd).total_seconds() / 3600.0)
        avg_resolution = []
        for pri in sorted(closed_by_pri):
            hrs = closed_by_pri[pri]
            avg_resolution.append({"priority": pri, "total_closed": len(hrs),
                                   "avg_hours": round(sum(hrs) / len(hrs), 1)})

        # overdue: casos abiertos que pasaron su SLA por prioridad
        sla_days = {"critical": 1, "high": 3, "medium": 7, "low": 30}
        overdue = {"critical_overdue": 0, "high_overdue": 0, "medium_overdue": 0,
                   "low_overdue": 0, "total_open": 0}
        for c in cases:
            if c.get("status") == "closed":
                continue
            overdue["total_open"] += 1
            pri = c.get("priority", "medium")
            d = _parse_dt(c.get("created_at"))
            if d and pri in sla_days and d < now - dt.timedelta(days=sla_days[pri]):
                overdue[f"{pri}_overdue"] += 1

        # by_priority: conteo por (prioridad, estado)
        bp = Counter((c.get("priority", "medium"), c.get("status", "open")) for c in cases)
        pri_rank = {"critical": 1, "high": 2, "medium": 3, "low": 4}
        by_priority = sorted(
            [{"priority": k[0], "status": k[1], "n": v} for k, v in bp.items()],
            key=lambda x: (pri_rank.get(x["priority"], 5), x["status"]))

        return resp(200, {"avg_resolution": avg_resolution, "overdue": [overdue],
                          "by_priority": by_priority, "all_done": True})
    except Exception as e:
        return resp(200, {"avg_resolution": [], "overdue": [], "by_priority": [],
                          "all_done": True, "warning": str(e)})


# ---------------------------------------------------------------------------
# Phase 7 — User Management
# ---------------------------------------------------------------------------

# Roles are a small fixed set (CRM has no role-editing UI).
ROLES = [
    {"id": 1, "name": "analyst", "description": "Analista AML"},
    {"id": 2, "name": "supervisor", "description": "Supervisor de Compliance"},
    {"id": 3, "name": "admin", "description": "Administrador"},
]
_ROLE_BY_ID = {r["id"]: r["name"] for r in ROLES}
_ROLE_BY_NAME = {r["name"]: r["id"] for r in ROLES}


def get_users():
    # S3-backed. Uses email as the stable id (route /users/{id}).
    try:
        users = []
        for u in _crm_list("users"):
            role_name = u.get("role_name", "analyst")
            users.append({
                "id": u.get("email", ""),
                "email": u.get("email", ""),
                "full_name": u.get("full_name") or u.get("email", ""),
                "is_active": bool(u.get("is_active", True)),
                "created_at": u.get("created_at", ""),
                "last_login_at": u.get("last_login_at", ""),
                "role_name": role_name,
                "role_id": _ROLE_BY_NAME.get(role_name, 1),
            })
        users.sort(key=lambda x: x["created_at"], reverse=True)
        return resp(200, {"users": users})
    except Exception as e:
        return resp(200, {"users": [], "warning": str(e)})


def get_roles():
    return resp(200, {"roles": ROLES})


def create_user(body: dict):
    email = str(body.get("email", "")).strip().lower()[:255]
    full_name = str(body.get("full_name", "")).strip()[:255]
    role_id = int(body.get("role_id", 1))
    if not email:
        return resp(400, {"error": "email is required"})
    _crm_put("users", email, {
        "email": email,
        "full_name": full_name or email,
        "is_active": True,
        "role_name": _ROLE_BY_ID.get(role_id, "analyst"),
        "created_at": _now_str(),
        "last_login_at": "",
    })
    _safe_audit(user_email="admin", action="create_user", entity_type="user", entity_id=email)
    return resp(201, {"message": "Usuario creado", "email": email})


def update_user(user_id: str, body: dict):
    # user_id is the email (what get_users returns as id).
    changes = {}
    if "full_name" in body:
        changes["full_name"] = str(body["full_name"]).strip()
    if "role_id" in body:
        changes["role_name"] = _ROLE_BY_ID.get(int(body["role_id"]), "analyst")
    if "is_active" in body:
        changes["is_active"] = bool(body["is_active"])
    if not changes:
        return resp(400, {"error": "nothing to update"})
    if _crm_update("users", user_id, changes) is None:
        return resp(404, {"error": f"User '{user_id}' not found"})
    _safe_audit(user_email="admin", action="update_user", entity_type="user", entity_id=user_id)
    return resp(200, {"message": "Usuario actualizado"})


def deactivate_user(user_id: str):
    if _crm_update("users", user_id, {"is_active": False}) is None:
        return resp(404, {"error": f"User '{user_id}' not found"})
    _safe_audit(user_email="admin", action="deactivate_user", entity_type="user", entity_id=user_id)
    return resp(200, {"message": "Usuario desactivado"})


# ---------------------------------------------------------------------------
# Phase 10 — Auto-case Rules (stored as JSON in S3)
# ---------------------------------------------------------------------------

def _load_rules() -> list[dict]:
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=AUTO_RULES_KEY)
        return json.loads(obj["Body"].read())
    except Exception:
        return []


def _save_rules(rules: list[dict]) -> None:
    s3.put_object(
        Bucket=S3_BUCKET,
        Key=AUTO_RULES_KEY,
        Body=json.dumps(rules, default=str).encode(),
        ContentType="application/json",
    )


def get_rules():
    return resp(200, {"rules": _load_rules()})


def create_rule(body: dict):
    rules = _load_rules()
    rule = {
        "id": str(uuid.uuid4()),
        "name": str(body.get("name", "")).strip()[:100],
        "report_name": str(body.get("report_name", "")).strip()[:100],
        "row_threshold": int(body.get("row_threshold", 1)),
        "field_name": str(body.get("field_name", "")).strip()[:100],
        "field_value": float(body["field_value"]) if str(body.get("field_value", "")).strip() != "" else None,
        "case_title_template": str(body.get("case_title_template", "Alerta automática: {report_name}")).strip()[:255],
        "priority": str(body.get("priority", "medium")),
        "assigned_to": str(body.get("assigned_to", "")).strip()[:255],
        "enabled": bool(body.get("enabled", True)),
        "created_at": dt.datetime.utcnow().isoformat(),
    }
    rules.append(rule)
    _save_rules(rules)
    return resp(201, {"message": "Regla creada", "rule": rule})


def update_rule(rule_id: str, body: dict):
    rules = _load_rules()
    for rule in rules:
        if rule["id"] == rule_id:
            for field in ["name", "report_name", "row_threshold", "field_name", "case_title_template", "priority", "assigned_to", "enabled"]:
                if field in body:
                    rule[field] = body[field]
            if "field_value" in body:
                rule["field_value"] = float(body["field_value"]) if str(body["field_value"]).strip() != "" else None
            _save_rules(rules)
            return resp(200, {"message": "Regla actualizada", "rule": rule})
    return resp(404, {"error": "Rule not found"})


def delete_rule(rule_id: str):
    rules = _load_rules()
    new_rules = [r for r in rules if r["id"] != rule_id]
    if len(new_rules) == len(rules):
        return resp(404, {"error": "Rule not found"})
    _save_rules(new_rules)
    return resp(200, {"message": "Regla eliminada"})


def _create_auto_case(report_name: str, rule: dict, title: str, description: str,
                       entity_type: str = "report", entity_id: str = "") -> None:
    case_id = str(uuid.uuid4())
    now = _now_str()
    assigned = str(rule.get("assigned_to", "")).strip()
    priority = rule.get("priority", "medium")
    # Casos viven en S3 (always-on) → escribir ahí, no en Redshift.
    _crm_put("cases", case_id, {
        "case_id": case_id,
        "title": title,
        "description": description,
        "status": "open",
        "priority": priority,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "report_name": report_name,
        "assigned_to": assigned,
        "created_by": "sistema@auto",
        "created_at": now,
        "updated_at": now,
        "closed_at": "",
        "notes": [],
    })
    if assigned:
        _case_assignment_email(assigned, case_id, title, priority, "sistema automático")


class _SafeFormatDict(dict):
    """.format_map() dict that leaves unknown {placeholders} as literal text
    instead of raising KeyError (case_title_template is user-authored)."""
    def __missing__(self, key):
        return "{" + key + "}"


def apply_auto_case_rules(report_name: str, rows: list[dict], run_id: str) -> None:
    """Called after a report completes — creates cases for matching enabled rules.

    Two modes, chosen per rule:
    - field_name/field_value set  → condición por fila: crea UN caso por cada
      fila donde float(row[field_name]) > field_value (ej. smurfing con
      small_payins_7d > 1000). No envía solicitud de documentos, solo abre
      el caso y notifica al analista asignado.
    - field_name vacío (legado)   → condición por reporte completo: crea UN
      solo caso si el total de filas del reporte >= row_threshold.
    """
    row_count = len(rows)
    try:
        rules = _load_rules()
        for rule in rules:
            if not rule.get("enabled", True):
                continue
            if rule.get("report_name") and rule["report_name"] != report_name:
                continue
            rule_name = str(rule.get("name", ""))
            field_name = str(rule.get("field_name", "")).strip()
            field_value = rule.get("field_value")

            if field_name and field_value is not None:
                for row in rows:
                    raw = row.get(field_name)
                    if raw is None:
                        continue
                    try:
                        val = float(raw)
                    except (TypeError, ValueError):
                        continue
                    if val <= float(field_value):
                        continue
                    entity_type = "customer" if "customer_id" in row else "company" if "company_id" in row else "report"
                    entity_id = str(row.get("customer_id") or row.get("company_id") or "")
                    ctx = _SafeFormatDict(row)
                    ctx.update(report_name=report_name, row_count=row_count, run_id=run_id)
                    title = rule.get("case_title_template", "Alerta automática: {report_name}").format_map(ctx)
                    description = (f'Creado automáticamente por regla "{rule_name}" — '
                                    f"{report_name}: {field_name}={raw} (run {run_id}).")
                    _create_auto_case(report_name, rule, title, description, entity_type, entity_id)
            else:
                ctx = _SafeFormatDict(report_name=report_name, row_count=row_count, run_id=run_id)
                if row_count < int(rule.get("row_threshold", 1)):
                    continue
                title = rule.get("case_title_template", "Alerta automática: {report_name}").format_map(ctx)
                description = (f'Creado automáticamente por regla "{rule_name}" — '
                                f"{report_name} con {row_count} filas (run {run_id}).")
                _create_auto_case(report_name, rule, title, description, "report", "")
    except Exception:
        pass
