"""
Compliance Redshift Reports — main Lambda handler.

Flow:
  1. Resume cluster if paused, wait until available.
  2. Execute the SQL for the requested report via Redshift Data API.
  3. Fetch results and build Excel + HTML summary.
  4. Upload Excel to S3 (encrypted).
  5. Send SES email with attachment + presigned link (best-effort).
  6. POST summary to Slack webhook.
  7. Pause cluster.

Supported reports (pass via event["report_name"] or REPORT_NAME env var):
  - high_risk_countries        : AML screening — outbound tx to FATF/OFAC countries
  - amount_ranges_by_country   : Volume distribution by USD range × destination country (7d)
  - top_customers_by_range_country : Top-15 customers per range × country (7d)

Environment variables expected (set by Terraform):
  CLUSTER_IDENTIFIER       — Redshift cluster identifier
  DATABASE_NAME            — DB name (e.g. dev)
  DB_USER                  — DB user — uses IAM auth via GetClusterCredentials
  S3_BUCKET                — output bucket
  SES_FROM_ADDRESS         — verified SES sender
  SES_TO_ADDRESSES         — comma-separated recipients
  SLACK_WEBHOOK_SECRET_ARN — Secrets Manager ARN holding the Slack webhook URL
  REPORT_NAME              — default report if not passed in event
  AUTO_PAUSE               — "true" to pause cluster after run
  RUNS_TABLE               — DynamoDB table name for run history (optional)
  CATALOG_TABLE            — DynamoDB table name for custom query catalog (optional)
"""

from __future__ import annotations

import base64
import datetime as dt
import io
import json
import logging
import os
import re
import time
import urllib.request
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import boto3
import yaml
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from aml_individual import build_aml_excel

try:
    from api_handler import apply_auto_case_rules as _apply_auto_case_rules
except Exception:
    def _apply_auto_case_rules(*args, **kwargs):  # noqa: ANN001
        pass

try:
    from api_handler import poll_document_replies as _poll_document_replies
except Exception:
    def _poll_document_replies(*args, **kwargs):  # noqa: ANN001
        return {"status": "error", "error": "poll_document_replies not importable"}

try:
    from api_handler import maybe_trigger_auto_document_requests as _trigger_auto_document_requests
except Exception:
    def _trigger_auto_document_requests(*args, **kwargs):  # noqa: ANN001
        return {"triggered": 0, "error": "not importable"}

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# ---------------------------------------------------------------------------
# AWS clients (created once per Lambda container)
# ---------------------------------------------------------------------------
redshift = boto3.client("redshift")
redshift_data = boto3.client("redshift-data")
s3 = boto3.client("s3")
ses = boto3.client("ses")
secrets = boto3.client("secretsmanager")
dynamodb = boto3.resource("dynamodb")

# ---------------------------------------------------------------------------
# Config from environment
# ---------------------------------------------------------------------------
CLUSTER_ID = os.environ["CLUSTER_IDENTIFIER"]
DATABASE = os.environ["DATABASE_NAME"]
DB_USER = os.environ["DB_USER"]
S3_BUCKET = os.environ["S3_BUCKET"]
SES_FROM = os.environ["SES_FROM_ADDRESS"]
SES_TO = [e.strip() for e in os.environ["SES_TO_ADDRESSES"].split(",") if e.strip()]
SLACK_SECRET_ARN = os.environ.get("SLACK_WEBHOOK_SECRET_ARN", "")
REPORT_NAME = os.environ.get("REPORT_NAME", "high_risk_countries")
AUTO_PAUSE = os.environ.get("AUTO_PAUSE", "true").lower() == "true"
RUNS_TABLE_NAME = os.environ.get("RUNS_TABLE", "")
CATALOG_TABLE_NAME = os.environ.get("CATALOG_TABLE", "")

BASE_DIR = Path(__file__).parent
QUERIES_DIR = BASE_DIR / "queries"
CONFIG_DIR = BASE_DIR / "config"
TEMPLATES_DIR = BASE_DIR

POLL_INTERVAL_SECONDS = 5
MAX_WAIT_RESUME_SECONDS = 780   # 13 min (Lambda timeout is 900s)
MAX_WAIT_QUERY_SECONDS = 540    # 9 min

# ---------------------------------------------------------------------------
# Report registry
# Add new reports here — no other changes needed for simple cases.
# ---------------------------------------------------------------------------
REPORT_CONFIGS: dict[str, dict] = {
    "priority_queue_test_alerts": {
        "display_name": "Priorización de Alertas — Datos de Prueba",
        "sql_file": "priority_queue_test_alerts.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "high_risk_countries": {
        "display_name": "High-Risk Countries Transactions",
        "sql_file": "high_risk_countries_transactions.sql",
        "needs_country_filter": True,   # substitutes {country_codes} and {only_successful}
        "needs_since_date": True,        # passes :since_date as Data API parameter
    },
    "amount_ranges_by_country": {
        "display_name": "Amount Ranges by Country (7d)",
        "sql_file": "amount_ranges_by_country.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "top_customers_by_range_country": {
        "display_name": "Top Customers by Range & Country (7d)",
        "sql_file": "top_customers_by_range_country.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "tax_haven_transactions": {
        "display_name": "Transacciones a Régimen Fiscal Preferencial (90d)",
        "sql_file": "tax_haven_transactions.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "tax_haven_funding": {
        "display_name": "Fondeos desde Régimen Fiscal Preferencial (7d)",
        "sql_file": "tax_haven_funding.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "payin_payout_accumulation": {
        "display_name": "Acumulación Pay In → Pay Out (7d)",
        "sql_file": "payin_payout_accumulation.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "small_payin_structuring": {
        "display_name": "Pay In Pequeños → Pay Out (Smurfing, 7d)",
        "sql_file": "small_payin_structuring.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "velocity_payin_payout": {
        "display_name": "Velocity Pay In ↔ Pay Out < 24h (7d)",
        "sql_file": "velocity_payin_payout.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "external_funder_single": {
        "display_name": "Tercero que Fondea Una Sola Cuenta (7d)",
        "sql_file": "external_funder_single.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "external_funder_multiple": {
        "display_name": "Tercero que Fondea Múltiples Cuentas (7d)",
        "sql_file": "external_funder_multiple.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "circular_transactions": {
        "display_name": "Circularidad DNI Cliente ↔ Beneficiario (90d)",
        "sql_file": "circular_transactions.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "structuring_detection": {
        "display_name": "Estructuración / Fraccionamiento (7d)",
        "sql_file": "structuring_detection.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "shared_beneficiary": {
        "display_name": "Beneficiario Compartido por Múltiples Remitentes (7d)",
        "sql_file": "shared_beneficiary.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "customer_metrics_7d": {
        "display_name": "Métricas por Cliente B2C (7d)",
        "sql_file": "customer_metrics_7d.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "beneficiary_concentration": {
        "display_name": "Concentración de Beneficiarios (7d)",
        "sql_file": "beneficiary_concentration.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "beneficiary_dispersion": {
        "display_name": "Dispersión de Beneficiarios (7d)",
        "sql_file": "beneficiary_dispersion.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "outbound_bank_change": {
        "display_name": "Cambio de Banco Outbound (30d vs 7d)",
        "sql_file": "outbound_bank_change.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "new_corridor_detection": {
        "display_name": "Corredor Nuevo para el Cliente (7d vs 90d)",
        "sql_file": "new_corridor_detection.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "high_volume_vs_historical": {
        "display_name": "Alto Volumen vs Histórico (7d vs 90d)",
        "sql_file": "high_volume_vs_historical.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "swift_mismatch_detection": {
        "display_name": "Mismatch SWIFT vs País Beneficiario (30d)",
        "sql_file": "swift_mismatch_detection.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "jumio_kyc_approval_rates": {
        "display_name": "Tasas de Aprobación / Rechazo KYC por Flujo",
        "sql_file": "jumio_kyc_approval_rates.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "jumio_duplicate_flows": {
        "display_name": "Documentos Jumio Duplicados / Flujos Múltiples",
        "sql_file": "jumio_duplicate_flows.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "b2c_as_legal_rep": {
        "display_name": "Clientes B2C como Representantes Legales",
        "sql_file": "b2c_as_legal_rep.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "top_companies_by_legal_reps": {
        "display_name": "Top 15 Empresas con Más Representantes Legales",
        "sql_file": "top_companies_by_legal_reps.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "age_anomaly_customers": {
        "display_name": "Clientes con Anomalía de Edad (<18 o >90 años)",
        "sql_file": "age_anomaly_customers.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "crypto_bridge_transactions": {
        "display_name": "Transacciones Bridge/Crypto (30d)",
        "sql_file": "crypto_bridge_transactions.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "crypto_bridge_cash_calls": {
        "display_name": "Cash Calls Bridge/Crypto (30d)",
        "sql_file": "crypto_bridge_cash_calls.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "crypto_high_risk_destinations": {
        "display_name": "Crypto hacia Países de Riesgo (30d)",
        "sql_file": "crypto_high_risk_destinations.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
    "crypto_full_bridge_activity": {
        "display_name": "Actividad Completa Bridge (30d)",
        "sql_file": "crypto_full_bridge_activity.sql",
        "needs_country_filter": False,
        "needs_since_date": False,
    },
}


# ---------------------------------------------------------------------------
# Cluster control
# ---------------------------------------------------------------------------
def get_cluster_status() -> str:
    resp = redshift.describe_clusters(ClusterIdentifier=CLUSTER_ID)
    return resp["Clusters"][0]["ClusterStatus"]


def _try_resume_cluster() -> None:
    """Issue resume_cluster, swallowing the error if it's already resuming/available
    (e.g. a concurrent run beat us to it) or busy with another operation."""
    try:
        logger.info("Resuming cluster %s", CLUSTER_ID)
        redshift.resume_cluster(ClusterIdentifier=CLUSTER_ID)
    except redshift.exceptions.InvalidClusterStateFault as e:
        logger.info("resume_cluster no-op (cluster already transitioning): %s", e)
    except Exception as e:  # noqa: BLE001
        logger.warning("resume_cluster call failed (will keep polling/retrying): %s", e)


def ensure_cluster_available() -> None:
    status = get_cluster_status()
    logger.info("Cluster %s status: %s", CLUSTER_ID, status)

    if status == "available":
        return

    if status == "paused":
        _try_resume_cluster()

    # Re-issue resume_cluster periodically while stuck on "paused" — a resume call
    # can silently no-op if it raced with another run's pause_cluster() finishing.
    # Without a retry, a stuck "paused" status never recovers on its own.
    deadline = time.time() + MAX_WAIT_RESUME_SECONDS
    last_resume_retry = time.time()
    RESUME_RETRY_INTERVAL = 30

    while time.time() < deadline:
        status = get_cluster_status()
        if status == "available":
            logger.info("Cluster is available")
            return
        if status == "paused" and time.time() - last_resume_retry >= RESUME_RETRY_INTERVAL:
            logger.warning("Still paused %ds after resume attempt — retrying", RESUME_RETRY_INTERVAL)
            _try_resume_cluster()
            last_resume_retry = time.time()
        logger.info("Waiting for cluster... (current: %s)", status)
        time.sleep(POLL_INTERVAL_SECONDS)

    raise TimeoutError(f"Cluster did not become available within {MAX_WAIT_RESUME_SECONDS}s")


def pause_cluster() -> None:
    """Pause the cluster, retrying if Redshift still has internal operations running."""
    MAX_PAUSE_ATTEMPTS = 6
    PAUSE_RETRY_WAIT = 20  # seconds between retries

    for attempt in range(1, MAX_PAUSE_ATTEMPTS + 1):
        try:
            status = get_cluster_status()
            if status != "available":
                logger.info("Cluster status is %s, skipping pause", status)
                return
            logger.info("Pausing cluster %s (attempt %d/%d)", CLUSTER_ID, attempt, MAX_PAUSE_ATTEMPTS)
            redshift.pause_cluster(ClusterIdentifier=CLUSTER_ID)
            logger.info("Cluster pause initiated successfully")
            return
        except redshift.exceptions.InvalidClusterStateFault:
            if attempt < MAX_PAUSE_ATTEMPTS:
                logger.info("Cluster busy, retrying pause in %ds...", PAUSE_RETRY_WAIT)
                time.sleep(PAUSE_RETRY_WAIT)
            else:
                logger.warning("Could not pause cluster after %d attempts — pause it manually", MAX_PAUSE_ATTEMPTS)
        except Exception as e:  # noqa: BLE001
            logger.exception("Failed to pause cluster: %s", e)
            return


# ---------------------------------------------------------------------------
# Whitelist support
# ---------------------------------------------------------------------------
def fetch_active_whitelist(report_name: str = "") -> list[dict]:
    """Fetch non-expired whitelist entries from the S3 JSON store (not Redshift),
    so it stays in sync with the CRM and never depends on cluster state.
    Returns global + report-specific entries."""
    try:
        import json as _json
        from concurrent.futures import ThreadPoolExecutor

        now = int(time.time())
        prefix = "crm/whitelist/"
        keys: list[str] = []
        paginator = s3.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix):
            for o in page.get("Contents", []):
                if o["Key"].endswith(".json"):
                    keys.append(o["Key"])
        if not keys:
            return []

        def _fetch(k):
            try:
                return _json.loads(s3.get_object(Bucket=S3_BUCKET, Key=k)["Body"].read())
            except Exception:
                return None

        with ThreadPoolExecutor(max_workers=16) as ex:
            items = [i for i in ex.map(_fetch, keys) if i is not None]

        rows: list[dict] = []
        for i in items:
            exp = int(i.get("expires_at", 0))
            if exp and exp <= now:
                continue  # vencida
            scope = i.get("scope", "global")
            rn = i.get("report_name", "")
            if scope == "global" or rn == report_name:
                rows.append({
                    "entity_field": i.get("entity_field"),
                    "entity_value": i.get("entity_value"),
                    "scope": scope,
                    "report_name": rn,
                })
        return rows

    except Exception as e:  # noqa: BLE001
        logger.warning("fetch_active_whitelist failed (non-blocking): %s", e)
        return []


def _final_select_columns(sql: str) -> set:
    """Return the set of column names exposed by the outermost SELECT.

    Walks the SQL character by character tracking parenthesis depth so that
    columns inside CTEs / subqueries are ignored.  For each top-level SELECT
    item we take the explicit AS alias when present, otherwise the bare
    identifier at the end of the expression (e.g. 't.customer_id' → 'customer_id').
    """
    sql_up = sql.upper()
    # Find the position of the LAST top-level SELECT (depth == 0)
    depth = 0
    final_pos = -1
    i = 0
    while i < len(sql):
        c = sql[i]
        if c == '(':
            depth += 1
        elif c == ')':
            depth -= 1
        elif depth == 0 and sql_up[i:i + 6] == 'SELECT':
            final_pos = i
        i += 1
    if final_pos == -1:
        return set()

    after_select = sql[final_pos + 6:]
    sql_up2 = after_select.upper()

    # Find the FROM that terminates the SELECT list (at depth 0)
    depth = 0
    from_pos = len(after_select)
    i = 0
    while i < len(after_select):
        c = after_select[i]
        if c == '(':
            depth += 1
        elif c == ')':
            depth -= 1
        elif depth == 0 and sql_up2[i:i + 4] == 'FROM':
            from_pos = i
            break
        i += 1

    select_clause = after_select[:from_pos]

    # Split by commas at depth 0
    items = []
    buf: list[str] = []
    depth = 0
    for c in select_clause:
        if c == '(':
            depth += 1
            buf.append(c)
        elif c == ')':
            depth -= 1
            buf.append(c)
        elif c == ',' and depth == 0:
            items.append(''.join(buf).strip())
            buf = []
        else:
            buf.append(c)
    if buf:
        items.append(''.join(buf).strip())

    columns: set = set()
    for item in items:
        item = item.strip()
        if not item or item == '*':
            continue
        # Explicit alias:  expr AS alias
        m = re.search(r'\bAS\s+(\w+)\s*$', item, re.IGNORECASE)
        if m:
            columns.add(m.group(1).lower())
        else:
            # Bare identifier or table.column — take the last word
            m2 = re.search(r'(\w+)\s*$', item)
            if m2:
                word = m2.group(1).lower()
                if word not in ('asc', 'desc', 'distinct'):
                    columns.add(word)
    return columns


def inject_whitelist_exclusions(sql: str, whitelist_entries: list[dict]) -> str:
    """Wrap SQL in a subquery that excludes whitelisted entities.

    Only injects a WHERE condition for fields that are actually present in
    the outermost SELECT output (i.e. available in _wt_base after wrapping).
    Fields that only appear inside CTEs or aggregate functions are skipped.
    """
    if not whitelist_entries:
        return sql
    # Group exclusions by field
    by_field: dict[str, list[str]] = {}
    for entry in whitelist_entries:
        field = entry.get("entity_field", "").strip()
        value = str(entry.get("entity_value", "")).strip()
        if field and value:
            by_field.setdefault(field, []).append(value)
    if not by_field:
        return sql

    available = _final_select_columns(sql)
    logger.info("Whitelist: final SELECT columns detected: %s", sorted(available))

    conditions = []
    for field, values in by_field.items():
        if field.lower() not in available:
            logger.info("Whitelist: skipping field '%s' — not in final SELECT output", field)
            continue
        quoted = ", ".join(f"'{v.replace(chr(39), chr(39)+chr(39))}'" for v in values)
        conditions.append(f"CAST({field} AS VARCHAR) NOT IN ({quoted})")
    if not conditions:
        return sql
    where_clause = " AND ".join(conditions)
    wrapped = f"SELECT * FROM (\n{sql}\n) _wt_base\nWHERE {where_clause}"
    logger.info("Whitelist injection: excluded %d entries across %d fields", len(whitelist_entries), len(by_field))
    return wrapped


# ---------------------------------------------------------------------------
# Query rendering + execution
# ---------------------------------------------------------------------------
def load_country_codes() -> list[dict]:
    with open(CONFIG_DIR / "high_risk_countries.yaml", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data["countries"]


def render_query(
    report_name: str,
    since_date: str,
    only_successful: bool,
    country_codes: list[str],
) -> tuple[str, list[dict]]:
    """Return (sql_string, data_api_params_list) for the given report.

    For built-in reports, reads from the queries/ directory.
    For custom reports, reads SQL from the DynamoDB catalog.
    """
    api_params: list[dict] = []

    if report_name in REPORT_CONFIGS:
        config = REPORT_CONFIGS[report_name]
        sql = (QUERIES_DIR / config["sql_file"]).read_text(encoding="utf-8")
        # Strip semicolons BEFORE whitelist injection wraps the SQL in a subquery
        sql = sql.strip().rstrip(";").strip()

        if config["needs_country_filter"]:
            quoted = ",".join(f"'{c}'" for c in country_codes)
            sql = sql.replace("{country_codes}", quoted)
            sql = sql.replace("{only_successful}", "TRUE" if only_successful else "FALSE")

        if config["needs_since_date"]:
            api_params.append({"name": "since_date", "value": since_date})
    else:
        # Custom query from DynamoDB catalog
        sql = _load_custom_sql(report_name)
        if not sql:
            raise ValueError(f"Report '{report_name}' not found in built-in registry or catalog")
        # Strip semicolons from custom queries too
        sql = sql.strip().rstrip(";").strip()

    # Inject whitelist exclusions
    whitelist_entries = fetch_active_whitelist(report_name)
    if whitelist_entries:
        sql = inject_whitelist_exclusions(sql, whitelist_entries)

    return sql, api_params


def execute_query(sql: str, api_params: list[dict] | None = None) -> list[dict]:
    logger.info("Submitting query to Redshift Data API")
    # Redshift Data API rejects trailing semicolons
    sql = sql.strip().rstrip(";").strip()
    kwargs: dict = dict(
        ClusterIdentifier=CLUSTER_ID,
        Database=DATABASE,
        DbUser=DB_USER,
        Sql=sql,
        WithEvent=False,
    )
    if api_params:
        kwargs["Parameters"] = api_params

    resp = redshift_data.execute_statement(**kwargs)
    statement_id = resp["Id"]
    logger.info("Statement id: %s", statement_id)

    deadline = time.time() + MAX_WAIT_QUERY_SECONDS
    while time.time() < deadline:
        desc = redshift_data.describe_statement(Id=statement_id)
        status = desc["Status"]
        if status == "FINISHED":
            break
        if status in ("FAILED", "ABORTED"):
            raise RuntimeError(f"Query {status}: {desc.get('Error', 'unknown error')}")
        time.sleep(POLL_INTERVAL_SECONDS)
    else:
        raise TimeoutError(f"Query did not finish within {MAX_WAIT_QUERY_SECONDS}s")

    return fetch_results(statement_id)


def enrich_rows_with_priority(rows: list[dict]) -> list[dict]:
    """Le pega 'prioridad' y 'risk_score' a cada fila que tenga customer_id o
    company_id, consultando compliance.priority_queue_b2c/b2b (una sola
    consulta en batch, no una por fila). Reportes agregados/resumen que no
    traen esas columnas quedan intactos — no todos los 29 reportes son
    listas por cliente, algunos son resúmenes por corredor/método de pago.

    Score PLACEHOLDER (promedio simple) — ver compliance.priority_queue_b2c/
    b2b para el detalle; se actualiza solo cuando lleguen los pesos reales.
    """
    if not rows:
        return rows

    sample = rows[0]
    if "prioridad" in sample:
        return rows  # ya viene con prioridad propia (ej. datos de prueba) — no pisar
    if "customer_id" in sample:
        id_col, view = "customer_id", "compliance.priority_queue_b2c"
    elif "company_id" in sample:
        id_col, view = "company_id", "compliance.priority_queue_b2b"
    else:
        return rows

    ids = {r.get(id_col) for r in rows if r.get(id_col) is not None}
    if not ids:
        return rows

    try:
        ids_sql = ", ".join(str(int(i)) for i in ids)
        priority_rows = execute_query(
            f"SELECT {id_col}, risk_score FROM {view} WHERE {id_col} IN ({ids_sql})"
        )
    except Exception:
        logger.exception("No se pudo enriquecer con prioridad (no bloquea el reporte)")
        return rows

    score_by_id = {str(pr[id_col]): pr.get("risk_score") for pr in priority_rows}
    for r in rows:
        score = score_by_id.get(str(r.get(id_col)))
        r["risk_score"] = score
        if score is None:
            r["prioridad"] = None
        else:
            score = float(score)
            r["prioridad"] = "P1" if score >= 75 else "P2" if score >= 50 else "P3"
    return rows


def fetch_results(statement_id: str) -> list[dict]:
    rows: list[dict] = []
    columns: list[str] = []
    next_token: str | None = None

    while True:
        kwargs = {"Id": statement_id}
        if next_token:
            kwargs["NextToken"] = next_token
        resp = redshift_data.get_statement_result(**kwargs)

        if not columns:
            columns = [c["name"] for c in resp["ColumnMetadata"]]

        for record in resp["Records"]:
            rows.append({columns[i]: _unwrap_value(cell) for i, cell in enumerate(record)})

        next_token = resp.get("NextToken")
        if not next_token:
            break

    logger.info("Fetched %d rows", len(rows))
    return rows


def _unwrap_value(cell: dict):
    """Redshift Data API returns each cell as a single-key dict like {'stringValue': 'x'}."""
    if "isNull" in cell and cell["isNull"]:
        return None
    for key in ("stringValue", "longValue", "doubleValue", "booleanValue", "blobValue"):
        if key in cell:
            return cell[key]
    return None


# ---------------------------------------------------------------------------
# AI summary — aggregates over the FULL result set (not just the preview), so
# the in-browser "Analizar con IA" reasons over all rows, not the first 10.
# ---------------------------------------------------------------------------
def _build_ai_summary(rows: list, top_n: int = 30) -> dict:
    if not rows:
        return {"total_rows": 0, "columns": [], "numeric_stats": {}, "top_rows": []}
    columns = list(rows[0].keys())
    numeric_stats: dict = {}
    for col in columns:
        vals = []
        for r in rows:
            try:
                vals.append(float(r.get(col)))
            except (TypeError, ValueError):
                pass
        if vals and len(vals) >= len(rows) / 2:   # columna mayormente numérica
            numeric_stats[col] = {
                "sum": round(sum(vals), 2),
                "avg": round(sum(vals) / len(vals), 2),
                "min": round(min(vals), 2),
                "max": round(max(vals), 2),
            }
    key_col = None
    if numeric_stats:
        key_col = max(numeric_stats, key=lambda c: numeric_stats[c]["max"])

        def _salience(r):
            try:
                return float(r.get(key_col))
            except (TypeError, ValueError):
                return float("-inf")
        top_rows = sorted(rows, key=_salience, reverse=True)[:top_n]
    else:
        top_rows = rows[:top_n]
    return {
        "total_rows": len(rows),
        "columns": columns,
        "numeric_stats": numeric_stats,
        "top_rows_metric": key_col,
        "top_rows": top_rows,
    }


# ---------------------------------------------------------------------------
# DynamoDB run tracking  (no-ops when RUNS_TABLE_NAME is empty)
# ---------------------------------------------------------------------------
def _update_run(run_id: str | None, **attrs) -> None:
    """Best-effort DynamoDB update — failures are logged but never propagate."""
    if not run_id or not RUNS_TABLE_NAME:
        return
    try:
        table = dynamodb.Table(RUNS_TABLE_NAME)
        update_expr = "SET " + ", ".join(f"#{k} = :{k}" for k in attrs)
        expr_names = {f"#{k}": k for k in attrs}
        expr_values = {f":{k}": v for k, v in attrs.items()}
        table.update_item(
            Key={"run_id": run_id},
            UpdateExpression=update_expr,
            ExpressionAttributeNames=expr_names,
            ExpressionAttributeValues=expr_values,
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("DynamoDB update_run failed (non-blocking): %s", e)


def _load_custom_sql(report_name: str) -> str | None:
    """Fetch SQL from the DynamoDB catalog for custom queries. Returns None if not found."""
    if not CATALOG_TABLE_NAME:
        return None
    try:
        table = dynamodb.Table(CATALOG_TABLE_NAME)
        item = table.get_item(Key={"report_name": report_name}).get("Item")
        return item.get("sql") if item else None
    except Exception as e:  # noqa: BLE001
        logger.warning("DynamoDB catalog lookup failed: %s", e)
        return None


# ---------------------------------------------------------------------------
# Report building
# ---------------------------------------------------------------------------
def build_summary(rows: list[dict], report_name: str) -> dict:
    """Dispatch to report-specific summary builder."""
    if report_name == "high_risk_countries":
        return _summary_high_risk(rows)
    if report_name == "amount_ranges_by_country":
        return _summary_amount_ranges(rows)
    if report_name == "top_customers_by_range_country":
        return _summary_top_customers(rows)
    return _summary_generic(rows)


def _summary_high_risk(rows: list[dict]) -> dict:
    total = len(rows)
    by_country: dict[str, dict] = {}
    swift_mismatches = 0
    total_usd = 0.0

    for r in rows:
        country = r.get("beneficiary_country_code") or "UNK"
        usd = float(r.get("destiny_amount_usd") or 0)
        total_usd += usd
        if r.get("swift_country_mismatch_flag"):
            swift_mismatches += 1
        bucket = by_country.setdefault(country, {"count": 0, "usd": 0.0})
        bucket["count"] += 1
        bucket["usd"] += usd

    top_countries = sorted(
        ({"country": k, **v} for k, v in by_country.items()),
        key=lambda x: x["usd"],
        reverse=True,
    )[:10]

    return {
        "total_transactions": total,
        "total_usd": total_usd,
        "distinct_countries": len(by_country),
        "swift_country_mismatches": swift_mismatches,
        "top_countries": top_countries,
    }


def _summary_amount_ranges(rows: list[dict]) -> dict:
    total_rows = len(rows)
    total_txs = sum(int(r.get("total_transactions") or 0) for r in rows)
    total_usd = sum(float(r.get("total_amount_usd") or 0) for r in rows)
    distinct_countries = len({r.get("beneficiary_country_code") for r in rows})

    # Top 5 country+range combos by transaction count
    top_combos = sorted(rows, key=lambda r: int(r.get("total_transactions") or 0), reverse=True)[:5]

    return {
        "total_rows": total_rows,
        "total_transactions": total_txs,
        "total_usd": total_usd,
        "distinct_countries": distinct_countries,
        "top_combos": [
            {
                "country": r.get("beneficiary_country_code"),
                "range": r.get("amount_range_usd"),
                "count": r.get("total_transactions"),
                "usd": r.get("total_amount_usd"),
            }
            for r in top_combos
        ],
    }


def _summary_top_customers(rows: list[dict]) -> dict:
    total_rows = len(rows)
    distinct_customers = len({r.get("customer_id") for r in rows})
    distinct_countries = len({r.get("beneficiary_country_code") for r in rows})
    total_usd = sum(float(r.get("total_amount_usd") or 0) for r in rows)

    # Top 5 customers by total transactions
    top_customers = sorted(rows, key=lambda r: int(r.get("total_transactions") or 0), reverse=True)[:5]

    return {
        "total_rows": total_rows,
        "distinct_customers": distinct_customers,
        "distinct_countries": distinct_countries,
        "total_usd": total_usd,
        "top_customers": [
            {
                "customer_id": r.get("customer_id"),
                "email": r.get("customer_email"),
                "country": r.get("beneficiary_country_code"),
                "range": r.get("amount_range_usd"),
                "count": r.get("total_transactions"),
                "usd": r.get("total_amount_usd"),
            }
            for r in top_customers
        ],
    }


def _summary_generic(rows: list[dict]) -> dict:
    return {"total_rows": len(rows)}


def build_excel(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    if not rows:
        ws["A1"] = "No data found for the selected period."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    columns = list(rows[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_transaction_search_excel(rows: list[dict], n_requested: int) -> bytes:
    """Generate Excel for transaction_search: one sheet, styled header, tipo_envio highlighted."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1B3A6B")
    nac_fill     = PatternFill("solid", fgColor="DCFCE7")   # verde claro
    int_fill     = PatternFill("solid", fgColor="DBEAFE")   # azul claro
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    if not rows:
        ws["A1"] = f"No se encontraron transacciones para los {n_requested} IDs consultados."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    columns = list(rows[0].keys())
    tipo_envio_col = columns.index("tipo_envio") + 1 if "tipo_envio" in columns else None

    # Header row
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col))
            # Colorea tipo_envio según valor
            if tipo_envio_col and col_idx == tipo_envio_col:
                val = row.get(col, "") or ""
                cell.fill = nac_fill if val == "Envío nacional" else int_fill
                cell.alignment = center_align

    # Column widths
    col_widths = {
        "transaction_id": 18, "customer_id": 14, "beneficiary_country_name": 26,
        "tipo_envio": 20, "beneficiary_dni": 16, "beneficiary_dni_type": 18,
        "beneficiary_name": 28, "beneficiary_first_name": 22, "beneficiary_last_name": 22,
        "beneficiary_email": 32, "beneficiary_id": 16, "origin_country": 18,
        "destiny_country": 18, "destiny_amount_usd": 20, "tx_status": 22, "start_date": 22,
    }
    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 20)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_to_s3(content: bytes, key: str, content_type: str) -> str:
    s3.put_object(
        Bucket=S3_BUCKET,
        Key=key,
        Body=content,
        ContentType=content_type,
        ServerSideEncryption="AES256",
    )
    url = s3.generate_presigned_url(
        "get_object",
        Params={"Bucket": S3_BUCKET, "Key": key},
        ExpiresIn=24 * 60 * 60,  # 24h
    )
    return url


def render_email_html(summary: dict, params: dict, s3_url: str) -> str:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template = env.get_template("email_template.html")
    return template.render(
        summary=summary,
        params=params,
        s3_url=s3_url,
        generated_at=dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
    )


# ---------------------------------------------------------------------------
# Delivery
# ---------------------------------------------------------------------------
def send_email(html_body: str, xlsx_bytes: bytes, xlsx_filename: str, subject: str) -> None:
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = SES_FROM
    msg["To"] = ", ".join(SES_TO)

    body = MIMEMultipart("alternative")
    body.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(body)

    att = MIMEApplication(xlsx_bytes)
    att.add_header("Content-Disposition", "attachment", filename=xlsx_filename)
    msg.attach(att)

    ses.send_raw_email(
        Source=SES_FROM,
        Destinations=SES_TO,
        RawMessage={"Data": msg.as_string()},
    )
    logger.info("Email sent to %s", SES_TO)


def post_slack(summary: dict, params: dict, s3_url: str, report_name: str) -> None:
    if not SLACK_SECRET_ARN:
        logger.info("No Slack secret configured, skipping Slack notification")
        return

    secret = secrets.get_secret_value(SecretId=SLACK_SECRET_ARN)
    webhook_url = secret["SecretString"].strip()
    display_name = REPORT_CONFIGS.get(report_name, {}).get("display_name", report_name)
    generated_at = dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    text_lines = [f"*Compliance Report — {display_name}*", f"Generated: {generated_at}", ""]

    if report_name == "high_risk_countries":
        text_lines += [
            f"Period since: `{params['since_date']}`",
            f"• Total transactions: *{summary['total_transactions']:,}*",
            f"• Total USD: *${summary['total_usd']:,.2f}*",
            f"• Distinct countries: *{summary['distinct_countries']}*",
            f"• SWIFT/country mismatches: *{summary['swift_country_mismatches']}* :warning:",
            "",
            "*Top 5 countries by USD:*",
        ]
        for c in summary["top_countries"][:5]:
            text_lines.append(f"  • {c['country']}: {c['count']} tx — ${c['usd']:,.2f}")

    elif report_name == "amount_ranges_by_country":
        text_lines += [
            f"• Combinations country × range: *{summary['total_rows']:,}*",
            f"• Total transactions: *{summary['total_transactions']:,}*",
            f"• Total USD: *${summary['total_usd']:,.2f}*",
            f"• Distinct countries: *{summary['distinct_countries']}*",
            "",
            "*Top 5 combos by transaction count:*",
        ]
        for c in summary.get("top_combos", []):
            text_lines.append(f"  • {c['country']} / {c['range']}: {c['count']} tx — ${float(c['usd'] or 0):,.2f}")

    elif report_name == "top_customers_by_range_country":
        text_lines += [
            f"• Total rows: *{summary['total_rows']:,}*",
            f"• Distinct customers: *{summary['distinct_customers']:,}*",
            f"• Distinct countries: *{summary['distinct_countries']}*",
            f"• Total USD: *${summary['total_usd']:,.2f}*",
            "",
            "*Top 5 customers by transaction count:*",
        ]
        for c in summary.get("top_customers", []):
            text_lines.append(
                f"  • {c['customer_id']} ({c['country']} / {c['range']}): {c['count']} tx"
            )

    else:
        text_lines.append(f"• Total rows: *{summary.get('total_rows', '?'):,}*")

    text_lines += ["", f"<{s3_url}|Download full report (expires in 24h)>"]

    payload = json.dumps({"text": "\n".join(text_lines)}).encode("utf-8")
    req = urllib.request.Request(
        webhook_url,
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        if resp.status != 200:
            raise RuntimeError(f"Slack webhook returned {resp.status}")
    logger.info("Slack notification posted")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def handler(event, context):  # noqa: ARG001
    logger.info("Event: %s", json.dumps(event, default=str))

    # report_name: from event payload (EventBridge / API / manual invoke) → fallback to env var
    report_name = event.get("report_name") or REPORT_NAME
    # run_id is injected by api_handler when a user triggers via the frontend
    run_id: str | None = event.get("run_id")

    # ── Módulo especial: escucha de respuestas de documentos por correo ────
    # No usa Redshift ni el ciclo normal de runs — es un job de polling IMAP
    # disparado por EventBridge cada ~10 min. Retorna directo, sin pasar por
    # el resto del flujo de reportes (encendido de clúster, Excel, etc.).
    if report_name == "poll_document_replies":
        return _poll_document_replies()

    # ── Módulo especial: Análisis AML Individual ──────────────────────────
    if report_name == "individual_aml_analysis":
        customer_ids = event.get("customer_ids", [])
        if not customer_ids:
            err = "customer_ids is required for individual_aml_analysis"
            _update_run(run_id, status="ERROR", error_message=err,
                        completed_at=dt.datetime.utcnow().isoformat())
            raise ValueError(err)
        # Período configurable — aplica SOLO a la fuente nueva (cash call pay-in).
        # Las 2 queries originales (remesas + CCA wallet_deposit) NO se tocan.
        days = event.get("days")
        try:
            days = int(days)
            if days not in (5, 15, 30, 60, 90):
                days = None
        except (TypeError, ValueError):
            days = None
        days_filter_cashcall = (
            f"AND cc.creation_date >= DATEADD(day, -{days}, CURRENT_DATE)" if days else ""
        )
        days_filter_qr = (
            f"AND t.paid_date_millis >= (EXTRACT(EPOCH FROM DATEADD(day, -{days}, CURRENT_DATE))::BIGINT * 1000)"
            if days else ""
        )

        # Tipo de entidad: 'b2c' (default, customer_v2) o 'b2b' (company.company).
        # Son archivos .sql DUPLICADOS por tipo — nunca se mezclan ni se
        # parametriza una sola query para ambos casos.
        entity_type = "b2b" if str(event.get("entity_type", "b2c")).lower() == "b2b" else "b2c"
        suffix = "_b2b" if entity_type == "b2b" else ""

        try:
            _update_run(run_id, status="RESUMING")
            ensure_cluster_available()

            # Build customer_ids SQL list
            ids_sql = ", ".join(str(int(i)) for i in customer_ids)

            def _render(filename: str, extra: str = "") -> str:
                sql = (QUERIES_DIR / filename).read_text(encoding="utf-8")
                sql = sql.strip().rstrip(";").replace("{customer_ids}", ids_sql)
                if "{days_filter}" in sql:
                    sql = sql.replace("{days_filter}", extra)
                return sql

            # Motor de scoring/flags (rows_out + rows_in): remesas + CCA wallet-deposit
            # (como siempre) + Cash Call pay-out + QR Payment — todo lo que NO es
            # "CCA Pay-In" entra al score, por acuerdo explícito. Verificado sin
            # duplicación: ni Cash Call DR ni QR_PAYMENT se solapan con remesas/wallet.
            rows_out = execute_query(_render(f"individual_aml_out{suffix}.sql"))
            rows_out += execute_query(_render(f"individual_cashcall_out{suffix}.sql", days_filter_cashcall))
            rows_out += execute_query(_render(f"individual_qrpayment{suffix}.sql", days_filter_qr))
            rows_in = execute_query(_render(f"individual_aml_in{suffix}.sql"))

            # Única fuente que queda SEPARADA (no entra al score): CCA Cash Call
            # Pay-In (treasury.cash_call, type='CR'). Va en su propia hoja.
            rows_cashcall_in = execute_query(
                _render(f"individual_cashcall_in{suffix}.sql", days_filter_cashcall)
            )

            xlsx_bytes = build_aml_excel(rows_out, rows_in, customer_ids, rows_cashcall_in=rows_cashcall_in)

            run_ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            key = f"individual_aml_analysis/{run_ts}_customers-{len(customer_ids)}.xlsx"

            s3_url = upload_to_s3(
                xlsx_bytes, key,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            total_rows = len(rows_out) + len(rows_in) + len(rows_cashcall_in)
            _update_run(
                run_id,
                status="DONE",
                completed_at=dt.datetime.utcnow().isoformat(),
                s3_key=key,
                row_count=total_rows,
                result_preview=json.dumps([], default=str),
            )
            return {"status": "ok", "report_name": report_name, "rows": total_rows, "s3_key": key}
        except Exception as e:
            logger.exception("Individual AML analysis failed: %s", e)
            _update_run(run_id, status="ERROR", completed_at=dt.datetime.utcnow().isoformat(), error_message=str(e))
            raise
        finally:
            keep_session = event.get("keep_session", False)
            if AUTO_PAUSE and not keep_session:
                pause_cluster()
    # ── Módulo especial: Análisis de Wallet por partner_account_id ───────────
    if report_name == "wallet_search":
        partner_account_ids = event.get("partner_account_ids", [])
        wallet_entity_type = event.get("entity_type", "b2c")
        if not partner_account_ids:
            err = "partner_account_ids is required for wallet_search"
            _update_run(run_id, status="ERROR", error_message=err,
                        completed_at=dt.datetime.utcnow().isoformat())
            raise ValueError(err)
        try:
            _update_run(run_id, status="RESUMING")
            ensure_cluster_available()
            _update_run(run_id, status="RUNNING")

            ids_sql = ", ".join("'" + str(pid).replace("'", "''") + "'" for pid in partner_account_ids)

            if wallet_entity_type == "b2b":
                sql = f"""
SELECT
    pg.partner_account_id,
    pg.customer_id,
    pgg.account_group_id,

    co.name AS company_name,
    co.username,
    co.identification_type,
    co.identification_number,

    co.kyc_stage_1,
    co.compliance_status,
    co.risk_level

FROM "db_prod"."product_gateway"."account" AS pg

INNER JOIN "db_prod"."product_gateway"."account_group" AS pgg
    ON pg.account_group_id = pgg.account_group_id

INNER JOIN "db_prod"."company"."company" AS co
    ON pg.customer_id = co.company_id

WHERE pg.partner_account_id IN ({ids_sql})
  AND UPPER(co.kyc_stage_1) = 'APPROVED'
  AND UPPER(co.identification_type) = 'RUT'

ORDER BY
    co.name
""".strip()
            else:
                sql = f"""
WITH latest_compliance AS (
    SELECT
        customer_id,
        compliance_status,
        ROW_NUMBER() OVER (
            PARTITION BY customer_id
            ORDER BY status_created_at DESC
        ) AS rn
    FROM "db_prod"."customer"."compliance"
),

latest_customer_kyc AS (
    SELECT
        customer_id,
        kyc_status,
        created_at,
        ROW_NUMBER() OVER (
            PARTITION BY customer_id
            ORDER BY created_at DESC
        ) AS rn
    FROM "db_prod"."customer"."customer_kyc"
),

latest_kyc_document AS (
    SELECT
        customer_id,
        document_number,
        document_type,
        country_code,
        ROW_NUMBER() OVER (
            PARTITION BY customer_id
            ORDER BY COALESCE(updated_at, created_at) DESC
        ) AS rn
    FROM "db_prod"."customer"."kyc_document"
)

SELECT
    pg.partner_account_id,
    pg.gmoney_account_id,
    pg.customer_id,

    pgg.account_type,
    pgg.currency_code,
    pgg.country_code AS account_country_code,

    c.name,
    c.last_name,
    c.email,

    kd.document_number,
    kd.document_type,
    kd.country_code,

    kyc.kyc_status,
    lc.compliance_status

FROM "db_prod"."product_gateway"."account" AS pg

INNER JOIN "db_prod"."product_gateway"."account_group" AS pgg
    ON pg.account_group_id = pgg.account_group_id

INNER JOIN "db_prod"."customer"."customer_v2" AS c
    ON pg.customer_id = c.customer_id

INNER JOIN latest_customer_kyc AS kyc
    ON pg.customer_id = kyc.customer_id
   AND kyc.rn = 1

INNER JOIN latest_kyc_document AS kd
    ON pg.customer_id = kd.customer_id
   AND kd.rn = 1

LEFT JOIN latest_compliance AS lc
    ON pg.customer_id = lc.customer_id
   AND lc.rn = 1

WHERE pg.partner_account_id IN ({ids_sql})
  AND UPPER(kyc.kyc_status) = 'APPROVED'
  AND UPPER(kd.document_type) = 'RUT'

ORDER BY
    c.name,
    c.last_name
""".strip()

            rows = execute_query(sql)
            xlsx_bytes = build_excel(rows)

            run_ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            key = f"wallet_search/{run_ts}_{wallet_entity_type}_ids-{len(partner_account_ids)}.xlsx"
            upload_to_s3(xlsx_bytes, key,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            _update_run(
                run_id,
                status="DONE",
                completed_at=dt.datetime.utcnow().isoformat(),
                s3_key=key,
                row_count=len(rows),
                result_preview=json.dumps(rows[:10], default=str),
            )
            return {"status": "ok", "report_name": report_name, "rows": len(rows), "s3_key": key}
        except Exception as e:
            logger.exception("Wallet search failed: %s", e)
            _update_run(run_id, status="ERROR", completed_at=dt.datetime.utcnow().isoformat(),
                        error_message=str(e))
            raise
        finally:
            keep_session = event.get("keep_session", False)
            if AUTO_PAUSE and not keep_session:
                pause_cluster()
    # ── Módulo especial: Búsqueda de Transacciones por ID (Remesas) ─────────
    if report_name == "transaction_search":
        transaction_ids = event.get("transaction_ids", [])
        if not transaction_ids:
            err = "transaction_ids is required for transaction_search"
            _update_run(run_id, status="ERROR", error_message=err,
                        completed_at=dt.datetime.utcnow().isoformat())
            raise ValueError(err)
        try:
            _update_run(run_id, status="RESUMING")
            ensure_cluster_available()
            _update_run(run_id, status="RUNNING")

            ids_sql = ", ".join(str(int(i)) for i in transaction_ids)
            sql = f"""
SELECT
    transaction_id,
    customer_id,
    beneficiary_country_name,
    CASE
        WHEN beneficiary_country_name = 'Chile' THEN 'Envío nacional'
        ELSE 'Envío internacional'
    END AS tipo_envio,
    beneficiary_dni,
    beneficiary_dni_type,
    beneficiary_name,
    beneficiary_first_name,
    beneficiary_last_name,
    beneficiary_email,
    beneficiary_id,
    origin_country,
    destiny_country,
    destiny_amount_usd,
    tx_status,
    start_date
FROM "db_prod"."transaction"."transaction"
WHERE transaction_id IN ({ids_sql})
ORDER BY start_date DESC
""".strip()

            rows = execute_query(sql)
            xlsx_bytes = _build_transaction_search_excel(rows, len(transaction_ids))

            run_ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            key = f"transaction_search/{run_ts}_txs-{len(transaction_ids)}.xlsx"
            upload_to_s3(xlsx_bytes, key,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            _update_run(
                run_id,
                status="DONE",
                completed_at=dt.datetime.utcnow().isoformat(),
                s3_key=key,
                row_count=len(rows),
                result_preview=json.dumps([], default=str),
            )
            return {"status": "ok", "report_name": report_name, "rows": len(rows), "s3_key": key}
        except Exception as e:
            logger.exception("Transaction search failed: %s", e)
            _update_run(run_id, status="ERROR", completed_at=dt.datetime.utcnow().isoformat(),
                        error_message=str(e))
            raise
        finally:
            keep_session = event.get("keep_session", False)
            if AUTO_PAUSE and not keep_session:
                pause_cluster()
    # ── Fin módulo especial ───────────────────────────────────────────────

    # Validate: built-in OR exists in DynamoDB catalog
    is_builtin = report_name in REPORT_CONFIGS
    if not is_builtin and not _load_custom_sql(report_name):
        err = f"Unknown report '{report_name}'. Valid built-ins: {list(REPORT_CONFIGS)}"
        _update_run(run_id, status="ERROR", error_message=err,
                    completed_at=dt.datetime.utcnow().isoformat())
        raise ValueError(err)

    config = REPORT_CONFIGS.get(report_name, {})
    display_name = config.get("display_name", report_name)
    logger.info("Running report: %s (run_id=%s)", display_name, run_id)

    # Resolve optional params (only used by high_risk_countries today)
    today = dt.date.today()
    default_since = today.replace(day=1).isoformat()
    since_date = event.get("since_date") or default_since
    only_successful = bool(event.get("only_successful", False))

    country_codes: list[str] = []
    if config.get("needs_country_filter"):
        countries = load_country_codes()
        country_codes = [c["code"] for c in countries]

    params = {
        "report_name": report_name,
        "since_date": since_date if config.get("needs_since_date") else "last_7_days",
        "only_successful": only_successful,
        "country_count": len(country_codes),
    }

    try:
        # Signal the frontend that we're waking the cluster
        _update_run(run_id, status="RESUMING")
        ensure_cluster_available()

        sql, api_params = render_query(report_name, since_date, only_successful, country_codes)
        rows = execute_query(sql, api_params)
        rows = enrich_rows_with_priority(rows)

        summary = build_summary(rows, report_name)
        xlsx_bytes = build_excel(rows)

        run_ts = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        # Key includes since_date for high_risk_countries; just timestamp for others
        if config.get("needs_since_date"):
            key = f"{report_name}/{run_ts}_since-{since_date}.xlsx"
        else:
            key = f"{report_name}/{run_ts}.xlsx"

        s3_url = upload_to_s3(
            xlsx_bytes, key,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        total_rows = summary.get("total_transactions") or summary.get("total_rows", 0)
        subject = f"[Compliance] {display_name} — {total_rows} rows"

        # Email + Slack are best-effort — never fail the run
        try:
            html = render_email_html(summary, params, s3_url)
            send_email(html, xlsx_bytes, Path(key).name, subject)
        except Exception as e:  # noqa: BLE001
            logger.warning("Email delivery failed (non-blocking): %s", e)

        try:
            post_slack(summary, params, s3_url, report_name)
        except Exception as e:  # noqa: BLE001
            logger.warning("Slack notification failed (non-blocking): %s", e)

        # Update DynamoDB run record to DONE (best-effort)
        result_preview = rows[:10]  # first 10 rows for in-browser preview
        _update_run(
            run_id,
            status="DONE",
            completed_at=dt.datetime.utcnow().isoformat(),
            s3_key=key,
            row_count=total_rows,
            result_preview=json.dumps(result_preview, default=str),
            # Agregados sobre TODAS las filas + las más extremas → la IA analiza el dataset completo
            ai_summary=json.dumps(_build_ai_summary(rows), default=str),
        )

        # Phase 10 — apply auto-case rules (non-blocking)
        try:
            _apply_auto_case_rules(report_name, rows, run_id)
        except Exception:
            pass

        # Trigger de automatización end-to-end: si el interruptor maestro de
        # priorización está prendido, dispara solicitud de documentos + caso
        # para las filas P1 de este reporte (non-blocking, nunca rompe la
        # ejecución del reporte). Apagado por defecto — ver Admin > Priorización.
        try:
            _trigger_auto_document_requests(report_name, rows)
        except Exception:
            pass

        return {
            "status": "ok",
            "report_name": report_name,
            "rows": total_rows,
            "s3_key": key,
            "params": params,
        }

    except Exception as e:  # noqa: BLE001
        logger.exception("Report run failed: %s", e)
        _update_run(
            run_id,
            status="ERROR",
            completed_at=dt.datetime.utcnow().isoformat(),
            error_message=str(e),
        )
        raise

    finally:
        keep_session = event.get("keep_session", False)
        if AUTO_PAUSE and not keep_session:
            pause_cluster()
